# ui_reports.py
# Contains: ReportWindow, SessionHistoryWindow, AnalyticsWindow

import os
import tempfile
import smtplib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidgetItem, QFrame, QDialog, QLineEdit, QFileDialog,
    QMessageBox, QScrollArea, QRadioButton, QButtonGroup, QComboBox,
    QApplication,
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon

import config
from config import (
    BG, WHITE, PRIMARY, PRIMARY_LIGHT, BORDER,
    TEXT_DARK, TEXT_MED, TEXT_GRAY,
    DB_CONNECTED, db,
    SMTP_SERVER, SMTP_PORT, SMTP_LOGIN, SMTP_PASSWORD, SMTP_FROM,
)
from ui_theme import (
    make_stat_card, make_badge, make_table,
)


class ReportWindow(QWidget):
    def __init__(self, main_win, session_id=None, attendance=None, enrolled_students=None, course_data=None, auto_email_ok=None):
        super().__init__()
        self.main_win          = main_win
        self.session_id        = session_id
        self.attendance        = attendance or {}
        self.enrolled_students = enrolled_students or []
        self.course_data       = course_data or {}
        self.auto_email_ok     = auto_email_ok
        self.setWindowTitle("AIAS — Session Report")
        import os
        from PyQt5.QtGui import QIcon
        _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aias_icon.ico")
        if os.path.exists(_icon_path):
            self.setWindowIcon(QIcon(_icon_path))
        self.setMinimumSize(1100, 700)
        self.resize(1100, 700)
        self.setStyleSheet(f"background:{BG};")
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(48, 36, 48, 28)
        root.setSpacing(18)

        rows = []
        for stu in self.enrolled_students:
            sid  = str(stu.get("StudentID", ""))
            name = stu.get("FullName", "")
            att = self.attendance.get(sid, {})
            rows.append((
                sid, name,
                att.get("first_seen", "—"),
                att.get("last_seen",  "—"),
                att.get("status",     "Absent"),
                att.get("note",       ""),
            ))

        course_title = self.course_data.get("CourseTitle", "")
        sched_raw    = self.course_data.get("Schedule", "")
        if isinstance(sched_raw, dict):
            sched = f"{sched_raw.get('Day','')} {sched_raw.get('StartTime','')}–{sched_raw.get('EndTime','')}"
        else:
            sched = str(sched_raw)
        n_students = len(rows)

        title = QLabel("Session ended")
        title.setStyleSheet(f"font-size:24px; font-weight:800; color:{TEXT_DARK};")
        sub = QLabel(f"{course_title} · {sched} · {n_students} students")
        sub.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY};")
        root.addWidget(title)
        root.addWidget(sub)

        n_present     = sum(1 for _, _, _, _, s, _ in rows if s == "Present")
        n_late        = sum(1 for _, _, _, _, s, _ in rows if s == "Late")
        n_early_leave = sum(1 for _, _, _, _, s, _ in rows if s == "Early Leave")
        n_absent      = sum(1 for _, _, _, _, s, _ in rows if s == "Absent")
        stats_row = QHBoxLayout()
        stats_row.setSpacing(16)
        stats_row.addWidget(make_stat_card("Present",     n_present,     "#22C55E"))
        stats_row.addWidget(make_stat_card("Late",        n_late,        "#F59E0B"))
        stats_row.addWidget(make_stat_card("Early Leave", n_early_leave, "#3B82F6"))
        stats_row.addWidget(make_stat_card("Absent",      n_absent,      "#EF4444"))
        stats_row.addStretch()
        root.addLayout(stats_row)

        # Bug 4 / Rec 6: inline auto-email notice
        if self.auto_email_ok is not None:
            if self.auto_email_ok:
                _n_txt = "✅  Auto-email sent successfully."
                _n_bg, _n_bd, _n_fg = "#F0FDF4", "#86EFAC", "#166534"
            else:
                _n_txt = "⚠  Auto-email failed — use Send Email below to retry."
                _n_bg, _n_bd, _n_fg = "#FFFBEB", "#FCD34D", "#92400E"
            _notice = QLabel(_n_txt)
            _notice.setStyleSheet(
                f"background:{_n_bg}; border:1px solid {_n_bd}; border-radius:8px;"
                f"padding:10px 16px; color:{_n_fg}; font-size:12px; font-weight:600;"
            )
            root.addWidget(_notice)

        tbl = make_table(
            ["Student ID", "Full Name", "First seen", "Last seen", "Status", "Note"],
            rows,
            col_widths={0: 140, 2: 100, 3: 100, 4: 130},
            stretch_col=5,
        )
        for r, (sid, name, first, last, status, note) in enumerate(rows):
            tbl.setItem(r, 0, QTableWidgetItem(sid))
            tbl.setItem(r, 1, QTableWidgetItem(name))
            tbl.setItem(r, 2, QTableWidgetItem(first))
            tbl.setItem(r, 3, QTableWidgetItem(last))
            tbl.setCellWidget(r, 4, make_badge(status))
            tbl.setItem(r, 5, QTableWidgetItem(note))
            tbl.setRowHeight(r, 44)
        root.addWidget(tbl)

        actions_row = QHBoxLayout()
        actions_row.setSpacing(20)
        for card_title, desc, btn_txt in [
            ("Export report", "Download as PDF or Excel", "Export"),
            ("Send by email", "Send report to your university email", "Send email"),
        ]:
            card = QFrame()
            card.setStyleSheet(
                f"QFrame {{ background:{WHITE}; border:1px solid {BORDER}; border-radius:10px; }}"
            )
            card.setFixedHeight(110)
            c_lay = QVBoxLayout(card)
            c_lay.setContentsMargins(22, 16, 22, 16)
            c_lay.setSpacing(4)
            ct = QLabel(card_title)
            ct.setStyleSheet(
                f"font-size:14px; font-weight:700; color:{TEXT_DARK}; border:none; background:transparent;"
            )
            cd = QLabel(desc)
            cd.setStyleSheet(
                f"font-size:12px; color:{TEXT_GRAY}; border:none; background:transparent;"
            )
            cb = QPushButton(btn_txt)
            cb.setFixedWidth(110)
            cb.setFixedHeight(34)
            cb.setStyleSheet(
                f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
                f"font-size:12px; font-weight:600;"
            )
            cb.setCursor(Qt.PointingHandCursor)
            if btn_txt == "Export":
                cb.clicked.connect(self._export_report)
            elif btn_txt == "Send email":
                cb.clicked.connect(self._send_email)
            c_lay.addWidget(ct)
            c_lay.addWidget(cd)
            c_lay.addSpacing(4)
            c_lay.addWidget(cb, alignment=Qt.AlignLeft)
            actions_row.addWidget(card)

        pdf_btn = QPushButton("📄  Export PDF")
        pdf_btn.setStyleSheet(
            f"QPushButton {{ background:#DC2626; color:white; border-radius:6px; "
            f"padding:8px 16px; font-size:12px; font-weight:600; }}"
            f"QPushButton:hover {{ background:#B91C1C; color:white; }}"
        )
        pdf_btn.setCursor(Qt.PointingHandCursor)
        pdf_btn.clicked.connect(self._export_pdf)
        actions_row.addWidget(pdf_btn)
        actions_row.addStretch()
        root.addLayout(actions_row)

        back_row = QHBoxLayout()
        back_row.addStretch()
        back_btn = QPushButton("← Back to Home")
        back_btn.setFixedHeight(40)
        back_btn.setStyleSheet(
            f"background:{WHITE}; color:{PRIMARY}; border:1px solid {PRIMARY}; border-radius:6px;"
            "padding:8px 16px; font-size:13px; font-weight:600;"
        )
        back_btn.setCursor(Qt.PointingHandCursor)
        back_btn.clicked.connect(self._back_to_home)
        back_row.addWidget(back_btn)
        root.addLayout(back_row)

    def _export_report(self):
        course_title = self.course_data.get("CourseTitle", "Attendance") if self.course_data else "Attendance"
        default_name = f"Attendance_{course_title}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        path, selected_filter = QFileDialog.getSaveFileName(
            self, "Save Report", default_name,
            "Excel Files (*.xlsx);;PDF Files (*.pdf)"
        )
        if not path:
            return

        if selected_filter == "PDF Files (*.pdf)" or path.lower().endswith(".pdf"):
            self._export_pdf_qt(path, course_title)
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Report"

            green_fill = PatternFill("solid", start_color="1B5E35")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            center = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("A1:F1")
            ws["A1"] = f"Attendance Report — {course_title}"
            ws["A1"].font = Font(bold=True, size=14, color="1A1A2E")
            ws["A1"].alignment = center

            if self.course_data:
                sched = self.course_data.get("Schedule", "")
                if isinstance(sched, dict):
                    sched_str = f"{sched.get('Day','')} {sched.get('StartTime','')}–{sched.get('EndTime','')}"
                else:
                    sched_str = str(sched)
                ws.merge_cells("A2:F2")
                ws["A2"] = f"Date: {datetime.now().strftime('%Y-%m-%d')}  |  Schedule: {sched_str}"
                ws["A2"].alignment = center

            ws.append([])

            headers = ["Student ID", "Full Name", "First Seen", "Last Seen", "Status", "Note"]
            ws.append(headers)
            header_row = ws.max_row
            for col in range(1, 7):
                cell = ws.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = green_fill
                cell.alignment = center

            status_colors = {
                "Present":     "EAF3DE",
                "Late":        "FAEEDA",
                "Absent":      "FCEBEB",
                "Early Leave": "E6F1FB",
            }

            for stu in self.enrolled_students:
                sid = str(stu.get("StudentID", ""))
                name = stu.get("FullName", "")
                rec = self.attendance.get(sid, {})
                status = rec.get("status", "Absent")
                first_seen = rec.get("first_seen", "—")
                last_seen = rec.get("last_seen", "—")
                note = rec.get("note", "")

                ws.append([sid, name, first_seen, last_seen, status, note])
                row = ws.max_row
                status_fill = PatternFill("solid", start_color=status_colors.get(status, "FFFFFF"))
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = center
                    if col == 5:
                        cell.fill = status_fill

            ws.column_dimensions["A"].width = 16
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 14
            ws.column_dimensions["F"].width = 28

            wb.save(path)
            QMessageBox.information(self, "Export successful", f"Report saved to:\n{path}")

        except Exception as e:
            QMessageBox.critical(self, "Export failed", f"Error: {e}")

    def _export_pdf(self):
        from PyQt5.QtWidgets import QFileDialog, QMessageBox
        import datetime, os

        path, _ = QFileDialog.getSaveFileName(
            self, "Save PDF Report",
            f"AIAS_Report_{datetime.date.today()}.pdf",
            "PDF Files (*.pdf)"
        )
        if not path:
            return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                            Table, TableStyle, HRFlowable)
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

            doc = SimpleDocTemplate(
                path,
                pagesize=A4,
                rightMargin=2*cm,
                leftMargin=2*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )

            story = []
            styles = getSampleStyleSheet()

            # --- HEADER ---
            header_style = ParagraphStyle(
                'Header',
                fontSize=18,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#1B5E35'),
                alignment=TA_CENTER,
                spaceAfter=4
            )
            sub_style = ParagraphStyle(
                'Sub',
                fontSize=11,
                fontName='Helvetica',
                textColor=colors.HexColor('#6B7280'),
                alignment=TA_CENTER,
                spaceAfter=2
            )
            label_style = ParagraphStyle(
                'Label',
                fontSize=11,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#1A1A2E'),
                spaceAfter=4
            )
            normal_style = ParagraphStyle(
                'Normal2',
                fontSize=10,
                fontName='Helvetica',
                textColor=colors.HexColor('#374151'),
                spaceAfter=2
            )

            story.append(Paragraph("Qassim University", header_style))
            story.append(Paragraph("College of Computer Science", sub_style))
            story.append(Paragraph("AI Attendance System — Session Report", sub_style))
            story.append(Spacer(1, 0.3*cm))
            story.append(HRFlowable(width="100%", thickness=2,
                                    color=colors.HexColor('#1B5E35')))
            story.append(Spacer(1, 0.4*cm))

            # --- COURSE INFO ---
            import datetime as dt
            try:
                course_title    = self.course_data.get("CourseTitle", "N/A") if self.course_data else "N/A"
                section_id      = self.course_data.get("SectionID", "N/A") if self.course_data else "N/A"
                instructor_name = (self.main_win.instructor_data.get("fullname", "N/A")
                                   if self.main_win else "N/A")
                session_date    = str(dt.date.today())
                sched_raw       = self.course_data.get("Schedule", {}) if self.course_data else {}
                if isinstance(sched_raw, dict):
                    start_time = sched_raw.get("StartTime", "")
                    end_time   = sched_raw.get("EndTime", "")
                else:
                    start_time = end_time = str(sched_raw)
                records = []
                for stu in self.enrolled_students:
                    sid  = str(stu.get("StudentID", ""))
                    name = stu.get("FullName", "")
                    att  = self.attendance.get(sid, {})
                    records.append({
                        "StudentID":    sid,
                        "FullName":     name,
                        "Status":       att.get("status", "Absent"),
                        "RecognizedAt": att.get("first_seen", ""),
                    })
            except Exception:
                course_title = section_id = instructor_name = "N/A"
                session_date = str(dt.date.today())
                start_time = end_time = ""
                records = []

            info_data = [
                ["Course:", course_title,   "Section:", section_id],
                ["Instructor:", instructor_name, "Date:", session_date],
                ["Start Time:", start_time,  "End Time:", end_time],
            ]
            info_table = Table(info_data, colWidths=[3*cm, 6*cm, 3*cm, 5*cm])
            info_table.setStyle(TableStyle([
                ('FONTNAME',  (0,0), (-1,-1), 'Helvetica'),
                ('FONTNAME',  (0,0), (0,-1),  'Helvetica-Bold'),
                ('FONTNAME',  (2,0), (2,-1),  'Helvetica-Bold'),
                ('FONTSIZE',  (0,0), (-1,-1), 10),
                ('TEXTCOLOR', (0,0), (-1,-1), colors.HexColor('#1A1A2E')),
                ('ROWBACKGROUNDS', (0,0), (-1,-1),
                 [colors.HexColor('#F8F9FA'), colors.white]),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('TOPPADDING',    (0,0), (-1,-1), 6),
            ]))
            story.append(info_table)
            story.append(Spacer(1, 0.5*cm))

            # --- SUMMARY STATS ---
            present = sum(1 for r in records if r.get("Status") in ("Present", "Late"))
            absent  = sum(1 for r in records if r.get("Status") == "Absent")
            late    = sum(1 for r in records if r.get("Status") == "Late")
            total   = len(records)
            rate    = f"{present/total*100:.1f}%" if total > 0 else "0%"

            story.append(Paragraph("Session Summary", label_style))
            summary_data = [
                ["Total Students", "Present", "Late", "Absent", "Attendance Rate"],
                [str(total), str(present), str(late), str(absent), rate]
            ]
            summary_table = Table(summary_data, colWidths=[3.4*cm]*5)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND',   (0,0), (-1,0), colors.HexColor('#1B5E35')),
                ('TEXTCOLOR',    (0,0), (-1,0), colors.white),
                ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTNAME',     (0,1), (-1,1), 'Helvetica-Bold'),
                ('FONTSIZE',     (0,0), (-1,-1), 11),
                ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
                ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
                ('ROWHEIGHT',    (0,0), (-1,-1), 24),
                ('BACKGROUND',   (0,1), (-1,1), colors.HexColor('#F0FDF4')),
                ('TEXTCOLOR',    (0,1), (-1,1), colors.HexColor('#1B5E35')),
                ('GRID',         (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
                ('ROUNDEDCORNERS', [4]),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.5*cm))

            # --- ATTENDANCE TABLE ---
            story.append(Paragraph("Attendance Details", label_style))

            table_data = [["#", "Student ID", "Full Name", "Status", "Time"]]
            for i, rec in enumerate(records, 1):
                status = rec.get("Status", "")
                table_data.append([
                    str(i),
                    rec.get("StudentID", ""),
                    rec.get("FullName", ""),
                    status,
                    rec.get("RecognizedAt", "")
                ])

            att_table = Table(table_data,
                              colWidths=[1*cm, 3.5*cm, 6*cm, 2.5*cm, 4*cm])
            att_table.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#1B5E35')),
                ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
                ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',      (0,0), (-1,-1), 9),
                ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('ROWHEIGHT',     (0,0), (-1,-1), 20),
                ('ROWBACKGROUNDS',(0,1), (-1,-1),
                 [colors.white, colors.HexColor('#F9FAFB')]),
                ('GRID',          (0,0), (-1,-1), 0.5,
                 colors.HexColor('#E5E7EB')),
                ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
            ]))

            # Color status column
            for i, rec in enumerate(records, 1):
                status = rec.get("Status", "")
                color = (
                    colors.HexColor('#16A34A') if status == "Present" else
                    colors.HexColor('#F59E0B') if status == "Late" else
                    colors.HexColor('#DC2626')
                )
                att_table.setStyle(TableStyle([
                    ('TEXTCOLOR', (3, i), (3, i), color),
                    ('FONTNAME',  (3, i), (3, i), 'Helvetica-Bold'),
                ]))

            story.append(att_table)
            story.append(Spacer(1, 1*cm))

            # --- FOOTER ---
            footer_style = ParagraphStyle(
                'Footer',
                fontSize=8,
                fontName='Helvetica',
                textColor=colors.HexColor('#9CA3AF'),
                alignment=TA_CENTER
            )
            story.append(HRFlowable(width="100%", thickness=0.5,
                                    color=colors.HexColor('#E5E7EB')))
            story.append(Spacer(1, 0.2*cm))
            story.append(Paragraph(
                f"Generated by AIAS — AI Attendance System  ·  "
                f"{dt.datetime.now().strftime('%Y-%m-%d %H:%M')}  ·  "
                f"Qassim University",
                footer_style
            ))

            doc.build(story)
            QMessageBox.information(self, "Success",
                                    f"PDF report saved to:\n{path}")

        except ImportError:
            QMessageBox.warning(self, "Missing Library",
                                "reportlab is required.\nRun: pip install reportlab")
        except Exception as e:
            QMessageBox.warning(self, "Export Failed", f"Error:\n{e}")

    def _export_pdf_qt(self, path, course_title):
        try:
            from PyQt5.QtPrintSupport import QPrinter
            from PyQt5.QtGui import QTextDocument

            sched_raw = self.course_data.get("Schedule", "") if self.course_data else ""
            if isinstance(sched_raw, dict):
                sched_str = f"{sched_raw.get('Day','')} {sched_raw.get('StartTime','')}–{sched_raw.get('EndTime','')}"
            else:
                sched_str = str(sched_raw)

            status_colors_html = {
                "Present":     "#EAF3DE", "Late":        "#FAEEDA",
                "Absent":      "#FCEBEB", "Early Leave":  "#E6F1FB",
            }

            rows_html = ""
            for stu in self.enrolled_students:
                sid  = str(stu.get("StudentID", ""))
                name = stu.get("FullName", "")
                rec  = self.attendance.get(sid, {})
                status     = rec.get("status", "Absent")
                first_seen = rec.get("first_seen", "—")
                last_seen  = rec.get("last_seen",  "—")
                note       = rec.get("note", "")
                bg = status_colors_html.get(status, "#FFFFFF")
                rows_html += (
                    f"<tr>"
                    f"<td>{sid}</td><td>{name}</td>"
                    f"<td>{first_seen}</td><td>{last_seen}</td>"
                    f"<td style='background:{bg};'>{status}</td>"
                    f"<td>{note}</td>"
                    f"</tr>"
                )

            html = f"""
            <html><body style='font-family:Arial,sans-serif; font-size:12px;'>
            <h2 style='color:#1A1A2E;'>Attendance Report — {course_title}</h2>
            <p style='color:#6B7280;'>Date: {datetime.now().strftime('%Y-%m-%d')} &nbsp;|&nbsp; Schedule: {sched_str}</p>
            <table border='1' cellspacing='0' cellpadding='6' width='100%'
                   style='border-collapse:collapse; border-color:#E5E7EB;'>
              <thead>
                <tr style='background:#1B5E35; color:white;'>
                  <th>Student ID</th><th>Full Name</th>
                  <th>First Seen</th><th>Last Seen</th><th>Status</th><th>Note</th>
                </tr>
              </thead>
              <tbody>{rows_html}</tbody>
            </table>
            </body></html>
            """

            doc = QTextDocument()
            doc.setHtml(html)

            printer = QPrinter()
            printer.setOutputFileName(path)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setPageMargins(15, 15, 15, 15, QPrinter.Millimeter)
            doc.print_(printer)

            QMessageBox.information(self, "Export successful", f"PDF saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "PDF Export failed", f"Error: {e}")

    def _send_email(self):
        to_email = ""
        if DB_CONNECTED:
            try:
                username = self.main_win.instructor_data.get("username", "")
                instr = db["Instructors"].find_one({"Username": username})
                if instr:
                    to_email = instr.get("UniversityEmail", "")
            except Exception as e:
                print(f"[AIAS] Email credentials load error: {e}")

        if to_email:
            self._do_send_email(to_email)
            return

        # First-time setup dialog (only asks for destination email)
        dlg = QDialog(self)
        dlg.setWindowTitle("Email Setup")
        dlg.setFixedWidth(420)
        dlg.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        t = QLabel("Email Setup")
        t.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(t)
        sub = QLabel("Enter the university email address to send reports to.")
        sub.setStyleSheet(f"font-size:12px; color:{TEXT_GRAY};")
        sub.setWordWrap(True)
        lay.addWidget(sub)

        to_lbl = QLabel("Send reports to:")
        to_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        to_field = QLineEdit()
        to_field.setPlaceholderText("name@university.edu")
        to_field.setFixedHeight(38)
        to_field.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:6px; padding:6px 10px; background:{WHITE}; color:{TEXT_DARK};"
        )
        lay.addWidget(to_lbl)
        lay.addWidget(to_field)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.clicked.connect(dlg.reject)
        send_btn = QPushButton("Save & Send")
        send_btn.setStyleSheet(
            "background:#1B5E35; color:white; border-radius:6px; padding:8px 20px; font-weight:700;"
        )
        send_btn.clicked.connect(dlg.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(send_btn)
        lay.addLayout(btn_row)

        if dlg.exec_() != QDialog.Accepted:
            return

        to_email = to_field.text().strip()
        if not to_email:
            QMessageBox.warning(self, "Missing info", "Please enter a destination email address.")
            return

        if DB_CONNECTED:
            try:
                username = self.main_win.instructor_data.get("username", "")
                db["Instructors"].update_one(
                    {"Username": username},
                    {"$set": {"UniversityEmail": to_email}},
                )
            except Exception as e:
                print(f"[AIAS] Failed to save email credentials: {e}")

        self._do_send_email(to_email)

    def _do_send_email(self, to_email):
        course_title = self.course_data.get("CourseTitle", "Attendance") if self.course_data else "Attendance"
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Report"

            green_fill  = PatternFill("solid", start_color="1B5E35")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            center      = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("A1:F1")
            ws["A1"] = f"Attendance Report — {course_title}"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = center

            ws.append([])
            ws.append(["Student ID", "Full Name", "First Seen", "Last Seen", "Status", "Note"])
            for col in range(1, 7):
                cell = ws.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = green_fill
                cell.alignment = center

            status_colors = {
                "Present": "EAF3DE", "Late": "FAEEDA",
                "Absent": "FCEBEB", "Early Leave": "E6F1FB",
            }
            for stu in self.enrolled_students:
                sid  = str(stu.get("StudentID", ""))
                rec  = self.attendance.get(sid, {})
                status     = rec.get("status",     "Absent")
                first_seen = rec.get("first_seen", "—")
                last_seen  = rec.get("last_seen",  "—")
                note       = rec.get("note",       "")
                ws.append([sid, stu.get("FullName", ""), first_seen, last_seen, status, note])
                ws.cell(row=ws.max_row, column=5).fill = PatternFill(
                    "solid", start_color=status_colors.get(status, "FFFFFF")
                )
            for col, w in zip("ABCDEF", [16, 28, 14, 14, 14, 28]):
                ws.column_dimensions[col].width = w
            wb.save(tmp.name)

            n_present    = sum(1 for r in self.attendance.values() if r.get("status") == "Present")
            n_late       = sum(1 for r in self.attendance.values() if r.get("status") == "Late")
            n_early      = sum(1 for r in self.attendance.values() if r.get("status") == "Early Leave")
            n_absent     = (
                sum(1 for r in self.attendance.values() if r.get("status") == "Absent")
                + (len(self.enrolled_students) - len(self.attendance))
            )

            msg = MIMEMultipart()
            msg["From"]    = SMTP_FROM
            msg["To"]      = to_email
            msg["Subject"] = (
                f"Attendance Report — {course_title} — "
                f"{datetime.now().strftime('%Y-%m-%d')}"
            )
            body = (
                f"Dear Instructor,\n\n"
                f"Please find attached the attendance report for {course_title}.\n\n"
                f"Date: {datetime.now().strftime('%Y-%m-%d')}\n"
                f"Total students: {len(self.enrolled_students)}\n"
                f"Present: {n_present}  Late: {n_late}  "
                f"Early Leave: {n_early}  Absent: {n_absent}\n\n"
                f"Best regards,\nAIAS — AI Attendance System"
            )
            msg.attach(MIMEText(body, "plain"))

            with open(tmp.name, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            safe = "".join(c for c in course_title if c.isalnum() or c in "-_ ")
            part.add_header(
                "Content-Disposition", f'attachment; filename="Attendance_{safe}.xlsx"'
            )
            msg.attach(part)

            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()
                server.login(SMTP_LOGIN, SMTP_PASSWORD)
                server.send_message(msg)

            QMessageBox.information(self, "Email sent", f"Report sent successfully to:\n{to_email}")

        except Exception as e:
            QMessageBox.critical(
                self, "Email failed",
                f"Error: {str(e)}\n\nMake sure you are using a valid Brevo SMTP Key.\n"
                "To update your settings, use Email Settings in the sidebar.",
            )
        finally:
            try:
                os.unlink(tmp.name)
            except Exception as _e:
                print(f"[AIAS] error: {_e}")

    def _back_to_home(self):
        self.hide()
        self.main_win.show()


class SessionHistoryWindow(QWidget):
    def __init__(self, main_win):
        super().__init__()
        self.main_win = main_win
        self.setWindowTitle("AIAS — Session History")
        import os
        from PyQt5.QtGui import QIcon
        _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aias_icon.ico")
        if os.path.exists(_icon_path):
            self.setWindowIcon(QIcon(_icon_path))
        self.setMinimumSize(1100, 700)
        self.resize(1100, 700)
        self.setStyleSheet(f"background:{BG};")
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(48, 36, 48, 28)
        root.setSpacing(18)

        title = QLabel("Session History")
        title.setStyleSheet(f"font-size:24px; font-weight:800; color:{TEXT_DARK};")
        sub = QLabel("Past completed sessions for your courses")
        sub.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY};")

        header_row = QHBoxLayout()
        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        title_col.addWidget(title)
        title_col.addWidget(sub)
        header_row.addLayout(title_col)
        header_row.addStretch()

        root.addLayout(header_row)

        sessions = []
        if DB_CONNECTED:
            try:
                instr_username = self.main_win.instructor_data.get("username", "")
                raw = list(
                    db["Sessions"].find(
                        {"InstructorID": instr_username, "Status": "completed"}
                    ).sort("StartTime", -1).limit(100)
                )

                # Bug 7: batch-fetch all courses and all logs in 2 queries instead of N+1
                section_ids = list({s.get("SectionID", "") for s in raw if s.get("SectionID")})
                course_map = {}
                if section_ids:
                    for c in db["Courses"].find({"SectionID": {"$in": section_ids}}):
                        _sec = c.get("SectionID", "")
                        course_map[_sec] = f"{_sec} — {c.get('CourseTitle', '')}"

                sess_ids = [s["_id"] for s in raw]
                log_counts = {}   # session_id -> {status: count}
                sess_logs  = {}   # session_id -> [log docs]
                if sess_ids:
                    for log in db["AttendanceLogs"].find({"SessionID": {"$in": sess_ids}}):
                        _s = log.get("SessionID")
                        if _s not in log_counts:
                            log_counts[_s] = {"Present": 0, "Late": 0, "Early Leave": 0, "Absent": 0}
                            sess_logs[_s]  = []
                        _st = log.get("Status", "Absent")
                        if _st in log_counts[_s]:
                            log_counts[_s][_st] += 1
                        sess_logs[_s].append(log)

                for s in raw:
                    _doc_id    = s.get("_id")
                    section_id = s.get("SectionID", "")
                    start_time = s.get("StartTime")
                    date_str   = start_time.strftime("%Y-%m-%d") if start_time else "—"
                    time_str   = start_time.strftime("%H:%M")    if start_time else "—"
                    course_title = course_map.get(section_id, section_id)
                    counts = log_counts.get(_doc_id, {"Present": 0, "Late": 0, "Early Leave": 0, "Absent": 0})
                    logs   = sess_logs.get(_doc_id, [])
                    sessions.append((
                        date_str, time_str, course_title,
                        counts["Present"], counts["Late"], counts["Early Leave"], counts["Absent"],
                        logs,
                    ))
            except Exception as e:
                print(f"[AIAS] History load error: {e}")

        cols = ["Date", "Time", "Course", "Present", "Late", "Early Leave", "Absent", ""]
        tbl  = make_table(cols, [], col_widths={0: 110, 1: 75, 7: 80}, stretch_col=2)
        tbl.setRowCount(len(sessions))

        for r, (date_s, time_s, course_t, n_p, n_l, n_e, n_a, logs) in enumerate(sessions):
            for col, val in enumerate([date_s, time_s, course_t, str(n_p), str(n_l), str(n_e), str(n_a)]):
                item = QTableWidgetItem(val)
                item.setFlags(Qt.ItemIsEnabled)
                tbl.setItem(r, col, item)
            view_btn = QPushButton("View")
            view_btn.setFixedHeight(30)
            view_btn.setStyleSheet(
                f"background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:5px;"
                "font-size:11px; font-weight:600;"
            )
            view_btn.setCursor(Qt.PointingHandCursor)
            view_btn.clicked.connect(
                lambda _, c=course_t, d=date_s, t=time_s, ls=logs: self._view_detail(c, d, t, ls)
            )
            tbl.setCellWidget(r, 7, view_btn)
            tbl.setRowHeight(r, 44)

        root.addWidget(tbl)

        back_row = QHBoxLayout()
        back_row.addStretch()
        back_btn = QPushButton("← Back")
        back_btn.setFixedHeight(40)
        back_btn.setStyleSheet(
            f"background:{WHITE}; color:{PRIMARY}; border:1px solid {PRIMARY}; border-radius:6px;"
            "padding:8px 16px; font-size:13px; font-weight:600;"
        )
        back_btn.setCursor(Qt.PointingHandCursor)
        back_btn.clicked.connect(self._go_back)
        back_row.addWidget(back_btn)
        root.addLayout(back_row)

    def _view_detail(self, course_title, date_str, time_str, logs):
        session_id = logs[0].get("SessionID") if logs else None

        dlg = QDialog(self)
        dlg.setWindowTitle(f"Session Detail — {course_title}")
        dlg.setMinimumSize(900, 500)
        dlg.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        hdr = QLabel(f"{course_title}  ·  {date_str}  {time_str}")
        hdr.setStyleSheet(f"font-size:15px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(hdr)

        rows_data = []
        for log in logs:
            sid = str(log.get("StudentID", ""))
            fullname = sid
            if DB_CONNECTED:
                try:
                    stu = db["Students"].find_one({"StudentID": sid})
                    if stu:
                        fullname = stu.get("FullName", sid)
                except Exception as e:
                    print(f"[AIAS] Student lookup error for {sid}: {e}")
            first_dt = log.get("FirstSeenAt")
            last_dt  = log.get("LastSeenAt")
            rows_data.append([
                sid, fullname,
                first_dt.strftime("%H:%M") if first_dt else "—",
                last_dt.strftime("%H:%M")  if last_dt  else "—",
                log.get("Status", "—"),
                log.get("Note",   ""),
            ])

        det_tbl = make_table(
            ["Student ID", "Full Name", "First Seen", "Last Seen", "Status", "Note", ""],
            rows_data,
            col_widths={0: 120, 2: 90, 3: 90, 4: 110, 5: 150, 6: 60},
            stretch_col=1,
        )
        det_tbl.setRowCount(len(rows_data))

        def _fill_row(r, row):
            det_tbl.setItem(r, 0, QTableWidgetItem(row[0]))
            det_tbl.setItem(r, 1, QTableWidgetItem(row[1]))
            det_tbl.setItem(r, 2, QTableWidgetItem(row[2]))
            det_tbl.setItem(r, 3, QTableWidgetItem(row[3]))
            det_tbl.setCellWidget(r, 4, make_badge(row[4]))
            det_tbl.setItem(r, 5, QTableWidgetItem(row[5]))
            edit_btn = QPushButton("Edit")
            edit_btn.setFixedHeight(28)
            edit_btn.setStyleSheet(
                "background:#E6F1FB; color:#0C447C; border-radius:5px;"
                "font-size:11px; font-weight:600;"
            )
            edit_btn.setCursor(Qt.PointingHandCursor)
            edit_btn.clicked.connect(
                lambda _, ri=r: self._edit_log_entry(session_id, ri, rows_data, det_tbl, _fill_row)
            )
            det_tbl.setCellWidget(r, 6, edit_btn)
            det_tbl.setRowHeight(r, 40)

        for r, row in enumerate(rows_data):
            _fill_row(r, row)
        lay.addWidget(det_tbl)

        btn_row = QHBoxLayout()
        export_btn = QPushButton("Export Excel")
        export_btn.setStyleSheet(
            f"background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px;"
            "padding:8px 16px; font-weight:600;"
        )
        export_btn.setCursor(Qt.PointingHandCursor)
        export_btn.clicked.connect(lambda: self._export_session(course_title, date_str, rows_data))
        close_btn = QPushButton("Close")
        close_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; padding:8px 20px; font-weight:600;"
        )
        close_btn.clicked.connect(dlg.accept)
        btn_row.addWidget(export_btn)
        btn_row.addStretch()
        btn_row.addWidget(close_btn)
        lay.addLayout(btn_row)
        dlg.exec_()

    def _edit_log_entry(self, session_id, row_idx, rows_data, tbl, fill_row_fn):
        row = rows_data[row_idx]
        sid, name, current_status, current_note = row[0], row[1], row[4], row[5]

        dlg = QDialog(self)
        dlg.setWindowTitle(f"Edit — {name}")
        dlg.setFixedWidth(340)
        dlg.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(24, 20, 24, 20)
        lay.setSpacing(10)

        title_lbl = QLabel(f"Edit attendance for:\n{name}")
        title_lbl.setStyleSheet(f"font-size:14px; font-weight:700; color:{TEXT_DARK};")
        title_lbl.setWordWrap(True)
        lay.addWidget(title_lbl)

        group = QButtonGroup(dlg)
        buttons = {}
        for status in ["Present", "Late", "Early Leave", "Absent"]:
            rb = QRadioButton(status)
            rb.setStyleSheet(f"font-size:13px; color:{TEXT_MED};")
            if status == current_status:
                rb.setChecked(True)
            group.addButton(rb)
            buttons[status] = rb
            lay.addWidget(rb)

        note_lbl = QLabel("Note (optional)")
        note_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        note_field = QLineEdit()
        note_field.setPlaceholderText("e.g. Medical excuse, arrived late by bus…")
        note_field.setFixedHeight(36)
        note_field.setText(current_note)
        lay.addWidget(note_lbl)
        lay.addWidget(note_field)

        btn_row = QHBoxLayout()
        cancel = QPushButton("Cancel")
        cancel.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:7px 14px;"
        )
        cancel.clicked.connect(dlg.reject)
        save = QPushButton("Save")
        save.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px;"
            "padding:7px 14px; font-weight:700;"
        )
        save.clicked.connect(dlg.accept)
        btn_row.addWidget(cancel)
        btn_row.addStretch()
        btn_row.addWidget(save)
        lay.addLayout(btn_row)

        if dlg.exec_() != QDialog.Accepted:
            return

        new_status = next((s for s, rb in buttons.items() if rb.isChecked()), current_status)
        new_note   = note_field.text().strip()
        if new_status == current_status and new_note == current_note:
            return

        if DB_CONNECTED and session_id is not None:
            try:
                db["AttendanceLogs"].update_one(
                    {"SessionID": session_id, "StudentID": sid},
                    {"$set": {"Status": new_status, "Note": new_note}},
                )
            except Exception as e:
                QMessageBox.critical(self, "DB Error", f"Could not save:\n{e}")
                return

        rows_data[row_idx][4] = new_status
        rows_data[row_idx][5] = new_note
        fill_row_fn(row_idx, rows_data[row_idx])

    def _export_session(self, course_title, date_str, rows_data):
        safe_course = "".join(c for c in course_title if c.isalnum() or c in " _-").strip()
        default_name = f"Attendance_{safe_course}_{date_str}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Attendance", default_name, "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance"

            green_fill  = PatternFill("solid", start_color="1B5E35")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            center      = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("A1:F1")
            ws["A1"] = f"Attendance Report — {course_title}  |  {date_str}"
            ws["A1"].font      = Font(bold=True, size=13, color="1A1A2E")
            ws["A1"].alignment = center

            headers = ["Student ID", "Full Name", "First Seen", "Last Seen", "Status", "Note"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=h)
                cell.fill      = green_fill
                cell.font      = header_font
                cell.alignment = center

            for row_idx, (sid, name, fs, ls, status, note) in enumerate(rows_data, 3):
                ws.cell(row=row_idx, column=1, value=sid)
                ws.cell(row=row_idx, column=2, value=name)
                ws.cell(row=row_idx, column=3, value=fs)
                ws.cell(row=row_idx, column=4, value=ls)
                ws.cell(row=row_idx, column=5, value=status)
                ws.cell(row=row_idx, column=6, value=note)

            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            wb.save(path)
            QMessageBox.information(self, "Exported", f"Saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    def _go_back(self):
        self.hide()
        self.main_win.show()


class AnalyticsWindow(QWidget):
    def __init__(self, main_win):
        super().__init__()
        self.main_win = main_win
        self.setWindowTitle("AIAS — Attendance Analytics")
        import os
        from PyQt5.QtGui import QIcon
        _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aias_icon.ico")
        if os.path.exists(_icon_path):
            self.setWindowIcon(QIcon(_icon_path))
        self.setMinimumSize(1100, 700)
        self.resize(1100, 700)
        self.setStyleSheet(f"background:{BG};")
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(48, 36, 48, 28)
        root.setSpacing(16)

        title = QLabel("Attendance Analytics")
        title.setStyleSheet(f"font-size:24px; font-weight:800; color:{TEXT_DARK};")
        sub = QLabel("Per-student attendance rates across all completed sessions")
        sub.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY};")

        header_row = QHBoxLayout()
        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        title_col.addWidget(title)
        title_col.addWidget(sub)
        header_row.addLayout(title_col)
        header_row.addStretch()

        export_btn = QPushButton("📥  Export to Excel")
        export_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; "
            "padding:8px 16px; font-size:12px; font-weight:600;"
        )
        export_btn.setCursor(Qt.PointingHandCursor)
        export_btn.clicked.connect(lambda: self._export_analytics())
        header_row.addWidget(export_btn)
        root.addLayout(header_row)

        courses_data = getattr(self.main_win, "_courses_data", [])
        ctrl_row = QHBoxLayout()
        course_lbl = QLabel("Course:")
        course_lbl.setStyleSheet(f"font-size:13px; font-weight:600; color:{TEXT_MED};")
        self._course_combo = QComboBox()
        self._course_combo.setFixedHeight(36)
        self._course_combo.setMinimumWidth(300)
        for c in courses_data:
            self._course_combo.addItem(
                f"{c.get('SectionID', '')} — {c.get('CourseTitle', '')}",
                c
            )
        ctrl_row.addWidget(course_lbl)
        ctrl_row.addSpacing(8)
        ctrl_row.addWidget(self._course_combo)
        ctrl_row.addStretch()
        root.addLayout(ctrl_row)

        def _make_updatable(label_text, color=TEXT_DARK):
            f = QFrame()
            f.setStyleSheet(
                f"QFrame{{background:{WHITE};border:1px solid {BORDER};border-radius:8px;}}"
            )
            f.setFixedHeight(82)
            f.setMinimumWidth(130)
            fl = QVBoxLayout(f)
            fl.setContentsMargins(16, 12, 16, 12)
            fl.setSpacing(2)
            vl = QLabel("—")
            vl.setStyleSheet(
                f"font-size:26px; font-weight:800; color:{color};"
                "border:none; background:transparent;"
            )
            tl = QLabel(label_text)
            tl.setStyleSheet(
                f"font-size:11px; color:{TEXT_GRAY}; border:none; background:transparent;"
            )
            fl.addWidget(vl)
            fl.addWidget(tl)
            return f, vl

        sessions_card, self._lbl_sessions = _make_updatable("Total Sessions",   PRIMARY)
        avg_card,      self._lbl_avg      = _make_updatable("Avg Attendance %", "#22C55E")
        atrisk_card,   self._lbl_atrisk   = _make_updatable("At Risk (<60%)",   "#EF4444")

        stats_row = QHBoxLayout()
        stats_row.setSpacing(16)
        stats_row.addWidget(sessions_card)
        stats_row.addWidget(avg_card)
        stats_row.addWidget(atrisk_card)
        stats_row.addStretch()
        root.addLayout(stats_row)

        self._analytics_rows = []
        self._tbl = make_table(
            ["Student ID", "Full Name", "Attended", "Total Sessions", "Attendance Rate"],
            [],
            col_widths={0: 140, 2: 100, 3: 130, 4: 160},
            stretch_col=1,
        )
        self._tbl.setCursor(Qt.PointingHandCursor)
        self._tbl.cellClicked.connect(self._on_analytics_row_click)
        root.addWidget(self._tbl)

        hint = QLabel("Click any row to see per-session breakdown.")
        hint.setStyleSheet(f"font-size:11px; color:{TEXT_GRAY};")
        root.addWidget(hint)

        back_row = QHBoxLayout()
        back_row.addStretch()
        back_btn = QPushButton("← Back")
        back_btn.setFixedHeight(40)
        back_btn.setStyleSheet(
            f"background:{WHITE}; color:{PRIMARY}; border:1px solid {PRIMARY}; border-radius:6px;"
            "padding:8px 16px; font-size:13px; font-weight:600;"
        )
        back_btn.setCursor(Qt.PointingHandCursor)
        back_btn.clicked.connect(self._go_back)
        back_row.addWidget(back_btn)
        root.addLayout(back_row)

        self._course_combo.currentIndexChanged.connect(self._load_analytics)
        if courses_data:
            self._load_analytics(0)

    def _load_analytics(self, _idx=0):
        course = self._course_combo.currentData()
        if not course:
            return

        section_id   = str(course.get("SectionID", ""))
        enrolled_ids = [str(s) for s in course.get("EnrolledStudents", [])]

        sessions = []
        if DB_CONNECTED:
            try:
                sessions = list(db["Sessions"].find(
                    {"SectionID": section_id, "Status": "completed"}
                ).sort("StartTime", 1))
            except Exception as e:
                print(f"[AIAS] Analytics session load: {e}")

        n_sessions = len(sessions)
        self._lbl_sessions.setText(str(n_sessions))

        # Build per-student attendance count and per-session breakdown
        attended        = {sid: 0 for sid in enrolled_ids}
        student_sessions = {sid: [] for sid in enrolled_ids}

        if DB_CONNECTED and sessions:
            sess_ids   = [s["_id"] for s in sessions]
            sess_dates = {s["_id"]: s.get("StartTime") for s in sessions}
            try:
                for log in db["AttendanceLogs"].find({"SessionID": {"$in": sess_ids}}):
                    _s = str(log.get("StudentID", ""))
                    if _s in student_sessions:
                        _dt = sess_dates.get(log.get("SessionID"))
                        student_sessions[_s].append({
                            "date":   _dt,
                            "status": log.get("Status", "Absent"),
                        })
                        if log.get("Status") in ("Present", "Late"):
                            attended[_s] += 1
            except Exception as e:
                print(f"[AIAS] Analytics log error: {e}")

        name_map = {}
        if DB_CONNECTED and enrolled_ids:
            try:
                for stu in db["Students"].find({"StudentID": {"$in": enrolled_ids}}):
                    name_map[str(stu.get("StudentID", ""))] = stu.get("FullName", "")
            except Exception as e:
                print(f"[AIAS] Analytics student name load error: {e}")

        rows = [
            (sid, name_map.get(sid, sid), attended.get(sid, 0), n_sessions,
             (attended.get(sid, 0) / n_sessions * 100) if n_sessions > 0 else 0.0,
             student_sessions.get(sid, []))
            for sid in enrolled_ids
        ]
        self._analytics_rows = rows

        self._tbl.setRowCount(len(rows))
        for r, (sid, name, cnt, total, rate, _sess) in enumerate(rows):
            self._tbl.setItem(r, 0, QTableWidgetItem(sid))
            self._tbl.setItem(r, 1, QTableWidgetItem(name))
            self._tbl.setItem(r, 2, QTableWidgetItem(str(cnt)))
            self._tbl.setItem(r, 3, QTableWidgetItem(str(total)))

            if rate >= 80:
                bg, fg = "#EAF3DE", "#27500A"
            elif rate >= 60:
                bg, fg = "#FAEEDA", "#633806"
            else:
                bg, fg = "#FCEBEB", "#791F1F"

            rate_lbl = QLabel(f"{rate:.0f}%")
            rate_lbl.setAlignment(Qt.AlignCenter)
            rate_lbl.setFixedHeight(26)
            rate_lbl.setStyleSheet(
                f"background:{bg}; color:{fg}; border-radius:4px;"
                "padding:2px 10px; font-size:12px; font-weight:700;"
            )
            outer = QWidget()
            outer.setStyleSheet("background:transparent;")
            ol = QHBoxLayout(outer)
            ol.setContentsMargins(8, 4, 8, 4)
            ol.addWidget(rate_lbl)
            self._tbl.setCellWidget(r, 4, outer)
            self._tbl.setRowHeight(r, 44)

        avg     = (sum(r[4] for r in rows) / len(rows)) if rows else 0.0
        at_risk = sum(1 for r in rows if r[4] < 60)
        self._lbl_avg.setText(f"{avg:.0f}%")
        self._lbl_atrisk.setText(str(at_risk))

    def _on_analytics_row_click(self, row, _col):
        if row < len(self._analytics_rows):
            sid, name, _cnt, _total, _rate, sess_data = self._analytics_rows[row]
            self._show_student_sessions(sid, name, sess_data)

    def _show_student_sessions(self, sid, name, sess_data):
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Session History — {name}")
        dlg.setMinimumSize(480, 400)
        dlg.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(24, 20, 24, 20)
        lay.setSpacing(12)

        hdr = QLabel(f"{name}  ({sid})")
        hdr.setStyleSheet(f"font-size:15px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(hdr)

        if not sess_data:
            empty = QLabel("No attendance records found.")
            empty.setAlignment(Qt.AlignCenter)
            empty.setStyleSheet(f"color:{TEXT_GRAY}; font-size:13px;")
            lay.addWidget(empty)
        else:
            det_tbl = make_table(["Date", "Status"], sess_data, col_widths={0: 160}, stretch_col=1)
            det_tbl.setRowCount(len(sess_data))
            for r, entry in enumerate(sess_data):
                dt = entry.get("date")
                det_tbl.setItem(r, 0, QTableWidgetItem(
                    dt.strftime("%Y-%m-%d %H:%M") if dt else "—"
                ))
                det_tbl.setCellWidget(r, 1, make_badge(entry.get("status", "—")))
                det_tbl.setRowHeight(r, 40)
            lay.addWidget(det_tbl)

        close_btn = QPushButton("Close")
        close_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; padding:8px 20px; font-weight:600;"
        )
        close_btn.clicked.connect(dlg.accept)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_row.addWidget(close_btn)
        lay.addLayout(btn_row)
        dlg.exec_()

    def _export_analytics(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from PyQt5.QtWidgets import QFileDialog, QMessageBox
        import datetime

        path, _ = QFileDialog.getSaveFileName(
            self, "Save Analytics Report",
            f"AIAS_Analytics_{datetime.date.today()}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Analytics"

        # Header row
        headers = ["Student ID", "Full Name", "Sessions Present", "Total Sessions", "Attendance %", "Status"]
        header_fill = PatternFill("solid", fgColor="1B5E35")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows — read from the analytics table widget
        tbl = self._tbl
        for row in range(tbl.rowCount()):
            for col in range(tbl.columnCount()):
                item = tbl.item(row, col)
                val = item.text() if item else ""
                ws.cell(row=row+2, column=col+1, value=val)

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        try:
            wb.save(path)
            QMessageBox.information(self, "Export Successful", f"Analytics exported to:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Export Failed", f"Could not save file:\n{e}")

    def _go_back(self):
        self.hide()
        self.main_win.show()
