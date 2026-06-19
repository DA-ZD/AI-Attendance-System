import sys
import os

os.environ["ORT_LOG_SEVERITY_LEVEL"] = "3"
os.environ["ORT_LOG_LEVEL"] = "3"
os.environ["INSIGHTFACE_LOG_LEVEL"] = "0"

# Add torch's lib dir to DLL search path before PyQt5 imports alter it.
try:
    import importlib.util as _ilu
    _spec = _ilu.find_spec("torch")
    if _spec and _spec.origin:
        _torch_lib = os.path.join(os.path.dirname(_spec.origin), "lib")
        if os.path.isdir(_torch_lib):
            os.add_dll_directory(_torch_lib)
except Exception as _e:
    print(f"[AIAS] warning: {_e}")

try:
    import cv2 as _cv2_pre
    import onnxruntime as _ort_pre
    import torch as _torch_pre
    from ultralytics import YOLO as _YOLO_pre
    from insightface.app import FaceAnalysis as _FA_pre
    print("[AIAS] AI libraries pre-loaded")
except Exception as _pre_e:
    print(f"[AIAS] AI pre-load warning: {_pre_e}")

import pymongo
import bcrypt
import smtplib
import tempfile
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QWidget, QFrame, QLabel, QLineEdit, QPushButton,
    QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QStackedWidget, QProgressBar,
    QComboBox, QSizePolicy, QFileDialog, QMessageBox,
    QDialog, QCheckBox, QScrollArea, QRadioButton, QButtonGroup,
    QGraphicsDropShadowEffect,
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QFont, QImage, QPixmap, QColor, QPainter
from concurrent.futures import ThreadPoolExecutor
import threading as _threading

import torch
import torch.nn as nn
import torch.nn.functional as _F
try:
    import pytorch_lightning as _pl
    _PL_AVAILABLE = True
except ImportError:
    _PL_AVAILABLE = False

MONGO_URI = "mongodb://localhost:27017"
DB_NAME = "AIAS_DB"
try:
    client = pymongo.MongoClient(MONGO_URI, serverSelectionTimeoutMS=3000)
    client.server_info()
    db = client[DB_NAME]
    DB_CONNECTED = True
except Exception:
    client = None
    db = None
    DB_CONNECTED = False

if DB_CONNECTED:
    try:
        db.command("collMod", "AttendanceLogs", validator={}, validationLevel="off")
        print("[AIAS] AttendanceLogs validator disabled")
    except Exception as _e:
        print(f"[AIAS] warning: {_e}")

SMTP_SERVER   = "smtp-relay.brevo.com"
SMTP_PORT     = 587

LATE_CUTOFF_MINUTES        = 10
EARLY_LEAVE_CUTOFF_MINUTES = 15
RECOGNITION_THRESHOLD      = 0.55

def _load_settings():
    global LATE_CUTOFF_MINUTES, EARLY_LEAVE_CUTOFF_MINUTES, RECOGNITION_THRESHOLD
    if not DB_CONNECTED:
        return
    try:
        doc = db["Settings"].find_one({"_id": "global"})
        if doc:
            LATE_CUTOFF_MINUTES        = doc.get("LateCutoffMinutes",       10)
            EARLY_LEAVE_CUTOFF_MINUTES = doc.get("EarlyLeaveCutoffMinutes", 15)
            RECOGNITION_THRESHOLD      = doc.get("RecognitionThreshold",    0.55)
    except Exception as e:
        print(f"[AIAS] Settings load error: {e}")

_load_settings()


def _ensure_admin_exists():
    if not DB_CONNECTED:
        return
    try:
        if not db["Instructors"].find_one({"Username": "admin"}):
            import secrets, string
            _chars = string.ascii_letters + string.digits
            _rand_pw = ''.join(secrets.choice(_chars) for _ in range(12))
            hashed = bcrypt.hashpw(_rand_pw.encode(), bcrypt.gensalt())
            print(f"[AIAS] Default admin account created. Password: {_rand_pw}")
            print(f"[AIAS] Please change this password after first login.")
            db["Instructors"].insert_one({
                "Username": "admin",
                "FullName": "Administrator",
                "Password": hashed,
                "Role": "admin",
                "AssignedSections": [],
            })
            print("[AIAS] Default admin account created.")
    except Exception as e:
        print(f"[AIAS] Admin setup error: {e}")


AI_LOADED     = False
yolo_model    = None
_gfpgan_model = None
_codeformer_model = None
_codeformer_net   = None
arcface_model = None
_adaface_model = None
_sr_model      = None
_adaface_db    = {}

_AI_LOADER_THREAD = None   # kept alive so Qt doesn't GC it


# ── AdaFace IR50 ──────────────────────────────────────────────────────────────

class _AdaBottleneck(nn.Module):
    expansion = 4
    def __init__(self, in_planes, planes, stride=1):
        super().__init__()
        self.conv1 = nn.Conv2d(in_planes, planes, 1, bias=False)
        self.bn1   = nn.BatchNorm2d(planes)
        self.conv2 = nn.Conv2d(planes, planes, 3, stride=stride, padding=1, bias=False)
        self.bn2   = nn.BatchNorm2d(planes)
        self.conv3 = nn.Conv2d(planes, planes*4, 1, bias=False)
        self.bn3   = nn.BatchNorm2d(planes*4)
        self.shortcut = nn.Sequential()
        if stride != 1 or in_planes != planes*4:
            self.shortcut = nn.Sequential(
                nn.Conv2d(in_planes, planes*4, 1, stride=stride, bias=False),
                nn.BatchNorm2d(planes*4))
    def forward(self, x):
        out = _F.relu(self.bn1(self.conv1(x)))
        out = _F.relu(self.bn2(self.conv2(out)))
        out = self.bn3(self.conv3(out))
        out += self.shortcut(x)
        return _F.relu(out)

class _AdaIResNet50(nn.Module):
    def __init__(self):
        super().__init__()
        self.in_planes = 64
        self.conv1  = nn.Conv2d(3, 64, 3, stride=1, padding=1, bias=False)
        self.bn1    = nn.BatchNorm2d(64)
        self.prelu  = nn.PReLU(64)
        self.layer1 = self._make(64,  3, 2)
        self.layer2 = self._make(128, 4, 2)
        self.layer3 = self._make(256, 6, 2)
        self.layer4 = self._make(512, 3, 2)
        self.bn2    = nn.BatchNorm2d(2048)
        self.drop   = nn.Dropout(0.4)
        self.fc     = nn.Linear(32768, 512)
        self.features = nn.BatchNorm1d(512)
    def _make(self, planes, n, stride):
        layers = [_AdaBottleneck(self.in_planes, planes, stride)]
        self.in_planes = planes * 4
        for _ in range(1, n):
            layers.append(_AdaBottleneck(self.in_planes, planes))
        return nn.Sequential(*layers)
    def forward(self, x):
        x = self.prelu(self.bn1(self.conv1(x)))
        x = self.layer1(x); x = self.layer2(x)
        x = self.layer3(x); x = self.layer4(x)
        x = self.bn2(x); x = self.drop(x)
        x = _F.adaptive_avg_pool2d(x, (4, 4))
        x = x.view(x.size(0), -1)
        x = self.fc(x); x = self.features(x)
        return x

def _load_adaface(ckpt_path="adaface_ir50_ms1mv2.ckpt"):
    """Load AdaFace model. Returns model or None if unavailable."""
    if not os.path.exists(ckpt_path):
        return None
    try:
        if _PL_AVAILABLE:
            torch.serialization.add_safe_globals([
                _pl.callbacks.model_checkpoint.ModelCheckpoint])
        ckpt = torch.load(ckpt_path, map_location="cpu", weights_only=False)
        sd = ckpt.get("state_dict", ckpt.get("model", ckpt))
        clean = {k.replace("model.", "").replace("module.", ""): v for k, v in sd.items()}
        model = _AdaIResNet50()
        for k, v in clean.items():
            if "fc.weight" in k:
                model.fc = nn.Linear(v.shape[1], v.shape[0]); break
        model.load_state_dict(clean, strict=False)
        model.eval()
        print("[AIAS] AdaFace loaded")
        return model
    except Exception as e:
        print(f"[AIAS] AdaFace load failed: {e}")
        return None

def _adaface_embedding(model, img_bgr):
    """Get normalized 512-dim embedding from BGR face crop."""
    import cv2, numpy as np
    try:
        img = cv2.resize(img_bgr, (112, 112))
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        img = img.astype(np.float32) / 127.5 - 1.0
        t = torch.from_numpy(img.transpose(2, 0, 1)).unsqueeze(0)
        with torch.no_grad():
            emb = model(t)
            emb = emb / emb.norm(dim=1, keepdim=True)
        return emb.squeeze(0).numpy()
    except Exception as e:
        print(f"[AIAS] AdaFace embedding error: {e}")
        return None


# ── Real-ESRGAN ───────────────────────────────────────────────────────────────

class _RDB(nn.Module):
    def __init__(self, nf=64, gc=32):
        super().__init__()
        self.c1=nn.Conv2d(nf,gc,3,1,1); self.c2=nn.Conv2d(nf+gc,gc,3,1,1)
        self.c3=nn.Conv2d(nf+2*gc,gc,3,1,1); self.c4=nn.Conv2d(nf+3*gc,gc,3,1,1)
        self.c5=nn.Conv2d(nf+4*gc,nf,3,1,1)
        self.act=nn.LeakyReLU(0.2,inplace=True)
    def forward(self,x):
        x1=self.act(self.c1(x)); x2=self.act(self.c2(torch.cat((x,x1),1)))
        x3=self.act(self.c3(torch.cat((x,x1,x2),1))); x4=self.act(self.c4(torch.cat((x,x1,x2,x3),1)))
        return self.c5(torch.cat((x,x1,x2,x3,x4),1))*0.2+x

class _RRDB(nn.Module):
    def __init__(self, nf, gc=32):
        super().__init__()
        self.r1=_RDB(nf,gc); self.r2=_RDB(nf,gc); self.r3=_RDB(nf,gc)
    def forward(self,x): return self.r3(self.r2(self.r1(x)))*0.2+x

class _RRDBNet(nn.Module):
    def __init__(self):
        super().__init__()
        self.cf=nn.Conv2d(3,64,3,1,1)
        self.body=nn.Sequential(*[_RRDB(64) for _ in range(23)])
        self.cb=nn.Conv2d(64,64,3,1,1)
        self.u1=nn.Conv2d(64,64,3,1,1); self.u2=nn.Conv2d(64,64,3,1,1)
        self.hr=nn.Conv2d(64,64,3,1,1); self.cl=nn.Conv2d(64,3,3,1,1)
        self.act=nn.LeakyReLU(0.2,inplace=True)
    def forward(self,x):
        f=self.cf(x); f=f+self.cb(self.body(f))
        f=self.act(self.u1(_F.interpolate(f,scale_factor=2,mode='nearest')))
        f=self.act(self.u2(_F.interpolate(f,scale_factor=2,mode='nearest')))
        return self.cl(self.act(self.hr(f)))

def _load_realesrgan(model_path="RealESRGAN_x4plus.pth"):
    """Load Real-ESRGAN model. Returns model or None if unavailable."""
    if not os.path.exists(model_path):
        return None
    try:
        import re
        model = _RRDBNet()
        raw = torch.load(model_path, map_location="cpu", weights_only=False)
        sd = raw.get("params_ema", raw.get("params", raw))

        # Remap checkpoint keys to match our _RRDBNet attribute names
        key_map = {
            "conv_first": "cf",
            "conv_body":  "cb",
            "conv_up1":   "u1",
            "conv_up2":   "u2",
            "conv_hr":    "hr",
            "conv_last":  "cl",
        }
        remapped = {}
        for k, v in sd.items():
            new_k = k
            for old, new in key_map.items():
                if new_k.startswith(old + "."):
                    new_k = new + "." + new_k[len(old) + 1:]
                    break
            new_k = re.sub(
                r'body\.(\d+)\.rdb(\d+)\.conv(\d+)',
                lambda m: f'body.{m.group(1)}.r{m.group(2)}.c{m.group(3)}',
                new_k
            )
            remapped[new_k] = v

        model.load_state_dict(remapped, strict=True)
        model.eval()
        print("[AIAS] Real-ESRGAN loaded")
        return model
    except Exception as e:
        print(f"[AIAS] Real-ESRGAN load failed: {e}")
        return None

def _load_gfpgan():
    """Load GFPGAN face enhancement model."""
    try:
        from gfpgan import GFPGANer
        import os
        model_path = "GFPGANv1.4.pth"
        if not os.path.exists(model_path):
            import urllib.request
            print("[AIAS] Downloading GFPGAN model...")
            url = "https://github.com/TencentARC/GFPGAN/releases/download/v1.3.4/GFPGANv1.4.pth"
            urllib.request.urlretrieve(url, model_path)
        model = GFPGANer(
            model_path=model_path,
            upscale=2,
            arch='clean',
            channel_multiplier=2,
            bg_upsampler=None
        )
        print("[AIAS] GFPGAN loaded")
        return model
    except Exception as e:
        print(f"[AIAS] GFPGAN load failed: {e}")
        return None

def _load_codeformer():
    """Load CodeFormer face restoration model."""
    try:
        import sys
        import os
        import torch
        # Add CodeFormer to path
        cf_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               "CodeFormer-master", "CodeFormer-master")
        if cf_path not in sys.path:
            sys.path.insert(0, cf_path)

        from basicsr.utils.download_util import load_file_from_url
        from facelib.utils.face_restoration_helper import FaceRestoreHelper
        from basicsr.archs.rrdbnet_arch import RRDBNet
        from basicsr.utils.realesrgan_utils import RealESRGANer

        # Try to import CodeFormer net
        try:
            from basicsr.archs.codeformer_arch import CodeFormer as CodeFormerNet
        except ImportError:
            # Try alternative import path
            sys.path.insert(0, os.path.join(cf_path, "basicsr", "archs"))
            from codeformer_arch import CodeFormer as CodeFormerNet

        # Download model if not exists
        model_path = os.path.join(cf_path, "weights", "CodeFormer", "codeformer.pth")
        os.makedirs(os.path.dirname(model_path), exist_ok=True)
        if not os.path.exists(model_path):
            print("[AIAS] Downloading CodeFormer model...")
            url = "https://github.com/sczhou/CodeFormer/releases/download/v0.1.0/codeformer.pth"
            load_file_from_url(url=url, model_dir=os.path.dirname(model_path),
                               progress=True, file_name="codeformer.pth")

        # Load net
        net = CodeFormerNet(dim_embd=512, codebook_size=1024, n_head=8,
                            n_layers=9, connect_list=['32', '64', '128', '256'])
        checkpoint = torch.load(model_path, map_location=torch.device('cpu'))
        net.load_state_dict(checkpoint['params_ema'])
        net.eval()

        # Load face helper
        face_helper = FaceRestoreHelper(
            upscale_factor=2,
            face_size=512,
            crop_ratio=(1, 1),
            det_model='retinaface_resnet50',
            save_ext='png',
            use_parse=True,
            device=torch.device('cpu')
        )

        print("[AIAS] CodeFormer loaded")
        return net, face_helper
    except Exception as e:
        print(f"[AIAS] CodeFormer load failed: {e}")
        return None, None

def _sr_upscale(sr_model, img_bgr, target_size=112):
    """Upscale small face crop using Real-ESRGAN."""
    import cv2, numpy as np
    if sr_model is None:
        return img_bgr
    h, w = img_bgr.shape[:2]
    if min(h, w) >= target_size:
        return img_bgr
    try:
        img = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB).astype(np.float32) / 255.0
        t = torch.from_numpy(img.transpose(2, 0, 1)).unsqueeze(0)
        with torch.no_grad():
            out = sr_model(t).squeeze(0).clamp(0, 1).numpy().transpose(1, 2, 0)
        out = (out * 255).astype(np.uint8)
        return cv2.cvtColor(out, cv2.COLOR_RGB2BGR)
    except Exception as e:
        print(f"[AIAS] SR upscale error: {e}")
        return img_bgr


def _ensure_ai_models(det_size=(640, 640)):
    """Load YOLO + ArcFace into globals if not already loaded. Returns (yolo, arcface, ok)."""
    global yolo_model, arcface_model, AI_LOADED, _adaface_model, _sr_model, _gfpgan_model, _codeformer_model, _codeformer_net
    if AI_LOADED and yolo_model is not None and arcface_model is not None:
        # Re-prepare with requested det_size (lightweight operation)
        try:
            arcface_model.prepare(ctx_id=-1, det_size=det_size)
        except Exception:
            pass
        if _adaface_model is None:
            _adaface_model = _load_adaface("adaface_ir50_ms1mv2.ckpt")
        if _sr_model is None:
            _sr_model = _load_realesrgan("RealESRGAN_x4plus.pth")
        if _gfpgan_model is None:
            _gfpgan_model = _load_gfpgan()
        if _codeformer_net is None:
            _codeformer_net, _codeformer_model = _load_codeformer()
        return yolo_model, arcface_model, True
    try:
        import logging
        logging.getLogger("insightface").setLevel(logging.ERROR)
        logging.getLogger("onnxruntime").setLevel(logging.ERROR)
        from ultralytics import YOLO
        from insightface.app import FaceAnalysis
        import os
        _yolo_path = "yolov11l-face.pt" if os.path.exists("yolov11l-face.pt") else "yolov8n-face.pt"
        yolo_model = YOLO(_yolo_path)
        print(f"[AIAS] Using YOLO model: {_yolo_path}")
        try:
            arcface_model = FaceAnalysis(name="buffalo_l", providers=["CPUExecutionProvider"])
        except TypeError:
            arcface_model = FaceAnalysis(name="buffalo_l")
        arcface_model.prepare(ctx_id=-1, det_size=det_size)
        AI_LOADED = True
        print(f"[AIAS] AI models loaded (det_size={det_size})")
        if _adaface_model is None:
            _adaface_model = _load_adaface("adaface_ir50_ms1mv2.ckpt")
        if _sr_model is None:
            _sr_model = _load_realesrgan("RealESRGAN_x4plus.pth")
        if _gfpgan_model is None:
            _gfpgan_model = _load_gfpgan()
        if _codeformer_net is None:
            _codeformer_net, _codeformer_model = _load_codeformer()
        return yolo_model, arcface_model, True
    except Exception as e:
        print(f"[AIAS] AI model load failed: {e}")
        return None, None, False

PRIMARY = "#1B5E35"
PRIMARY_LIGHT = "#C8E0D0"
BG = "#F8F9FA"
WHITE = "#FFFFFF"
TEXT_DARK = "#1A1A2E"
TEXT_MED = "#374151"
TEXT_GRAY = "#6B7280"
BORDER = "#E5E7EB"

# Dark mode color palette
DARK_PRIMARY       = "#4ecf8e"
DARK_PRIMARY_LIGHT = "#0D2B18"
DARK_BG            = "#0A0A0F"
DARK_WHITE         = "#111118"
DARK_TEXT_DARK     = "#F1F1F5"
DARK_TEXT_MED      = "#C4C4D4"
DARK_TEXT_GRAY     = "#7777AA"
DARK_BORDER        = "#222235"

IS_DARK_MODE = False

BADGE_MAP = {
    "Present": ("#EAF3DE", "#27500A"),
    "Late": ("#FAEEDA", "#633806"),
    "Absent": ("#FCEBEB", "#791F1F"),
    "Early Leave": ("#E6F1FB", "#0C447C"),
    "Enrolled": ("#EAF3DE", "#27500A"),
    "Ready": ("#EAF3DE", "#27500A"),
    "No photo": ("#FAEEDA", "#633806"),
}

COURSES  = []
STUDENTS = []

GLOBAL_QSS = f"""
QWidget {{
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 13px;
}}
QLineEdit {{
    border: 1px solid {BORDER};
    border-radius: 6px;
    padding: 8px 12px;
    background: {WHITE};
    color: {TEXT_DARK};
}}
QLineEdit:focus {{
    border: 1.5px solid {PRIMARY};
}}
QPushButton {{
    border: none;
    border-radius: 6px;
    padding: 8px 16px;
    font-size: 13px;
    font-weight: 600;
}}
QTableWidget {{
    border: none;
    outline: none;
    background: {WHITE};
    alternate-background-color: #FAFAFA;
    selection-background-color: #D4EBE1;
    selection-color: {TEXT_DARK};
    gridline-color: transparent;
}}
QTableWidget::item {{
    padding: 6px 10px;
    border: none;
}}
QHeaderView::section {{
    background: #F5F5F5;
    color: {TEXT_GRAY};
    border: none;
    border-bottom: 1px solid {BORDER};
    padding: 8px 10px;
    font-size: 12px;
    font-weight: 600;
}}
QScrollBar:vertical {{
    border: none;
    background: {BG};
    width: 8px;
    border-radius: 4px;
}}
QScrollBar::handle:vertical {{
    background: #D1D5DB;
    border-radius: 4px;
    min-height: 20px;
}}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
    height: 0;
}}
QComboBox {{
    border: 1px solid {BORDER};
    border-radius: 6px;
    padding: 6px 10px;
    background: {WHITE};
    color: {TEXT_DARK};
}}
QComboBox::drop-down {{
    border: none;
    padding-right: 8px;
}}
QProgressBar {{
    background: #E5E7EB;
    border-radius: 4px;
    border: none;
}}
QProgressBar::chunk {{
    background: {PRIMARY};
    border-radius: 4px;
}}
"""


def apply_theme(app, dark=False):
    global IS_DARK_MODE, PRIMARY, PRIMARY_LIGHT, BG, WHITE
    global TEXT_DARK, TEXT_MED, TEXT_GRAY, BORDER, GLOBAL_QSS

    IS_DARK_MODE = dark

    if dark:
        PRIMARY       = DARK_PRIMARY
        PRIMARY_LIGHT = DARK_PRIMARY_LIGHT
        BG            = DARK_BG
        WHITE         = DARK_WHITE
        TEXT_DARK     = DARK_TEXT_DARK
        TEXT_MED      = DARK_TEXT_MED
        TEXT_GRAY     = DARK_TEXT_GRAY
        BORDER        = DARK_BORDER
    else:
        PRIMARY       = "#1B5E35"
        PRIMARY_LIGHT = "#C8E0D0"
        BG            = "#F8F9FA"
        WHITE         = "#FFFFFF"
        TEXT_DARK     = "#1A1A2E"
        TEXT_MED      = "#374151"
        TEXT_GRAY     = "#6B7280"
        BORDER        = "#E5E7EB"

    alt_row   = "#141420" if dark else "#FAFAFA"
    sel_row   = "#1A2E24" if dark else "#D4EBE1"
    header_bg = "#111118" if dark else "#F5F5F5"
    scroll_handle = "#444466" if dark else "#D1D5DB"
    input_bg  = WHITE
    btn_cancel_bg = "#1C1C2E" if dark else "#F3F4F6"

    qss = f"""
QWidget, QDialog, QMainWindow {{
    background-color: {BG};
    color: {TEXT_DARK};
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 13px;
}}
QFrame {{
    background-color: {BG};
    color: {TEXT_DARK};
}}
QLabel {{
    background-color: transparent;
    color: {TEXT_DARK};
}}
QLineEdit {{
    border: 1px solid {BORDER};
    border-radius: 6px;
    padding: 8px 12px;
    background-color: {WHITE};
    color: {TEXT_DARK};
}}
QLineEdit:focus {{
    border: 1.5px solid {PRIMARY};
}}
QPushButton {{
    border: none;
    border-radius: 6px;
    padding: 8px 16px;
    font-size: 13px;
    font-weight: 600;
    background-color: {WHITE};
    color: {TEXT_DARK};
}}
QPushButton:hover {{
    background-color: {PRIMARY_LIGHT};
    color: {PRIMARY};
}}
QPushButton:pressed {{
    padding: 9px 15px 7px 17px;
    background-color: {PRIMARY};
    color: #ffffff;
}}
QPushButton[class="primary"]:hover {{
    background-color: {PRIMARY};
    color: #ffffff;
    opacity: 0.88;
}}
QComboBox {{
    border: 1px solid {BORDER};
    border-radius: 6px;
    padding: 6px 10px;
    background-color: {WHITE};
    color: {TEXT_DARK};
}}
QComboBox QAbstractItemView {{
    background-color: {WHITE};
    color: {TEXT_DARK};
    border: 1px solid {BORDER};
    selection-background-color: {PRIMARY_LIGHT};
    selection-color: {TEXT_DARK};
}}
QComboBox::drop-down {{
    border: none;
    padding-right: 8px;
}}
QTableWidget {{
    border: none;
    outline: none;
    background-color: {WHITE};
    color: {TEXT_DARK};
    gridline-color: transparent;
    alternate-background-color: {BG};
}}
QTableWidget::item {{
    padding: 6px 10px;
    border: none;
    color: {TEXT_DARK};
    background-color: transparent;
}}
QTableWidget::item:selected {{
    background-color: {PRIMARY_LIGHT};
    color: {TEXT_DARK};
}}
QHeaderView::section {{
    background-color: {BG};
    color: {TEXT_GRAY};
    border: none;
    border-bottom: 1px solid {BORDER};
    padding: 8px 10px;
    font-size: 12px;
    font-weight: 600;
}}
QScrollBar:vertical {{
    border: none;
    background-color: {BG};
    width: 8px;
    border-radius: 4px;
}}
QScrollBar::handle:vertical {{
    background-color: {BORDER};
    border-radius: 4px;
    min-height: 20px;
}}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
    height: 0;
}}
QScrollArea {{
    border: none;
    background-color: transparent;
}}
QScrollArea > QWidget > QWidget {{
    background-color: {BG};
}}
QProgressBar {{
    background-color: {BORDER};
    border-radius: 4px;
    border: none;
}}
QProgressBar::chunk {{
    background-color: {PRIMARY};
    border-radius: 4px;
}}
QCheckBox {{
    color: {TEXT_DARK};
    spacing: 8px;
    background-color: transparent;
}}
QCheckBox::indicator {{
    width: 16px;
    height: 16px;
    border: 1px solid {BORDER};
    border-radius: 3px;
    background-color: {WHITE};
}}
QCheckBox::indicator:checked {{
    background-color: {PRIMARY};
    border-color: {PRIMARY};
}}
QRadioButton {{
    color: {TEXT_DARK};
    spacing: 8px;
    background-color: transparent;
}}
QRadioButton::indicator:checked {{
    background-color: {PRIMARY};
    border-color: {PRIMARY};
}}
QMessageBox {{
    background-color: {BG};
    color: {TEXT_DARK};
}}
QStackedWidget {{
    background-color: {BG};
}}
"""
    app.setStyleSheet(qss)
    for widget in app.topLevelWidgets():
        widget.setStyleSheet(f"background-color:{BG};")
        force_theme_on_all_widgets(widget)
        widget.update()


def force_theme_on_all_widgets(root_widget):
    from PyQt5.QtGui import QPalette, QColor

    # Stylesheet string replacements: baked hex values → current theme values.
    # Handles both bg/text/border in either direction (light↔dark).
    replacements = [
        # Backgrounds — light variants → BG
        ("background-color:#F8F9FA",  f"background-color:{BG}"),
        ("background-color: #F8F9FA", f"background-color:{BG}"),
        ("background:#F8F9FA",        f"background:{BG}"),
        ("background: #F8F9FA",       f"background:{BG}"),
        ("background-color:#F5F5F5",  f"background-color:{BG}"),
        ("background:#F5F5F5",        f"background:{BG}"),
        ("background-color:#FAFAFA",  f"background-color:{BG}"),
        ("background:#FAFAFA",        f"background:{BG}"),
        ("background-color:#F3F4F6",  f"background-color:{BG}"),
        ("background:#F3F4F6",        f"background:{BG}"),
        # Backgrounds — dark variant → BG
        ("background-color:#0A0A0F",  f"background-color:{BG}"),
        ("background:#0A0A0F",        f"background:{BG}"),
        # White/surface — both directions
        ("background-color:#FFFFFF",  f"background-color:{WHITE}"),
        ("background:#FFFFFF",        f"background:{WHITE}"),
        ("background-color:#111118",  f"background-color:{WHITE}"),
        ("background:#111118",        f"background:{WHITE}"),
        ("background:white",          f"background:{WHITE}"),
        ("background: white",         f"background:{WHITE}"),
        # Text — both directions
        ("color:#1A1A2E",  f"color:{TEXT_DARK}"),
        ("color: #1A1A2E", f"color:{TEXT_DARK}"),
        ("color:#F1F1F5",  f"color:{TEXT_DARK}"),
        ("color:#374151",  f"color:{TEXT_MED}"),
        ("color:#C4C4D4",  f"color:{TEXT_MED}"),
        ("color:#6B7280",  f"color:{TEXT_GRAY}"),
        ("color:#7777AA",  f"color:{TEXT_GRAY}"),
        # Borders — both directions
        ("border:1px solid #E5E7EB",        f"border:1px solid {BORDER}"),
        ("border: 1px solid #E5E7EB",       f"border:1px solid {BORDER}"),
        ("border:1px solid #222235",        f"border:1px solid {BORDER}"),
        ("border-bottom:1px solid #E5E7EB", f"border-bottom:1px solid {BORDER}"),
        ("border-bottom: 1px solid #E5E7EB",f"border-bottom:1px solid {BORDER}"),
        ("border-bottom:1px solid #222235", f"border-bottom:1px solid {BORDER}"),
        # PRIMARY — both directions
        ("background:#1B5E35",  f"background:{PRIMARY}"),
        ("background:#4ecf8e",  f"background:{PRIMARY}"),
        ("background:#166E42",  f"background:{PRIMARY}"),
        # PRIMARY_LIGHT — both directions
        ("background:#C8E0D0",  f"background:{PRIMARY_LIGHT}"),
        ("background:#0D2B18",  f"background:{PRIMARY_LIGHT}"),
    ]

    palette = QPalette()
    palette.setColor(QPalette.Window,          QColor(BG))
    palette.setColor(QPalette.WindowText,      QColor(TEXT_DARK))
    palette.setColor(QPalette.Base,            QColor(WHITE))
    palette.setColor(QPalette.AlternateBase,   QColor(BG))
    palette.setColor(QPalette.Text,            QColor(TEXT_DARK))
    palette.setColor(QPalette.Button,          QColor(WHITE))
    palette.setColor(QPalette.ButtonText,      QColor(TEXT_DARK))
    palette.setColor(QPalette.Highlight,       QColor(PRIMARY))
    palette.setColor(QPalette.HighlightedText, QColor("#ffffff"))

    def patch(w):
        w.setPalette(palette)
        w.setAutoFillBackground(True)
        ss = w.styleSheet()
        if ss:
            for old, new in replacements:
                if old in ss:
                    ss = ss.replace(old, new)
            w.setStyleSheet(ss)
        w.update()

    patch(root_widget)
    for child in root_widget.findChildren(QWidget):
        patch(child)
    root_widget.repaint()


def h_sep():
    line = QFrame()
    line.setFrameShape(QFrame.HLine)
    line.setStyleSheet(f"background:{BORDER}; max-height:1px; border:none;")
    return line


def make_badge(text):
    bg, fg = BADGE_MAP.get(text, ("#E5E7EB", TEXT_MED))
    outer = QWidget()
    outer.setStyleSheet("background:transparent;")
    ol = QHBoxLayout(outer)
    ol.setContentsMargins(8, 4, 8, 4)
    lbl = QLabel(text)
    lbl.setAlignment(Qt.AlignCenter)
    lbl.setFixedHeight(26)
    lbl.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
    lbl.setStyleSheet(
        f"background:{bg}; color:{fg}; border-radius:4px;"
        f"padding:2px 10px; font-size:12px; font-weight:600;"
    )
    ol.addWidget(lbl)
    return outer


def make_stat_card(title, value, value_color=TEXT_DARK):
    frame = QFrame()
    frame.setStyleSheet(
        f"QFrame {{ background:{WHITE}; border:1px solid {BORDER}; border-radius:8px; }}"
    )
    frame.setFixedHeight(82)
    frame.setMinimumWidth(130)
    lay = QVBoxLayout(frame)
    lay.setContentsMargins(16, 12, 16, 12)
    lay.setSpacing(2)
    v = QLabel(str(value))
    v.setStyleSheet(
        f"font-size:26px; font-weight:800; color:{value_color}; border:none; background:transparent;"
    )
    t = QLabel(title)
    t.setStyleSheet(f"font-size:11px; color:{TEXT_GRAY}; border:none; background:transparent;")
    lay.addWidget(v)
    lay.addWidget(t)
    return frame


def make_avatar(initials, size=44, bg=PRIMARY, fg=WHITE):
    lbl = QLabel(initials)
    lbl.setFixedSize(size, size)
    lbl.setAlignment(Qt.AlignCenter)
    lbl.setStyleSheet(
        f"background:{bg}; color:{fg}; border-radius:{size // 2}px;"
        f"font-size:{size // 2 - 2}px; font-weight:800;"
    )
    return lbl


def make_sidebar_base(initials, name, role):
    sidebar = QFrame()
    sidebar.setFixedWidth(200)
    sidebar.setStyleSheet(
        f"QFrame {{ background:{WHITE}; border-right:1px solid {BORDER}; border-top:none;"
        f"border-left:none; border-bottom:none; }}"
    )
    lay = QVBoxLayout(sidebar)
    lay.setContentsMargins(16, 24, 16, 16)
    lay.setSpacing(6)

    av = make_avatar(initials)
    name_lbl = QLabel(name)
    name_lbl.setAlignment(Qt.AlignCenter)
    name_lbl.setStyleSheet(
        f"font-weight:700; color:{TEXT_DARK}; font-size:13px; background:transparent; border:none;"
    )
    role_lbl = QLabel(role)
    role_lbl.setAlignment(Qt.AlignCenter)
    role_lbl.setStyleSheet(
        f"color:{TEXT_GRAY}; font-size:11px; background:transparent; border:none;"
    )
    lay.addWidget(av, alignment=Qt.AlignCenter)
    lay.addWidget(name_lbl)
    lay.addWidget(role_lbl)
    lay.addSpacing(10)
    lay.addWidget(h_sep())
    lay.addSpacing(6)
    return sidebar, lay


def make_table(cols, rows_data, col_widths=None, stretch_col=1):
    tbl = QTableWidget(len(rows_data), len(cols))
    tbl.setStyleSheet(f"QTableWidget {{ background-color:{WHITE}; color:{TEXT_DARK}; }} QTableWidget::item {{ color:{TEXT_DARK}; }} QHeaderView::section {{ background-color:{BG}; color:{TEXT_GRAY}; border-bottom:1px solid {BORDER}; }}")
    tbl.setHorizontalHeaderLabels(cols)
    tbl.horizontalHeader().setSectionResizeMode(stretch_col, QHeaderView.Stretch)
    if col_widths:
        for ci, w in col_widths.items():
            if ci != stretch_col:
                tbl.setColumnWidth(ci, w)
    tbl.verticalHeader().setVisible(False)
    tbl.setAlternatingRowColors(True)
    tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
    tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
    tbl.setFocusPolicy(Qt.NoFocus)
    tbl.setShowGrid(False)
    return tbl


def apply_clahe(img_bgr):
    """
    Apply adaptive CLAHE to a BGR face image.
    Automatically adjusts clip limit based on image brightness.
    Returns BGR image.
    """
    import cv2
    import numpy as np

    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    brightness = np.mean(gray)

    if brightness < 60:
        clip_limit = 4.0
    elif brightness < 110:
        clip_limit = 2.5
    elif brightness < 170:
        clip_limit = 1.5
    else:
        clip_limit = 1.0

    lab = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=clip_limit, tileGridSize=(8, 8))
    l_clahe = clahe.apply(l)
    lab_clahe = cv2.merge([l_clahe, a, b])
    return cv2.cvtColor(lab_clahe, cv2.COLOR_LAB2BGR)


def upscale_face_crop(img_bgr, target_size=112):
    """
    Upscale a small face crop to target_size using bicubic interpolation.
    Only upscales if the face is smaller than target_size.
    Returns BGR image.
    """
    import cv2
    h, w = img_bgr.shape[:2]
    min_dim = min(h, w)

    if min_dim >= target_size:
        return img_bgr

    scale = target_size / min_dim
    new_w = int(w * scale)
    new_h = int(h * scale)

    return cv2.resize(img_bgr, (new_w, new_h), interpolation=cv2.INTER_CUBIC)


def compute_dynamic_threshold(face_crop_bgr):
    """
    Compute a dynamic recognition threshold based on face image quality.
    Returns a float threshold between 0.28 and 0.52.

    Lower quality → lower threshold (try harder to recognize)
    Higher quality → higher threshold (be more strict)
    """
    import cv2
    import numpy as np

    h, w = face_crop_bgr.shape[:2]
    face_area = h * w

    # ── Factor 1: Face size (distance proxy) ──────────────────────────
    if face_area >= 150 * 150:
        size_score = 1.0
    elif face_area >= 80 * 80:
        size_score = 0.75
    elif face_area >= 40 * 40:
        size_score = 0.5
    elif face_area >= 20 * 20:
        size_score = 0.3
    else:
        size_score = 0.15

    # ── Factor 2: Sharpness (motion blur / focus) ──────────────────────
    gray = cv2.cvtColor(face_crop_bgr, cv2.COLOR_BGR2GRAY)
    laplacian_var = cv2.Laplacian(gray, cv2.CV_64F).var()

    if laplacian_var >= 500:
        sharp_score = 1.0
    elif laplacian_var >= 200:
        sharp_score = 0.8
    elif laplacian_var >= 80:
        sharp_score = 0.6
    elif laplacian_var >= 30:
        sharp_score = 0.4
    else:
        sharp_score = 0.2

    # ── Factor 3: Brightness ───────────────────────────────────────────
    brightness = np.mean(gray)

    if 80 <= brightness <= 200:
        bright_score = 1.0
    elif 50 <= brightness < 80:
        bright_score = 0.7
    elif 200 < brightness <= 230:
        bright_score = 0.8
    elif brightness < 50:
        bright_score = 0.4
    else:
        bright_score = 0.3

    # ── Combined quality score (weighted) ─────────────────────────────
    quality = (size_score * 0.5) + (sharp_score * 0.3) + (bright_score * 0.2)

    # ── Map quality to threshold ───────────────────────────────────────
    threshold = 0.22 + 0.20 * quality
    return float(np.clip(threshold, 0.22, 0.42))


class AIModelLoader(QThread):
    """Loads YOLO + ArcFace once at startup in a background thread."""
    done = pyqtSignal(bool)

    def __init__(self):
        super().__init__()

    def run(self):
        _, _, ok = _ensure_ai_models(det_size=(640, 640))
        global _adaface_model, _adaface_db
        # Load AdaFace embeddings from MongoDB (cached)
        if _adaface_model is not None:
            try:
                import numpy as _np_ada
                if DB_CONNECTED:
                    students = list(db["Students"].find(
                        {"AdaFaceEmbedding": {"$exists": True}},
                        {"StudentID": 1, "AdaFaceEmbedding": 1}
                    ))
                    for stu in students:
                        sid = str(stu.get("StudentID", ""))
                        ada_emb = stu.get("AdaFaceEmbedding", [])
                        if ada_emb and len(ada_emb) == 512:
                            arr = _np_ada.array(ada_emb, dtype=_np_ada.float32)
                            norm = _np_ada.linalg.norm(arr)
                            if norm > 0:
                                _adaface_db[sid] = arr / norm
                    print(f"[AIAS] AdaFace DB loaded from MongoDB: {len(_adaface_db)} students")
            except Exception as e:
                print(f"[AIAS] AdaFace DB load error: {e}")
        self.done.emit(ok)


class AnimatedButton(QPushButton):
    """QPushButton with smooth hover color transition."""

    def __init__(self, text="", parent=None):
        super().__init__(text, parent)
        self._anim_timer = QTimer(self)
        self._anim_timer.timeout.connect(self._tick)
        self._hover_progress = 0.0   # 0.0 = normal, 1.0 = fully hovered
        self._hovered = False
        self.setCursor(Qt.PointingHandCursor)

    def enterEvent(self, event):
        self._hovered = True
        if not self._anim_timer.isActive():
            self._anim_timer.start(16)
        super().enterEvent(event)

    def leaveEvent(self, event):
        self._hovered = False
        if not self._anim_timer.isActive():
            self._anim_timer.start(16)
        super().leaveEvent(event)

    def _tick(self):
        speed = 0.12
        if self._hovered:
            self._hover_progress = min(1.0, self._hover_progress + speed)
        else:
            self._hover_progress = max(0.0, self._hover_progress - speed)

        if (self._hovered and self._hover_progress >= 1.0) or \
           (not self._hovered and self._hover_progress <= 0.0):
            self._anim_timer.stop()

        self._apply_hover_style()

    def _apply_hover_style(self):
        pass  # Styling is handled via QSS :hover pseudo-class below


class AppWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("AIAS — AI Attendance System")
        import os
        from PyQt5.QtGui import QIcon
        _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aias_icon.ico")
        if os.path.exists(_icon_path):
            self.setWindowIcon(QIcon(_icon_path))

        screen = QApplication.primaryScreen().geometry()
        self.resize(960, 640)
        self.move(
            (screen.width() - self.width()) // 2,
            (screen.height() - self.height()) // 2
        )

        # Splash state
        self._showing_splash = True
        self._sp_opacity     = 0.0
        self._sp_scale       = 0.7
        self._sp_progress    = 0.0
        self._sp_scan_y      = 0.25
        self._sp_scan_dir    = 1
        self._sp_corner_op   = 1.0
        self._sp_corner_dir  = -1
        self._sp_ring_scale  = [1.0, 1.0, 1.0]
        self._sp_ring_dir    = [1, 1, 1]
        self._sp_dot_scale   = [1.0, 1.0]
        self._sp_dot_dir     = [1, -1]
        self._sp_done        = False
        self._sp_fade_out    = 1.0

        # Particle system
        import random, math
        self._particles = []
        for _ in range(100):
            angle = random.uniform(0, math.pi * 2)
            speed = random.uniform(0.0003, 0.0012)
            self._particles.append({
                "px":          random.uniform(0.0, 1.0),
                "py":          random.uniform(0.0, 1.0),
                "vx":          math.cos(angle) * speed,
                "vy":          math.sin(angle) * speed,
                "size":        random.uniform(1.5, 3.5),
                "opacity":     random.uniform(0.2, 0.75),
                "pulse":       random.uniform(0, 6.28),
                "pulse_speed": random.uniform(0.015, 0.045),
            })

        # Login page
        self._login_page = LoginWindow(embedded=True)
        self._login_page.login_success.connect(self._on_login_success)

        # Single-window navigation stack
        self._stack = QStackedWidget(self)
        self._stack.setStyleSheet("background:transparent;")
        self._stack.addWidget(self._login_page)
        self._stack.hide()

        self._overlay_opacity = 1.0
        self._overlay_active  = False

        # Animation timer
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._tick)
        self._timer.start(33)

        self.setStyleSheet("background:#0a0f0a;")
        self.setAttribute(Qt.WA_OpaquePaintEvent, True)

    def resizeEvent(self, event):
        self._stack.setGeometry(0, 0, self.width(), self.height())

    def _tick(self):
        if not self._sp_done:
            # Animate splash
            if self._sp_opacity < 1.0:
                self._sp_opacity = min(1.0, self._sp_opacity + 0.04)
            if self._sp_scale < 1.0:
                self._sp_scale = min(1.0, self._sp_scale + 0.02)
            self._sp_progress = min(100.0, self._sp_progress + 0.9)

            self._sp_scan_y += 0.01 * self._sp_scan_dir
            if self._sp_scan_y >= 0.72: self._sp_scan_dir = -1
            elif self._sp_scan_y <= 0.22: self._sp_scan_dir = 1

            self._sp_corner_op += 0.05 * self._sp_corner_dir
            if self._sp_corner_op <= 0.2: self._sp_corner_dir = 1
            elif self._sp_corner_op >= 1.0: self._sp_corner_dir = -1

            speeds = [0.008, 0.006, 0.004]
            for i in range(3):
                self._sp_ring_scale[i] += speeds[i] * self._sp_ring_dir[i]
                if self._sp_ring_scale[i] >= 1.06: self._sp_ring_dir[i] = -1
                elif self._sp_ring_scale[i] <= 0.96: self._sp_ring_dir[i] = 1

            for i in range(2):
                self._sp_dot_scale[i] += 0.06 * self._sp_dot_dir[i]
                if self._sp_dot_scale[i] >= 1.5: self._sp_dot_dir[i] = -1
                elif self._sp_dot_scale[i] <= 0.6: self._sp_dot_dir[i] = 1

            if self._sp_progress >= 100.0:
                QTimer.singleShot(2500, self._start_fade_out)
                self._sp_done = True

            # Update particles
            import math
            for pt in self._particles:
                pt["px"]    += pt["vx"]
                pt["py"]    += pt["vy"]
                pt["pulse"] += pt["pulse_speed"]
                pt["opacity"] = 0.3 + 0.5 * abs(math.sin(pt["pulse"]))
                # Wrap around using percentages
                if pt["px"] < 0.0: pt["px"] += 1.0
                if pt["px"] > 1.0: pt["px"] -= 1.0
                if pt["py"] < 0.0: pt["py"] += 1.0
                if pt["py"] > 1.0: pt["py"] -= 1.0

            self.update()
        else:
            # Fading out
            if self._sp_fade_out > 0.0:
                import math
                self._sp_fade_out = max(0.0, self._sp_fade_out - 0.007)
                self.update()
            else:
                if not hasattr(self, '_splash_finished'):
                    self._splash_finished = True
                    self._finish_splash()
                # Keep particles running in background
                import math
                for pt in self._particles:
                    pt["px"]    += pt["vx"]
                    pt["py"]    += pt["vy"]
                    pt["pulse"] += pt["pulse_speed"]
                    pt["opacity"] = 0.3 + 0.5 * abs(math.sin(pt["pulse"]))
                    if pt["px"] < 0.0: pt["px"] += 1.0
                    if pt["px"] > 1.0: pt["px"] -= 1.0
                    if pt["py"] < 0.0: pt["py"] += 1.0
                    if pt["py"] > 1.0: pt["py"] -= 1.0
                self.update()

    def _start_fade_out(self):
        self._showing_splash = False

    def _finish_splash(self):
        self.setStyleSheet("background:#0a0f0a;")
        self._stack.setGeometry(0, 0, self.width(), self.height())
        self._stack.setCurrentWidget(self._login_page)
        self._stack.show()
        self._stack.raise_()
        # Start overlay fade
        self._overlay_opacity = 1.0
        self._overlay_active  = True
        self._fade_in_step    = 0.0
        self._fade_timer = QTimer(self)
        self._fade_timer.timeout.connect(self._fade_in_login)
        self._fade_timer.start(16)

    def _on_login_success(self, role, username, fullname):
        try:
            if role == "instructor":
                instr = db["Instructors"].find_one({"Username": username})
                assigned = instr.get("AssignedSections", []) if instr else []
                new_page = MainWindow(username, fullname, assigned, embedded=True)
            else:
                new_page = AdminPanelWindow(embedded=True)
        except Exception as e:
            print(f"[AIAS] login error: {e}")
            return

        # Remove old pages except login
        while self._stack.count() > 1:
            widget = self._stack.widget(1)
            self._stack.removeWidget(widget)
            widget.deleteLater()

        self._stack.addWidget(new_page)
        self._stack.setCurrentWidget(new_page)
        self.setStyleSheet(f"background:{BG};")
        new_page.update()

    def _fade_in_login(self):
        self._fade_in_step = min(1.0, self._fade_in_step + 0.008)
        t = self._fade_in_step
        # Cubic ease-in-out
        if t < 0.5:
            eased = 4 * t * t * t
        else:
            eased = 1 - ((-2 * t + 2) ** 3) / 2
        # Overlay goes from 1.0 (black) to 0.0 (transparent)
        self._overlay_opacity = 1.0 - eased
        self.update()
        if self._fade_in_step >= 1.0:
            self._fade_timer.stop()
            self._overlay_active  = False
            self._overlay_opacity = 0.0
            self.update()

    def paintEvent(self, event):
        # Draw black overlay during login fade-in
        if self._overlay_active:
            from PyQt5.QtGui import QPainter, QColor
            p = QPainter(self)
            p.setRenderHint(QPainter.Antialiasing)
            p.setOpacity(self._overlay_opacity)
            p.setBrush(QColor("#0a0f0a"))
            p.setPen(Qt.NoPen)
            p.drawRect(0, 0, self.width(), self.height())
            p.end()
            return

        if not self._showing_splash and self._sp_fade_out <= 0.0:
            # Still draw particles behind login
            if self._overlay_active:
                return
            from PyQt5.QtGui import QPainter, QColor
            import math
            p = QPainter(self)
            p.setRenderHint(QPainter.Antialiasing)
            p.setBrush(QColor("#0a0f0a"))
            p.setPen(Qt.NoPen)
            p.drawRect(0, 0, self.width(), self.height())
            W, H = self.width(), self.height()
            for pt in self._particles:
                px = pt["px"] * W
                py = pt["py"] * H
                size = pt["size"]
                p.setOpacity(pt["opacity"] * 0.4)
                p.setBrush(QColor("#1B5E35"))
                p.setPen(Qt.NoPen)
                p.drawEllipse(int(px-size*2), int(py-size*2), int(size*4), int(size*4))
                p.setOpacity(pt["opacity"] * 0.6)
                p.setBrush(QColor("#4ecf8e"))
                p.drawEllipse(int(px-size*0.5), int(py-size*0.5), int(size), int(size))
            p.end()
            return

        from PyQt5.QtGui import QPainter, QColor, QPen, QBrush, QFont, QLinearGradient, QRadialGradient
        from PyQt5.QtCore import QRectF, QPointF

        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)

        fade = self._sp_fade_out if not self._showing_splash else 1.0
        p.setOpacity(self._sp_opacity * fade)

        W, H = self.width(), self.height()
        cx, cy = W // 2, H // 2 - 30

        # Background
        p.setBrush(QColor("#0a0f0a"))
        p.setPen(Qt.NoPen)
        p.drawRect(0, 0, W, H)

        # Draw particles
        import math
        for pt in self._particles:
            pt_opacity = pt["opacity"] * self._sp_opacity * fade
            if pt_opacity <= 0:
                continue
            # Convert percentage to actual pixels
            px = pt["px"] * W
            py = pt["py"] * H
            size = pt["size"]
            # Outer glow
            p.setOpacity(pt_opacity * 0.3)
            p.setBrush(QColor("#1B5E35"))
            p.setPen(Qt.NoPen)
            p.drawEllipse(
                int(px - size * 2), int(py - size * 2),
                int(size * 4), int(size * 4)
            )
            # Inner bright dot
            p.setOpacity(pt_opacity)
            p.setBrush(QColor("#4ecf8e"))
            p.drawEllipse(
                int(px - size * 0.5), int(py - size * 0.5),
                int(size), int(size)
            )
        p.setOpacity(self._sp_opacity * fade)

        # Rings
        ring_sizes = [340, 440, 540]
        ring_ops   = [0.18, 0.10, 0.06]
        for i, (sz, op) in enumerate(zip(ring_sizes, ring_ops)):
            scaled = sz * self._sp_ring_scale[i]
            pen = QPen(QColor("#1B5E35"))
            pen.setWidthF(1.0)
            p.setPen(pen)
            p.setBrush(Qt.NoBrush)
            p.setOpacity(self._sp_opacity * fade * op)
            p.drawEllipse(QRectF(cx - scaled/2, cy - scaled/2, scaled, scaled))

        p.setOpacity(self._sp_opacity * fade)

        # Logo circle
        zoom = 1.0 + (1.0 - self._sp_fade_out) * 0.30
        logo_r = 100 * self._sp_scale * zoom
        grad = QRadialGradient(cx, cy - 20, logo_r)
        grad.setColorAt(0, QColor("#1B5E35"))
        grad.setColorAt(1, QColor("#0d3520"))
        p.setBrush(QBrush(grad))
        pen2 = QPen(QColor("#2d8a57"))
        pen2.setWidthF(2.0)
        p.setPen(pen2)
        p.drawEllipse(QRectF(cx - logo_r, cy - logo_r, logo_r*2, logo_r*2))

        # Dashed ring
        pen3 = QPen(QColor(180, 220, 200, 60))
        pen3.setWidthF(1.2)
        pen3.setStyle(Qt.DashLine)
        p.setPen(pen3)
        p.setBrush(Qt.NoBrush)
        outer_r = logo_r + 8
        p.drawEllipse(QRectF(cx - outer_r, cy - outer_r, outer_r*2, outer_r*2))

        # User head
        head_r  = 28 * self._sp_scale
        head_cx = cx
        head_cy = cy - 22 * self._sp_scale
        p.setPen(Qt.NoPen)
        p.setBrush(QColor(180, 220, 200, 195))
        p.drawEllipse(QRectF(head_cx - head_r, head_cy - head_r, head_r*2, head_r*2))

        # Scan line
        scan_y = head_cy - head_r + (head_r*2 * self._sp_scan_y)
        p.setClipRect(QRectF(head_cx - head_r, head_cy - head_r, head_r*2, head_r*2).toRect())
        sp = QPen(QColor(255, 255, 255, 210))
        sp.setWidthF(1.5)
        p.setPen(sp)
        p.drawLine(QPointF(head_cx - head_r + 2, scan_y), QPointF(head_cx + head_r - 2, scan_y))
        p.setClipping(False)

        # Dots
        for i, (dx, dy) in enumerate([(head_cx - head_r*0.3, head_cy + head_r*0.05),
                                       (head_cx + head_r*0.3, head_cy + head_r*0.05)]):
            ds = 3.5 * self._sp_dot_scale[i] * self._sp_scale
            p.setPen(Qt.NoPen)
            p.setBrush(QColor(13, 53, 32, 220))
            p.drawEllipse(QRectF(dx-ds, dy-ds, ds*2, ds*2))

        # Body
        body_w = 74 * self._sp_scale
        body_h = 30 * self._sp_scale
        body_y = head_cy + head_r + 3 * self._sp_scale
        p.setBrush(QColor(180, 220, 200, 165))
        p.setPen(Qt.NoPen)
        p.drawChord(QRectF(cx - body_w/2, body_y, body_w, body_h*2), 0, 180*16)

        # Corners
        cc = QColor(255, 255, 255, int(255 * self._sp_corner_op))
        cp = QPen(cc)
        cp.setWidthF(2.5)
        cp.setCapStyle(Qt.RoundCap)
        p.setPen(cp)
        p.setBrush(Qt.NoBrush)
        cw  = 16 * self._sp_scale
        fx1 = head_cx - head_r - 10 * self._sp_scale
        fx2 = head_cx + head_r + 10 * self._sp_scale
        fy1 = head_cy - head_r - 10 * self._sp_scale
        fy2 = body_y + body_h + 4 * self._sp_scale
        p.drawLine(QPointF(fx1, fy1+cw), QPointF(fx1, fy1)); p.drawLine(QPointF(fx1, fy1), QPointF(fx1+cw, fy1))
        p.drawLine(QPointF(fx2-cw, fy1), QPointF(fx2, fy1)); p.drawLine(QPointF(fx2, fy1), QPointF(fx2, fy1+cw))
        p.drawLine(QPointF(fx1, fy2-cw), QPointF(fx1, fy2)); p.drawLine(QPointF(fx1, fy2), QPointF(fx1+cw, fy2))
        p.drawLine(QPointF(fx2-cw, fy2), QPointF(fx2, fy2)); p.drawLine(QPointF(fx2, fy2), QPointF(fx2, fy2-cw))

        # AIAS text
        text_y = body_y + body_h + 18 * self._sp_scale
        f1 = QFont("Segoe UI", int(18 * self._sp_scale), QFont.Bold)
        p.setFont(f1)
        p.setPen(QColor("#ffffff"))
        p.drawText(QRectF(cx-80, text_y, 160, 26), Qt.AlignCenter, "AIAS")
        f2 = QFont("Segoe UI", int(6 * self._sp_scale))
        f2.setLetterSpacing(QFont.AbsoluteSpacing, 2.5)
        p.setFont(f2)
        p.setPen(QColor(200, 230, 210, 180))
        p.drawText(QRectF(cx-90, text_y+24, 180, 18), Qt.AlignCenter, "AI ATTENDANCE SYSTEM")

        # Title
        title_y = cy + logo_r + 28
        p.setOpacity(self._sp_opacity * (fade ** 3) * min(1.0, self._sp_progress / 30))
        f3 = QFont("Segoe UI", 22, QFont.Bold)
        p.setFont(f3)
        p.setPen(QColor("#ffffff"))
        p.drawText(QRectF(0, title_y, W, 36), Qt.AlignCenter, "AI Attendance System")
        f4 = QFont("Segoe UI", 9)
        f4.setLetterSpacing(QFont.AbsoluteSpacing, 3.5)
        p.setFont(f4)
        p.setPen(QColor(180, 220, 200, 140))
        p.drawText(QRectF(0, title_y+40, W, 20), Qt.AlignCenter, "POWERED BY FACE RECOGNITION")

        # Loading bar
        p.setOpacity(self._sp_opacity * (fade ** 3) * min(1.0, self._sp_progress / 20))
        bar_w = 240; bar_h = 3
        bar_x = (W - bar_w) / 2
        bar_y = title_y + 76
        p.setBrush(QColor(255, 255, 255, 18))
        p.setPen(Qt.NoPen)
        p.drawRoundedRect(QRectF(bar_x, bar_y, bar_w, bar_h), 2, 2)
        fill_w = bar_w * (self._sp_progress / 100)
        lg = QLinearGradient(bar_x, 0, bar_x + bar_w, 0)
        lg.setColorAt(0, QColor("#1B5E35"))
        lg.setColorAt(1, QColor("#4ecf8e"))
        p.setBrush(QBrush(lg))
        p.drawRoundedRect(QRectF(bar_x, bar_y, fill_w, bar_h), 2, 2)
        f5 = QFont("Segoe UI", 8)
        f5.setLetterSpacing(QFont.AbsoluteSpacing, 2)
        p.setFont(f5)
        p.setPen(QColor(180, 220, 200, 90))
        p.drawText(QRectF(0, bar_y+10, W, 16), Qt.AlignCenter, "INITIALIZING...")

        p.end()


class LoginWindow(QWidget):
    login_success = pyqtSignal(str, str, str)  # role, username, fullname

    def __init__(self, embedded=False):
        super().__init__()
        self._embedded = embedded
        if not embedded:
            self.setWindowTitle("AIAS — Login")
            self.setMinimumSize(600, 680)
            self.setStyleSheet("background:transparent;")
        self._role = "instructor"
        self._build()

    def _build(self):
        outer = QVBoxLayout(self)
        outer.setAlignment(Qt.AlignCenter)
        outer.setContentsMargins(40, 40, 40, 40)

        # Card
        card = QFrame()
        card.setFixedWidth(400)
        card.setMaximumHeight(560)
        card.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Preferred)
        card.setStyleSheet(
            "QFrame { background:#ffffff; border-radius:16px; }"
        )
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(30)
        shadow.setOffset(0, 4)
        shadow.setColor(QColor(0, 0, 0, 40))
        card.setGraphicsEffect(shadow)

        lay = QVBoxLayout(card)
        lay.setContentsMargins(32, 32, 32, 32)
        lay.setSpacing(0)

        # 1. Icon + Name row
        header_row = QHBoxLayout()
        header_row.setSpacing(12)

        icon_lbl = QLabel("AI")
        icon_lbl.setFixedSize(48, 48)
        icon_lbl.setAlignment(Qt.AlignCenter)
        icon_lbl.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:10px;"
            "font-size:14px; font-weight:700; border:none;"
        )

        name_col = QVBoxLayout()
        name_col.setSpacing(2)
        app_name = QLabel("AIAS")
        app_name.setStyleSheet(
            "font-size:22px; font-weight:700; color:#1A1A2E;"
            "background:transparent; border:none;"
        )
        app_sub = QLabel("AI Attendance System")
        app_sub.setStyleSheet(
            "font-size:12px; color:#6B7280; background:transparent; border:none;"
        )
        name_col.addWidget(app_name)
        name_col.addWidget(app_sub)

        header_row.addWidget(icon_lbl)
        header_row.addLayout(name_col)
        header_row.addStretch()
        lay.addLayout(header_row)
        lay.addSpacing(24)

        # 2. Badge
        badge_lbl = QLabel("Qassim University")
        badge_lbl.setFixedHeight(26)
        badge_lbl.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        badge_lbl.setStyleSheet(
            f"background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:12px;"
            "padding:4px 12px; font-size:11px; font-weight:600; border:none;"
        )
        lay.addWidget(badge_lbl)
        lay.addSpacing(20)

        # 3. Welcome text
        welcome_lbl = QLabel("Welcome back")
        welcome_lbl.setStyleSheet(
            "font-size:26px; font-weight:700; color:#1A1A2E;"
            "background:transparent; border:none;"
        )
        signin_sub = QLabel("Sign in to access the dashboard")
        signin_sub.setStyleSheet(
            "font-size:13px; color:#6B7280; background:transparent; border:none;"
        )
        lay.addWidget(welcome_lbl)
        lay.addWidget(signin_sub)
        lay.addSpacing(24)

        # 4. Role toggle
        toggle_frame = QFrame()
        toggle_frame.setFixedHeight(44)
        toggle_frame.setStyleSheet(
            f"QFrame {{ background:{BG}; border-radius:10px; border:none; }}"
        )
        toggle_lay = QHBoxLayout(toggle_frame)
        toggle_lay.setContentsMargins(4, 4, 4, 4)
        toggle_lay.setSpacing(4)
        self._btn_instructor = QPushButton("Instructor")
        self._btn_admin = QPushButton("Admin")
        for btn in (self._btn_instructor, self._btn_admin):
            btn.setCursor(Qt.PointingHandCursor)
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._btn_instructor.clicked.connect(lambda: self._set_role("instructor"))
        self._btn_admin.clicked.connect(lambda: self._set_role("admin"))
        toggle_lay.addWidget(self._btn_instructor)
        toggle_lay.addWidget(self._btn_admin)
        lay.addWidget(toggle_frame)
        lay.addSpacing(20)
        self._apply_toggle_style()

        # 5. Username
        u_lbl = QLabel("Username")
        u_lbl.setStyleSheet(
            "font-size:13px; font-weight:600; color:#1A1A2E;"
            "background:transparent; border:none;"
        )
        self.username_field = QLineEdit()
        self.username_field.setPlaceholderText("Enter your username")
        self.username_field.setFixedHeight(48)
        self.username_field.setStyleSheet(
            "QLineEdit { background:#ffffff; color:#1A1A2E; border:1.5px solid #E5E7EB; "
            "border-radius:10px; padding:12px 16px; font-size:14px; }"
            "QLineEdit:focus { border:1.5px solid #1B5E35; }"
        )
        self.username_field.setPlaceholderText("Enter your username")
        lay.addWidget(u_lbl)
        lay.addSpacing(6)
        self.username_field.returnPressed.connect(self._login)
        lay.addWidget(self.username_field)
        lay.addSpacing(16)

        # 6. Password
        p_lbl = QLabel("Password")
        p_lbl.setStyleSheet(
            "font-size:13px; font-weight:600; color:#1A1A2E;"
            "background:transparent; border:none;"
        )
        self.password_field = QLineEdit()
        self.password_field.setPlaceholderText("Enter your password")
        self.password_field.setEchoMode(QLineEdit.Password)
        self.password_field.setFixedHeight(48)
        self.password_field.setStyleSheet(
            "QLineEdit { background:#ffffff; color:#1A1A2E; border:1.5px solid #E5E7EB; "
            "border-radius:10px; padding:12px 16px; font-size:14px; }"
            "QLineEdit:focus { border:1.5px solid #1B5E35; }"
        )
        self.password_field.setPlaceholderText("Enter your password")
        self.password_field.returnPressed.connect(self._login)
        lay.addWidget(p_lbl)
        lay.addSpacing(6)
        lay.addWidget(self.password_field)
        lay.addSpacing(24)

        # 7. Sign In button
        sign_btn = QPushButton("Sign In")
        sign_btn.setFixedHeight(48)
        sign_btn.setStyleSheet(
            f"QPushButton {{ background:{PRIMARY}; color:white; border-radius:10px; "
            f"height:48px; font-size:15px; font-weight:700; }}"
            f"QPushButton:hover {{ background:#145a30; color:white; }}"
            f"QPushButton:pressed {{ background:#0d3d20; color:white; }}"
        )
        sign_btn.setCursor(Qt.PointingHandCursor)
        sign_btn.clicked.connect(self._login)
        lay.addWidget(sign_btn)
        lay.addSpacing(12)

        # 8. Error label
        self.err_label = QLabel("")
        self.err_label.setAlignment(Qt.AlignCenter)
        self.err_label.setStyleSheet(
            "color:#DC2626; font-size:12px; background:transparent; border:none;"
        )
        self.err_label.setFixedHeight(18)
        lay.addWidget(self.err_label)

        outer.addWidget(card, alignment=Qt.AlignCenter)

    def _set_role(self, role):
        self._role = role
        self._apply_toggle_style()

    def _apply_toggle_style(self):
        active_ss = (
            f"background:{WHITE}; color:{PRIMARY}; border:1.5px solid {PRIMARY};"
            "border-radius:8px; font-size:13px; font-weight:600;"
        )
        inactive_ss = (
            f"background:transparent; color:{TEXT_GRAY}; border:none;"
            "font-size:13px; font-weight:500;"
        )
        if self._role == "instructor":
            self._btn_instructor.setStyleSheet(active_ss)
            self._btn_admin.setStyleSheet(inactive_ss)
        else:
            self._btn_instructor.setStyleSheet(inactive_ss)
            self._btn_admin.setStyleSheet(active_ss)

    def _login(self):
        u = self.username_field.text().strip()
        p = self.password_field.text()
        if not DB_CONNECTED:
            self.err_label.setText("Database not connected.")
            return
        try:
            doc = db["Instructors"].find_one({"Username": u})
            if not doc:
                self.err_label.setText("Invalid username or password.")
                return

            stored = doc.get("Password", "")
            if isinstance(stored, str):
                # Migrate plain-text password to bcrypt on first login (Bug 2)
                if stored == p:
                    hashed = bcrypt.hashpw(p.encode("utf-8"), bcrypt.gensalt())
                    db["Instructors"].update_one(
                        {"Username": u}, {"$set": {"Password": hashed}}
                    )
                    pw_ok = True
                else:
                    pw_ok = False
            else:
                pw_ok = bcrypt.checkpw(p.encode("utf-8"), stored)

            if not pw_ok:
                self.err_label.setText("Invalid username or password.")
                return

            role = doc.get("Role", "instructor")

            # Enforce toggle selection matches DB role
            selected = self._role  # "admin" or "instructor"
            if selected == "admin" and role != "admin":
                self.err_label.setText("This account does not have Admin privileges.")
                return
            if selected == "instructor" and role == "admin":
                self.err_label.setText("Please select 'Admin' to log in as Administrator.")
                return

            if role == "admin":
                self.login_success.emit("admin", "", "")
            else:
                fullname = doc.get("FullName", u)
                self.login_success.emit("instructor", u, fullname)
        except Exception as e:
            print(f"Exception: {e}")
            self.err_label.setText("Database error. Please try again.")

    def keyPressEvent(self, event):
        from PyQt5.QtCore import Qt
        if event.key() == Qt.Key_Escape:
            pass
        else:
            super().keyPressEvent(event)


class CourseCard(QFrame):
    def __init__(self, code, name, schedule, active=False, n_students=0):
        super().__init__()
        self.setFixedHeight(86)
        self.setCursor(Qt.PointingHandCursor)
        self._active = active
        self._apply_style()

        lay = QVBoxLayout(self)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(2)

        top = QLabel(f"<b>{code}</b> — {name}")
        top.setStyleSheet(
            f"font-size:12px; color:{TEXT_DARK}; background:transparent; border:none;"
        )
        bot = QLabel(schedule)
        bot.setStyleSheet(
            f"font-size:11px; color:{TEXT_GRAY}; background:transparent; border:none;"
        )
        count_lbl = QLabel(f"\U0001f465 {n_students} students")
        count_lbl.setStyleSheet(
            f"font-size:11px; color:{TEXT_GRAY}; background:transparent; border:none;"
        )
        lay.addWidget(top)
        lay.addWidget(bot)
        lay.addWidget(count_lbl)

    def _apply_style(self):
        if self._active:
            self.setStyleSheet(
                f"QFrame {{ background:{PRIMARY_LIGHT}; border:1.5px solid {PRIMARY};"
                f"border-radius:8px; }}"
            )
        else:
            self.setStyleSheet(
                f"QFrame {{ background:{WHITE}; border:1px solid {BORDER};"
                f"border-radius:8px; }}"
            )

    def set_active(self, val):
        self._active = val
        self._apply_style()


class MainWindow(QWidget):
    def __init__(self, username, fullname, sections, embedded=False):
        super().__init__()
        self._embedded = embedded
        self.instructor_data = {
            "username": username,
            "fullname": fullname,
            "assigned_sections": sections,
        }
        self._courses_data = []
        if not embedded:
            self.setWindowTitle("AIAS — Course Selection")
            import os
            from PyQt5.QtGui import QIcon
            _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aias_icon.ico")
            if os.path.exists(_icon_path):
                self.setWindowIcon(QIcon(_icon_path))
            self.setMinimumSize(1100, 700)
            self.resize(1100, 700)
        self.setStyleSheet(f"background:{BG};")
        self._active_idx = 0
        self._build()

    def _build(self):
        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        assigned = self.instructor_data.get("assigned_sections", [])
        fullname = self.instructor_data.get("fullname", "Instructor")

        import ast
        if isinstance(assigned, str):
            try:
                assigned = ast.literal_eval(assigned)
            except Exception:
                assigned = [assigned]

        if DB_CONNECTED:
            try:
                self._courses_data = list(
                    db["Courses"].find({"SectionID": {"$in": assigned}})
                ) if assigned else []
            except Exception:
                self._courses_data = []

        parts = fullname.split()
        initials = (parts[0][0] + parts[-1][0]).upper() if len(parts) >= 2 else fullname[:2].upper()
        sidebar, s_lay = make_sidebar_base(initials, fullname, "Instructor")

        dash_btn = QPushButton("📊  Dashboard")
        dash_btn.setStyleSheet(
            f"QPushButton {{ background:{PRIMARY}; color:#ffffff; border-radius:6px; "
            f"padding:8px 12px; font-size:12px; font-weight:600; }}"
            f"QPushButton:hover {{ background:#145a30; color:#ffffff; }}"
        )
        dash_btn.setCursor(Qt.PointingHandCursor)
        dash_btn.clicked.connect(self._show_dashboard)
        s_lay.addWidget(dash_btn)
        s_lay.addSpacing(8)

        sec_lbl = QLabel("MY COURSES")
        sec_lbl.setStyleSheet(
            f"font-size:10px; font-weight:700; color:{TEXT_GRAY}; letter-spacing:1px;"
            f"background:transparent; border:none;"
        )
        s_lay.addWidget(sec_lbl)
        s_lay.addSpacing(4)

        self._course_cards = []
        for i, course in enumerate(self._courses_data):
            code = str(course.get("SectionID", ""))
            name = course.get("CourseTitle", "")
            sched_raw = course.get("Schedule", "")
            if isinstance(sched_raw, dict):
                sched = f"{sched_raw.get(chr(68) + 'ay', '')} {sched_raw.get('StartTime', '')}–{sched_raw.get('EndTime', '')}"
            else:
                sched = str(sched_raw)
            section_id = course.get("SectionID", "")
            try:
                n_students = db["Students"].count_documents({"Sections": section_id}) if section_id and DB_CONNECTED else 0
            except Exception:
                n_students = 0
            card = CourseCard(code, name, sched, active=(i == 0), n_students=n_students)
            card.mousePressEvent = lambda e, idx=i: self._select_course(idx)
            self._course_cards.append(card)
            s_lay.addWidget(card)

        s_lay.addStretch()

        history_btn = QPushButton("Session History")
        history_btn.setToolTip("View all past attendance sessions")
        history_btn.setStyleSheet(
            f"QPushButton {{ background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px; "
            f"padding:8px 12px; font-size:12px; font-weight:600; }}"
            f"QPushButton:hover {{ background:{PRIMARY}; color:#ffffff; }}"
            f"QPushButton:pressed {{ background:{PRIMARY}; color:#ffffff; padding:9px 11px 7px 13px; }}"
        )
        history_btn.setCursor(Qt.PointingHandCursor)
        history_btn.clicked.connect(self._open_history)
        s_lay.addWidget(history_btn)
        s_lay.addSpacing(4)

        analytics_btn = QPushButton("Analytics")
        analytics_btn.setToolTip("View attendance statistics and at-risk students")
        analytics_btn.setStyleSheet(
            f"QPushButton {{ background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px; "
            f"padding:8px 12px; font-size:12px; font-weight:600; }}"
            f"QPushButton:hover {{ background:{PRIMARY}; color:#ffffff; }}"
            f"QPushButton:pressed {{ background:{PRIMARY}; color:#ffffff; padding:9px 11px 7px 13px; }}"
        )
        analytics_btn.setCursor(Qt.PointingHandCursor)
        analytics_btn.clicked.connect(self._open_analytics)
        s_lay.addWidget(analytics_btn)
        s_lay.addSpacing(4)

        email_btn = QPushButton("Email Settings")
        email_btn.setToolTip("Configure email settings for sending reports")
        email_btn.setStyleSheet(
            f"QPushButton {{ background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px; "
            f"padding:8px 12px; font-size:12px; font-weight:600; }}"
            f"QPushButton:hover {{ background:{PRIMARY}; color:#ffffff; }}"
            f"QPushButton:pressed {{ background:{PRIMARY}; color:#ffffff; padding:9px 11px 7px 13px; }}"
        )
        email_btn.setCursor(Qt.PointingHandCursor)
        email_btn.clicked.connect(self._open_email_settings)
        s_lay.addWidget(email_btn)
        s_lay.addSpacing(6)

        change_pass_btn = QPushButton("🔒  Change Password")
        change_pass_btn.setToolTip("Change your login password")
        change_pass_btn.setStyleSheet(
            f"QPushButton {{ background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px; "
            f"padding:8px 12px; font-size:12px; font-weight:600; }}"
            f"QPushButton:hover {{ background:{PRIMARY}; color:#ffffff; }}"
            f"QPushButton:pressed {{ background:{PRIMARY}; color:#ffffff; padding:9px 11px 7px 13px; }}"
        )
        change_pass_btn.setCursor(Qt.PointingHandCursor)
        change_pass_btn.clicked.connect(self._change_password)
        s_lay.addWidget(change_pass_btn)
        s_lay.addSpacing(4)

        dark_btn = QPushButton("🌙  Dark Mode" if not IS_DARK_MODE else "☀  Light Mode")
        dark_btn.setStyleSheet(
            f"background:{'#2A3D35' if IS_DARK_MODE else PRIMARY_LIGHT}; "
            f"color:{'#4ecf8e' if IS_DARK_MODE else PRIMARY}; "
            "border-radius:6px; padding:8px 12px; font-size:12px; font-weight:600;"
        )
        dark_btn.setCursor(Qt.PointingHandCursor)
        dark_btn.clicked.connect(lambda: self._toggle_dark_mode(dark_btn))
        s_lay.addWidget(dark_btn)
        s_lay.addSpacing(4)

        logout_btn = QPushButton("Logout")
        logout_btn.setToolTip("Sign out and return to login screen")
        logout_btn.setStyleSheet(
            "QPushButton { background:#DC2626; color:white; border-radius:6px; "
            "padding:8px 12px; font-size:12px; font-weight:700; }"
            "QPushButton:hover { background:#B91C1C; color:white; }"
            "QPushButton:pressed { background:#991B1B; color:white; }"
        )
        logout_btn.setCursor(Qt.PointingHandCursor)
        logout_btn.clicked.connect(self._logout)
        s_lay.addWidget(logout_btn)

        main_w = QWidget()
        main_w.setStyleSheet(f"background:{BG};")
        m_lay = QVBoxLayout(main_w)
        m_lay.setContentsMargins(32, 28, 32, 24)
        m_lay.setSpacing(18)

        self._title_lbl = QLabel("")
        self._title_lbl.setStyleSheet(
            f"font-size:20px; font-weight:800; color:{TEXT_DARK};"
        )
        self._sub_lbl = QLabel("")
        self._sub_lbl.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY};")
        m_lay.addWidget(self._title_lbl)
        m_lay.addWidget(self._sub_lbl)

        stats_row = QHBoxLayout()
        stats_row.setSpacing(16)
        enrolled_card = QFrame()
        enrolled_card.setStyleSheet(
            f"QFrame {{ background:{WHITE}; border:1px solid {BORDER}; border-radius:8px; }}"
        )
        enrolled_card.setFixedHeight(82)
        enrolled_card.setMinimumWidth(130)
        ec_lay = QVBoxLayout(enrolled_card)
        ec_lay.setContentsMargins(16, 12, 16, 12)
        ec_lay.setSpacing(2)
        self._enrolled_val = QLabel("0")
        self._enrolled_val.setStyleSheet(
            f"font-size:26px; font-weight:800; color:{PRIMARY}; border:none; background:transparent;"
        )
        ec_lbl = QLabel("Enrolled")
        ec_lbl.setStyleSheet(
            f"font-size:11px; color:{TEXT_GRAY}; border:none; background:transparent;"
        )
        ec_lay.addWidget(self._enrolled_val)
        ec_lay.addWidget(ec_lbl)
        stats_row.addWidget(enrolled_card)

        def _make_updatable_card(title, initial):
            f = QFrame()
            f.setStyleSheet(f"QFrame {{ background:{WHITE}; border:1px solid {BORDER}; border-radius:8px; }}")
            f.setFixedHeight(82)
            f.setMinimumWidth(130)
            fl = QVBoxLayout(f)
            fl.setContentsMargins(16, 12, 16, 12)
            fl.setSpacing(2)
            vl = QLabel(str(initial))
            vl.setStyleSheet(f"font-size:26px; font-weight:800; color:{TEXT_DARK}; border:none; background:transparent;")
            tl = QLabel(title)
            tl.setStyleSheet(f"font-size:11px; color:{TEXT_GRAY}; border:none; background:transparent;")
            fl.addWidget(vl)
            fl.addWidget(tl)
            return f, vl

        sessions_card, self._sessions_held_lbl = _make_updatable_card("Sessions held", 0)
        last_card,     self._last_session_lbl  = _make_updatable_card("Last session",  "—")
        stats_row.addWidget(sessions_card)
        stats_row.addWidget(last_card)
        stats_row.addStretch()
        m_lay.addLayout(stats_row)

        self._tbl = make_table(
            ["Student ID", "Full Name", "Status"],
            [],
            col_widths={0: 140, 2: 140},
            stretch_col=1,
        )
        m_lay.addWidget(self._tbl)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        start_btn = QPushButton("▶  Start session")
        start_btn.setToolTip("Start a live face-recognition attendance session")
        start_btn.setFixedHeight(40)
        start_btn.setStyleSheet(
            f"QPushButton {{ background:{PRIMARY}; color:white; border-radius:8px; "
            f"padding:12px 24px; font-size:14px; font-weight:700; }}"
            f"QPushButton:hover {{ background:#145a30; color:white; }}"
            f"QPushButton:pressed {{ background:#0d3d20; color:white; }}"
        )
        start_btn.setCursor(Qt.PointingHandCursor)
        start_btn.clicked.connect(self._start_session)
        btn_row.addWidget(start_btn)
        m_lay.addLayout(btn_row)

        if not self._courses_data:
            for w in [self._tbl, start_btn]:
                w.setVisible(False)
            no_course_lbl = QLabel("No courses assigned to your account.\nPlease contact the admin.")
            no_course_lbl.setAlignment(Qt.AlignCenter)
            no_course_lbl.setStyleSheet(
                f"font-size:15px; color:{TEXT_GRAY}; background:transparent;"
            )
            m_lay.insertWidget(2, no_course_lbl)

        root.addWidget(sidebar)
        root.addWidget(main_w)

        if self._courses_data:
            self._select_course(0)

    def _select_course(self, idx):
        for i, c in enumerate(self._course_cards):
            c.set_active(i == idx)
        self._active_idx = idx
        course = self._courses_data[idx]
        code = str(course.get("SectionID", ""))
        name = course.get("CourseTitle", "")
        sched_raw = course.get("Schedule", "")
        if isinstance(sched_raw, dict):
            sched = f"{sched_raw.get('Day', '')} {sched_raw.get('StartTime', '')}–{sched_raw.get('EndTime', '')}"
        else:
            sched = str(sched_raw)
        enrolled_ids = course.get("EnrolledStudents", [])
        students = []
        if DB_CONNECTED and enrolled_ids:
            try:
                students = list(db["Students"].find({"StudentID": {"$in": enrolled_ids}}))
            except Exception as _e:
                print(f"[AIAS] warning: {_e}")
        self._title_lbl.setText(name)
        self._sub_lbl.setText(f"{code} · {sched} · {len(students)} students")
        self._enrolled_val.setText(str(len(students)))
        self._tbl.setRowCount(len(students))
        for r, stu in enumerate(students):
            self._tbl.setItem(r, 0, QTableWidgetItem(str(stu.get("StudentID", ""))))
            self._tbl.setItem(r, 1, QTableWidgetItem(stu.get("FullName", "")))
            self._tbl.setCellWidget(r, 2, make_badge("Enrolled"))
            self._tbl.setRowHeight(r, 44)

        if DB_CONNECTED:
            try:
                n_sessions = db["Sessions"].count_documents({"SectionID": code, "Status": "completed"})
                last_sess  = db["Sessions"].find_one({"SectionID": code, "Status": "completed"}, sort=[("EndTime", -1)])
                self._sessions_held_lbl.setText(str(n_sessions))
                self._last_session_lbl.setText(last_sess["EndTime"].strftime("%Y-%m-%d") if last_sess else "—")
            except Exception as _e:
                print(f"[AIAS] warning: {_e}")

    def _start_session(self):
        if not self._courses_data:
            return
        course     = self._courses_data[self._active_idx]
        session_id = None
        enrolled_students = []

        if DB_CONNECTED:
            try:
                enrolled_ids      = course.get("EnrolledStudents", [])
                enrolled_students = list(db["Students"].find({"StudentID": {"$in": enrolled_ids}}))
            except Exception as e:
                print(f"DB error: {e}")

        # Rec 2: warn if no student has a face embedding
        if enrolled_students:
            n_with_face = sum(1 for s in enrolled_students if s.get("FaceEmbedding"))
            if n_with_face == 0:
                reply = QMessageBox.warning(
                    self, "No Face Data",
                    f"None of the {len(enrolled_students)} enrolled students have face embeddings.\n"
                    "Attendance will need to be entered manually.\n\nProceed anyway?",
                    QMessageBox.Yes | QMessageBox.No,
                )
                if reply != QMessageBox.Yes:
                    return

        if DB_CONNECTED:
            try:
                session_doc = {
                    "CourseID":           course.get("CourseID", course.get("SectionID")),
                    "SectionID":          course.get("SectionID"),
                    "InstructorID":       self.instructor_data.get("username", ""),
                    "StartTime":          datetime.now(),
                    "Status":             "active",
                    "AttendanceInterval": 5,
                }
                result     = db["Sessions"].insert_one(session_doc)
                session_id = result.inserted_id
            except Exception as e:
                print(f"DB error: {e}")

        cam_idx = choose_camera_dialog(self)
        if cam_idx is None:
            return  # user cancelled

        start_btn = self.sender()
        if start_btn:
            start_btn.setText("▶  Starting...")
            start_btn.setEnabled(False)
            QApplication.processEvents()

        self.hide()
        self._live = LiveSessionWindow(self, session_id, enrolled_students, course, camera_index=cam_idx)
        self._live.show()

        if start_btn:
            start_btn.setText("▶  Start session")
            start_btn.setEnabled(True)

    def _show_dashboard(self):
        from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout,
                                     QLabel, QFrame, QScrollArea, QWidget)
        from PyQt5.QtCore import Qt
        import datetime

        dlg = QDialog(self)
        dlg.setWindowTitle("Dashboard")
        dlg.setMinimumSize(820, 580)
        dlg.setStyleSheet(f"background:{BG};")

        import os
        from PyQt5.QtGui import QIcon
        _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aias_icon.ico")
        if os.path.exists(_icon_path):
            dlg.setWindowIcon(QIcon(_icon_path))

        main_lay = QVBoxLayout(dlg)
        main_lay.setContentsMargins(28, 24, 28, 24)
        main_lay.setSpacing(20)

        # Title
        _full_name = self.instructor_data.get("fullname", "Instructor")
        _username  = self.instructor_data.get("username", "")
        title = QLabel("📊  Instructor Dashboard")
        title.setStyleSheet(f"font-size:22px; font-weight:800; color:{TEXT_DARK}; background:transparent;")
        sub = QLabel(f"Welcome, {_full_name}  ·  {datetime.date.today().strftime('%B %d, %Y')}")
        sub.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY}; background:transparent;")
        main_lay.addWidget(title)
        main_lay.addWidget(sub)

        # Load data from DB
        try:
            db = get_db()
            # Get all sections for this instructor
            instructor = db["Instructors"].find_one({"Username": _username})
            sections = instructor.get("Sections", []) if instructor else []

            total_students = 0
            total_sessions = 0
            total_present  = 0
            total_records  = 0
            absent_count   = {}

            for sec_id in sections:
                # Students in section
                n_students = db["Students"].count_documents({"Sections": sec_id})
                total_students += n_students

                # Sessions for this section
                sessions = list(db["AttendanceSessions"].find({"SectionID": sec_id}))
                total_sessions += len(sessions)

                for session in sessions:
                    records = session.get("Records", [])
                    for rec in records:
                        total_records += 1
                        status = rec.get("Status", "")
                        if status in ("Present", "Late"):
                            total_present += 1
                        elif status == "Absent":
                            sid = rec.get("StudentID", "")
                            absent_count[sid] = absent_count.get(sid, 0) + 1

            attendance_rate = (total_present / total_records * 100) if total_records > 0 else 0

            # Top absentees
            top_absent = sorted(absent_count.items(), key=lambda x: x[1], reverse=True)[:5]
            top_absent_data = []
            for sid, cnt in top_absent:
                student = db["Students"].find_one({"StudentID": sid})
                name = student.get("FullName", sid) if student else sid
                top_absent_data.append((name, cnt))

        except Exception as e:
            print(f"[AIAS] Dashboard error: {e}")
            sections = []
            total_students = total_sessions = attendance_rate = 0
            top_absent_data = []

        # Stats cards row
        cards_row = QHBoxLayout()
        cards_row.setSpacing(16)

        def make_stat_card(value, label, color):
            card = QFrame()
            card.setStyleSheet(
                f"QFrame {{ background:{WHITE}; border-radius:12px; "
                f"border-left:4px solid {color}; }}"
            )
            card.setMinimumHeight(90)
            lay = QVBoxLayout(card)
            lay.setContentsMargins(20, 14, 20, 14)
            v_lbl = QLabel(str(value))
            v_lbl.setStyleSheet(f"font-size:28px; font-weight:800; color:{color}; background:transparent;")
            l_lbl = QLabel(label)
            l_lbl.setStyleSheet(f"font-size:12px; color:{TEXT_GRAY}; background:transparent;")
            lay.addWidget(v_lbl)
            lay.addWidget(l_lbl)
            return card

        cards_row.addWidget(make_stat_card(total_students, "Total Students", PRIMARY))
        cards_row.addWidget(make_stat_card(total_sessions, "Sessions Held", "#2563EB"))
        cards_row.addWidget(make_stat_card(f"{attendance_rate:.1f}%", "Attendance Rate", "#16A34A"))
        cards_row.addWidget(make_stat_card(len(sections), "My Courses", "#9333EA"))
        main_lay.addLayout(cards_row)

        # Attendance rate bar
        bar_frame = QFrame()
        bar_frame.setStyleSheet(f"QFrame {{ background:{WHITE}; border-radius:12px; }}")
        bar_lay = QVBoxLayout(bar_frame)
        bar_lay.setContentsMargins(20, 16, 20, 16)
        bar_title = QLabel("Overall Attendance Rate")
        bar_title.setStyleSheet(f"font-size:14px; font-weight:700; color:{TEXT_DARK}; background:transparent;")
        bar_lay.addWidget(bar_title)

        from PyQt5.QtWidgets import QProgressBar
        progress = QProgressBar()
        progress.setMinimum(0)
        progress.setMaximum(100)
        progress.setValue(int(attendance_rate))
        color = PRIMARY if attendance_rate >= 75 else "#F59E0B" if attendance_rate >= 50 else "#DC2626"
        progress.setStyleSheet(f"""
            QProgressBar {{ background:#E5E7EB; border-radius:8px; height:18px; text-align:center; color:{TEXT_DARK}; font-weight:600; }}
            QProgressBar::chunk {{ background:{color}; border-radius:8px; }}
        """)
        bar_lay.addWidget(progress)
        main_lay.addWidget(bar_frame)

        # Most absent students
        if top_absent_data:
            absent_frame = QFrame()
            absent_frame.setStyleSheet(f"QFrame {{ background:{WHITE}; border-radius:12px; }}")
            absent_lay = QVBoxLayout(absent_frame)
            absent_lay.setContentsMargins(20, 16, 20, 16)
            absent_lay.setSpacing(10)

            absent_title = QLabel("⚠️  Most Absent Students")
            absent_title.setStyleSheet(f"font-size:14px; font-weight:700; color:{TEXT_DARK}; background:transparent;")
            absent_lay.addWidget(absent_title)

            for name, cnt in top_absent_data:
                row = QHBoxLayout()
                name_lbl = QLabel(f"• {name}")
                name_lbl.setStyleSheet(f"font-size:13px; color:{TEXT_DARK}; background:transparent;")
                cnt_lbl = QLabel(f"{cnt} absences")
                cnt_lbl.setStyleSheet(f"font-size:13px; color:#DC2626; font-weight:600; background:transparent;")
                row.addWidget(name_lbl)
                row.addStretch()
                row.addWidget(cnt_lbl)
                absent_lay.addLayout(row)

            main_lay.addWidget(absent_frame)

        main_lay.addStretch()
        dlg.exec_()

    def _open_history(self):
        self.hide()
        self._history_win = SessionHistoryWindow(self)
        self._history_win.show()

    def _open_analytics(self):
        self.hide()
        self._analytics_win = AnalyticsWindow(self)
        self._analytics_win.show()

    def _open_email_settings(self):
        univ_email = ""
        if DB_CONNECTED:
            try:
                instr = db["Instructors"].find_one(
                    {"Username": self.instructor_data.get("username", "")}
                )
                if instr:
                    univ_email = instr.get("UniversityEmail", "")
            except Exception as e:
                print(f"[AIAS] Email settings load error: {e}")

        dlg = QDialog(self)
        dlg.setWindowTitle("Email Settings")
        dlg.setFixedWidth(440)
        dlg.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(14)

        t = QLabel("Email Settings")
        t.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(t)

        if univ_email:
            status_html = (
                f"<b style='color:#22C55E;'>✔ Configured</b><br>"
                f"<span style='color:#374151;'>Send to: {univ_email}</span>"
            )
            status_color = "#F0FDF4"
            border_color = "#86EFAC"
        else:
            status_html = (
                "<b style='color:#F59E0B;'>⚠ Not configured</b><br>"
                "<span style='color:#6B7280;'>You'll be prompted to set up email the first time you send a report.</span>"
            )
            status_color = "#FFFBEB"
            border_color = "#FCD34D"

        status_lbl = QLabel()
        status_lbl.setText(status_html)
        status_lbl.setStyleSheet(
            f"background:{status_color}; border:1px solid {border_color};"
            "border-radius:8px; padding:12px;"
        )
        status_lbl.setWordWrap(True)
        lay.addWidget(status_lbl)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        close_btn = QPushButton("Close")
        close_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        close_btn.clicked.connect(dlg.reject)

        reset_btn = QPushButton("Reset / Clear All")
        reset_btn.setStyleSheet(
            "background:#FEE2E2; color:#DC2626; border-radius:6px; padding:8px 14px; font-weight:600;"
        )
        reset_btn.clicked.connect(lambda: dlg.done(2))

        update_btn = QPushButton("Update Settings")
        update_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; padding:8px 16px; font-weight:700;"
        )
        update_btn.clicked.connect(dlg.accept)

        btn_row.addWidget(close_btn)
        btn_row.addStretch()
        btn_row.addWidget(reset_btn)
        btn_row.addWidget(update_btn)
        lay.addLayout(btn_row)

        result = dlg.exec_()

        if result == QDialog.Accepted:
            self._open_email_settings_edit(univ_email)
        elif result == 2:
            reply = QMessageBox.question(
                self, "Confirm Reset",
                "Clear the saved destination email?\n"
                "You'll need to re-enter it the next time you send a report.",
                QMessageBox.Yes | QMessageBox.No,
            )
            if reply == QMessageBox.Yes and DB_CONNECTED:
                try:
                    db["Instructors"].update_one(
                        {"Username": self.instructor_data.get("username", "")},
                        {"$unset": {"UniversityEmail": ""}},
                    )
                    QMessageBox.information(self, "Cleared", "Email destination has been cleared.")
                except Exception as e:
                    QMessageBox.critical(self, "Error", str(e))

    def _open_email_settings_edit(self, current_univ_email=""):
        dlg = QDialog(self)
        dlg.setWindowTitle("Update Email Settings")
        dlg.setFixedWidth(420)
        dlg.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        t = QLabel("Update Email Settings")
        t.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(t)

        sub = QLabel("Enter the university email address to send reports to.")
        sub.setStyleSheet(f"font-size:12px; color:{TEXT_GRAY};")
        sub.setWordWrap(True)
        lay.addWidget(sub)

        lbl = QLabel("Send reports to:")
        lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        univ_field = QLineEdit()
        univ_field.setPlaceholderText("name@university.edu")
        univ_field.setFixedHeight(38)
        univ_field.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:6px; padding:6px 10px; background:{WHITE}; color:{TEXT_DARK};"
        )
        univ_field.setText(current_univ_email)
        lay.addWidget(lbl)
        lay.addWidget(univ_field)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.clicked.connect(dlg.reject)
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; padding:8px 20px; font-weight:700;"
        )
        save_btn.clicked.connect(dlg.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        if dlg.exec_() != QDialog.Accepted:
            return

        univ_email = univ_field.text().strip()
        if not univ_email:
            QMessageBox.warning(self, "Missing Field", "Please enter a destination email address.")
            return

        if DB_CONNECTED:
            try:
                db["Instructors"].update_one(
                    {"Username": self.instructor_data.get("username", "")},
                    {"$set": {"UniversityEmail": univ_email}},
                )
                QMessageBox.information(self, "Saved", "Email settings saved successfully.")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save settings:\n{e}")

    def _toggle_dark_mode(self, btn):
        new_dark = not IS_DARK_MODE
        apply_theme(QApplication.instance(), dark=new_dark)
        btn.setText("☀  Light Mode" if new_dark else "🌙  Dark Mode")
        btn.setStyleSheet(
            f"QPushButton {{ background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px; "
            f"padding:8px 12px; font-size:12px; font-weight:600; }}"
            f"QPushButton:hover {{ background:{PRIMARY}; color:#ffffff; }}"
            f"QPushButton:pressed {{ background:{PRIMARY}; color:#ffffff; padding:9px 11px 7px 13px; }}"
        )
        for widget in QApplication.instance().topLevelWidgets():
            force_theme_on_all_widgets(widget)
            widget.setStyleSheet(f"background-color:{BG};")
            widget.update()
            widget.repaint()

    def _change_password(self):
        from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QLabel,
                                     QLineEdit, QPushButton, QHBoxLayout)
        import bcrypt

        dlg = QDialog(self)
        dlg.setWindowTitle("Change Password")
        dlg.setStyleSheet(f"background:{BG};")
        dlg.setMinimumWidth(340)

        lay = QVBoxLayout(dlg)
        lay.setSpacing(14)
        lay.setContentsMargins(28, 28, 28, 28)

        title = QLabel("Change Password")
        title.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title)

        def make_field(label_text, placeholder):
            lbl = QLabel(label_text)
            lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_DARK};")
            field = QLineEdit()
            field.setPlaceholderText(placeholder)
            field.setEchoMode(QLineEdit.Password)
            field.setStyleSheet(
                f"border:1px solid {BORDER}; border-radius:8px; padding:9px 12px; "
                f"background:{WHITE}; color:{TEXT_DARK}; font-size:13px;"
            )
            lay.addWidget(lbl)
            lay.addWidget(field)
            return field

        current_field = make_field("Current Password", "Enter current password")
        new_field     = make_field("New Password", "Enter new password")
        confirm_field = make_field("Confirm New Password", "Re-enter new password")

        err_lbl = QLabel("")
        err_lbl.setStyleSheet("font-size:11px; color:#DC2626;")
        err_lbl.setWordWrap(True)
        lay.addWidget(err_lbl)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; "
            "border-radius:6px; padding:8px 16px; font-weight:600;"
        )
        cancel_btn.clicked.connect(dlg.reject)

        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; "
            "padding:8px 16px; font-weight:600;"
        )

        def do_change():
            current  = current_field.text().strip()
            new_pass = new_field.text().strip()
            confirm  = confirm_field.text().strip()

            if not current or not new_pass or not confirm:
                err_lbl.setText("Please fill in all fields.")
                return
            if new_pass != confirm:
                err_lbl.setText("New passwords do not match.")
                return
            if len(new_pass) < 6:
                err_lbl.setText("Password must be at least 6 characters.")
                return

            # Verify current password
            username = self.instructor_data.get("username", "")
            user = db["Instructors"].find_one({"Username": username})
            if not user:
                err_lbl.setText("User not found.")
                return

            stored = user.get("Password", "")
            try:
                if isinstance(stored, str):
                    stored = stored.encode()
                if not bcrypt.checkpw(current.encode(), stored):
                    err_lbl.setText("Current password is incorrect.")
                    return
            except Exception:
                err_lbl.setText("Error verifying password.")
                return

            # Save new password
            new_hash = bcrypt.hashpw(new_pass.encode(), bcrypt.gensalt())
            db["Instructors"].update_one(
                {"Username": username},
                {"$set": {"Password": new_hash.decode()}}
            )
            dlg.accept()
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.information(self, "Success", "Password changed successfully.")

        save_btn.clicked.connect(do_change)
        btn_row.addWidget(cancel_btn)
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)
        dlg.exec_()

    def _logout(self):
        for widget in QApplication.instance().topLevelWidgets():
            if isinstance(widget, AppWindow):
                widget._login_page.username_field.clear()
                widget._login_page.password_field.clear()
                widget._stack.setCurrentWidget(widget._login_page)
                widget.setStyleSheet("background:#0a0f0a;")
                break
        # Remove this page from stack
        parent_stack = self.parent()
        if parent_stack:
            try:
                stack = parent_stack.parent()
                if hasattr(stack, 'removeWidget'):
                    stack.removeWidget(self)
            except Exception as e:
                print(f"[AIAS] Logout stack cleanup error: {e}")
        self.deleteLater()

    def closeEvent(self, event):
        # Stop any active live session worker when main window closes
        try:
            if hasattr(self, '_live') and self._live is not None:
                try:
                    if hasattr(self._live, '_worker') and self._live._worker is not None:
                        self._live._worker.stop()
                        self._live._worker.quit()
                        self._live._worker.wait(2000)
                except Exception as e:
                    print(f"[AIAS] closeEvent worker stop error: {e}")
                try:
                    self._live.close()
                except Exception as e:
                    print(f"[AIAS] closeEvent live close error: {e}")
        except Exception as e:
            print(f"[AIAS] closeEvent error: {e}")
        event.accept()


def get_available_cameras(max_check=5):
    """Returns a list of (index, label) for all available cameras."""
    import cv2 as _cv2
    cameras = []
    for i in range(max_check):
        cap = _cv2.VideoCapture(i, _cv2.CAP_DSHOW)
        if cap is not None and cap.isOpened():
            cameras.append((i, f"Camera {i}" if i > 0 else f"Camera {i} (Default)"))
            cap.release()
    return cameras

def choose_camera_dialog(parent=None):
    """Shows a dialog to let the user pick a camera. Returns selected index or 0."""
    cameras = get_available_cameras()
    if len(cameras) <= 1:
        return 0  # only one camera, use it directly

    from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QComboBox, QPushButton, QHBoxLayout
    dlg = QDialog(parent)
    dlg.setWindowTitle("Select Camera")
    dlg.setStyleSheet(f"background:{BG};")
    dlg.setMinimumWidth(300)

    lay = QVBoxLayout(dlg)
    lay.setSpacing(14)
    lay.setContentsMargins(24, 24, 24, 24)

    title = QLabel("Select Camera")
    title.setStyleSheet(f"font-size:15px; font-weight:700; color:{TEXT_DARK};")
    lay.addWidget(title)

    sub = QLabel("Choose which camera to use for face recognition:")
    sub.setStyleSheet(f"font-size:12px; color:{TEXT_GRAY};")
    sub.setWordWrap(True)
    lay.addWidget(sub)

    combo = QComboBox()
    combo.setStyleSheet(
        f"border:1px solid {BORDER}; border-radius:6px; padding:6px 10px; "
        f"background:{WHITE}; color:{TEXT_DARK}; font-size:13px;"
    )
    for idx, label in cameras:
        combo.addItem(label, idx)
    lay.addWidget(combo)

    btn_row = QHBoxLayout()
    cancel_btn = QPushButton("Cancel")
    cancel_btn.setStyleSheet(
        f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; "
        "border-radius:6px; padding:8px 16px; font-weight:600;"
    )
    cancel_btn.clicked.connect(dlg.reject)

    confirm_btn = QPushButton("Start Session")
    confirm_btn.setStyleSheet(
        f"background:{PRIMARY}; color:white; border-radius:6px; "
        "padding:8px 16px; font-weight:600;"
    )
    confirm_btn.clicked.connect(dlg.accept)

    btn_row.addWidget(cancel_btn)
    btn_row.addWidget(confirm_btn)
    lay.addLayout(btn_row)

    result = dlg.exec_()
    if result == QDialog.Accepted:
        return combo.currentData()
    return None  # user cancelled


class _ProcessingWorker(QThread):
    result_ready = pyqtSignal(dict)

    def __init__(self, frame, cv2_mod, np_mod, yolo, arcface, parent_worker):
        super().__init__()
        self._frame   = frame
        self._cv2     = cv2_mod
        self._np      = np_mod
        self._yolo    = yolo
        self._arcface = arcface
        self._worker  = parent_worker

    def run(self):
        try:
            result = self._worker._process_frame(
                self._frame, self._cv2, self._np, self._yolo, self._arcface
            )
            self.result_ready.emit(result)
        except Exception as e:
            print(f"[AIAS] ProcessingWorker error: {e}")
            self.result_ready.emit({})


def _detect_faces_multiscale(_cv2, _np, _yolo, frame, conf=0.25):
    """
    Multi-scale face detection: runs YOLO at 1x, 2x, and 4x zoom.
    4x zoom catches faces at 5-7 meters from a ceiling-mounted camera.
    All coordinates mapped back to original frame space.
    """
    h, w = frame.shape[:2]
    all_boxes = []

    # ── 1. Native resolution ──────────────────────────────────────────
    try:
        res = _yolo(frame, verbose=False, imgsz=640, conf=conf)
        for box in res[0].boxes:
            x1, y1, x2, y2 = box.xyxy[0].cpu().numpy().astype(int)
            all_boxes.append((x1, y1, x2, y2))
    except Exception:
        pass

    # ── 2. 2x upscale ─────────────────────────────────────────────────
    try:
        up2 = _cv2.resize(frame, (w * 2, h * 2))
        res2 = _yolo(up2, verbose=False, imgsz=640, conf=conf)
        for box in res2[0].boxes:
            bx1, by1, bx2, by2 = box.xyxy[0].cpu().numpy()
            all_boxes.append((
                int(bx1 / 2), int(by1 / 2),
                int(bx2 / 2), int(by2 / 2)
            ))
    except Exception:
        pass

    # ── 3. 4x upscale — catches faces at 5-7 meters ───────────────────
    try:
        up4 = _cv2.resize(frame, (w * 4, h * 4))
        res4 = _yolo(up4, verbose=False, imgsz=640, conf=conf)
        for box in res4[0].boxes:
            bx1, by1, bx2, by2 = box.xyxy[0].cpu().numpy()
            all_boxes.append((
                int(bx1 / 4), int(by1 / 4),
                int(bx2 / 4), int(by2 / 4)
            ))
    except Exception:
        pass

    if not all_boxes:
        return []

    # ── 4. IoU deduplication ──────────────────────────────────────────
    def _iou(a, b):
        ix1 = max(a[0], b[0]); iy1 = max(a[1], b[1])
        ix2 = min(a[2], b[2]); iy2 = min(a[3], b[3])
        inter = max(0, ix2 - ix1) * max(0, iy2 - iy1)
        if inter == 0:
            return 0.0
        area_a = (a[2]-a[0]) * (a[3]-a[1])
        area_b = (b[2]-b[0]) * (b[3]-b[1])
        return inter / (area_a + area_b - inter + 1e-6)

    kept = []
    used = [False] * len(all_boxes)
    for i, box_a in enumerate(all_boxes):
        if used[i]:
            continue
        for j in range(i + 1, len(all_boxes)):
            if not used[j] and _iou(box_a, all_boxes[j]) > 0.35:
                used[j] = True
        kept.append(box_a)
        used[i] = True

    return kept


class CaptureWorker(QThread):
    recognized  = pyqtSignal(list)
    frame_ready = pyqtSignal(QImage, dict)
    camera_ok   = pyqtSignal(bool)

    def __init__(self, interval_minutes, session_id, enrolled_students, camera_index=0):
        super().__init__()
        self.interval_minutes   = interval_minutes
        self.session_id         = session_id
        self.enrolled_students  = enrolled_students
        self._camera_index      = camera_index
        self._running           = True
        self._emb_buffer        = {}
        self._buffer_size       = 1
        self._frame_count       = 0
        self._attendance        = {}
        self._processing_worker = None
        self._recognizing       = False
        self._interval_ms       = 5 * 60 * 1000
        self._timer_reset       = False
        self._capture_active    = False
        self._score_accumulator    = {}
        self._cumulative_threshold = 0.20
        self._last_det_boxes   = []          # cached boxes from last detection
        self._last_det_time    = 0           # timestamp of last detection (ms)
        self._det_interval_ms  = 1500        # run detection every 1500ms
        self._detecting        = False       # prevent overlapping detection threads
        self._recognition_results = __import__('queue').Queue()  # thread-safe result passing

    def set_interval(self, ms):
        """Change capture interval in milliseconds while running."""
        self._interval_ms = ms

    def reset_timer(self):
        """Reset capture timer to zero — next capture fires after interval."""
        self._timer_reset = True

    def run(self):
        import time

        try:
            import cv2 as _cv2
            import numpy as _np
        except Exception as e:
            print(f"[AIAS] cv2/numpy not available: {e}")
            return

        _yolo, _arcface, _ai_ok = _ensure_ai_models(det_size=(1280, 1280))
        if _ai_ok:
            print("[AIAS] CaptureWorker using pre-loaded AI models")

        try:
            cap = _cv2.VideoCapture(self._camera_index)
            if not cap.isOpened():
                cap = None
        except Exception as e:
            print(f"[AIAS] Camera init failed: {e}")
            cap = None

        self.camera_ok.emit(cap is not None)

        last_capture = int(time.time() * 1000)

        while self._running:
            if cap:
                ret, frame = cap.read()
            else:
                ret = False

            if ret:
                # ── Draw face bounding boxes on every frame ──
                if _ai_ok:
                    # ── Run detection in background every 500ms ──
                    _now_ms = int(time.time() * 1000)
                    if (not self._detecting and not self._recognizing and
                            _now_ms - self._last_det_time >= self._det_interval_ms):
                        self._detecting = True
                        self._last_det_time = _now_ms
                        _det_frame = frame.copy()
                        def _run_det(_f=_det_frame):
                            try:
                                boxes = _detect_faces_multiscale(_cv2, _np, _yolo, _f, conf=0.25)
                                self._last_det_boxes = boxes
                            except Exception:
                                pass
                            finally:
                                self._detecting = False
                        _threading.Thread(target=_run_det, daemon=True).start()

                    # ── Draw cached boxes on current frame (non-blocking) ──
                    for (_bx1, _by1, _bx2, _by2) in self._last_det_boxes:
                        try:
                            _cv2.rectangle(frame, (_bx1, _by1), (_bx2, _by2), (0, 220, 100), 2)
                            _cv2.putText(
                                frame, "Detecting...",
                                (_bx1, max(_by1 - 8, 12)),
                                _cv2.FONT_HERSHEY_SIMPLEX, 0.5,
                                (0, 220, 100), 1, _cv2.LINE_AA
                            )
                        except Exception:
                            pass

                rgb  = _cv2.cvtColor(frame, _cv2.COLOR_BGR2RGB)
                h, w, ch = rgb.shape
                qimg = QImage(rgb.data, w, h, ch * w, QImage.Format_RGB888)

                # Handle timer reset from user clicking "Start Capture"
                if self._timer_reset:
                    last_capture = int(time.time() * 1000)
                    self._timer_reset = False
                    self._capture_active = True

                interval_ms = self._interval_ms
                now_ms = int(time.time() * 1000)
                if self._capture_active and now_ms - last_capture >= interval_ms and _ai_ok:
                    last_capture = now_ms
                    # Non-blocking recognition — pass current frame directly
                    if not self._recognizing:
                        self._recognizing = True
                        _frame_copy = frame.copy()  # snapshot current frame — no cap.read() conflict
                        _yolo_ref   = _yolo
                        _arc_ref    = _arcface

                        def _run_recognition(_f=_frame_copy):
                            try:
                                result = self._process_frame(_f, _cv2, _np, _yolo_ref, _arc_ref)
                                if result:
                                    self._recognition_results.put(list(result.keys()))
                            except Exception as e:
                                print(f"[AIAS] Recognition error: {e}")
                            finally:
                                self._recognizing = False

                        _threading.Thread(target=_run_recognition, daemon=True).start()
                else:
                    self.frame_ready.emit(qimg, {})

            self._frame_count += 1

            self.msleep(33)

        if cap:
            cap.release()

    def _process_frame(self, frame, _cv2, _np, _yolo, _arcface):
        recognized_this_frame = {}
        ms_boxes = _detect_faces_multiscale(_cv2, _np, _yolo, frame, conf=0.25)
        if not ms_boxes:
            return recognized_this_frame

        for (x1, y1, x2, y2) in ms_boxes:
            try:
                x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                if x2 <= x1 or y2 <= y1:
                    continue
                face_crop_raw = frame[y1:y2, x1:x2]
                # Apply GFPGAN enhancement if available
                try:
                    if _gfpgan_model is not None and face_crop_raw.shape[0] > 10 and face_crop_raw.shape[1] > 10:
                        _, _, gfpgan_output = _gfpgan_model.enhance(
                            face_crop_raw,
                            has_aligned=False,
                            only_center_face=True,
                            paste_back=True
                        )
                        if gfpgan_output is not None:
                            face_crop_raw = gfpgan_output
                except Exception:
                    pass

                # Apply CodeFormer on top of GFPGAN for best quality
                try:
                    import torch
                    if _codeformer_net is not None and _codeformer_model is not None:
                        _codeformer_model.clean_all()
                        _codeformer_model.read_image(face_crop_raw)
                        _codeformer_model.get_face_landmarks_5(
                            only_center_face=True, resize=640, eye_dist_threshold=5
                        )
                        _codeformer_model.align_warp_face()
                        for idx in range(_codeformer_model.cropped_faces.__len__()):
                            cropped_face_t = _codeformer_model.cropped_faces[idx]
                            cropped_face_t = torch.from_numpy(cropped_face_t).float() / 255.
                            cropped_face_t = (cropped_face_t * 2 - 1).permute(2, 0, 1).unsqueeze(0)
                            with torch.no_grad():
                                output = _codeformer_net(cropped_face_t, w=0.7, adain=True)[0]
                            restored = output.squeeze(0).permute(1, 2, 0)
                            restored = (restored + 1) / 2
                            restored = restored.clamp(0, 1).numpy()
                            restored = (restored * 255).astype('uint8')
                            import cv2 as _cv2_cf
                            restored_bgr = _cv2_cf.cvtColor(restored, _cv2_cf.COLOR_RGB2BGR)
                            _codeformer_model.add_restored_face(restored_bgr, cropped_face_t)
                        _codeformer_model.paste_faces_to_input_image()
                        cf_output = _codeformer_model.output
                        if cf_output is not None and cf_output.shape[0] > 10:
                            face_crop_raw = cf_output
                except Exception:
                    pass
                face_crop = upscale_face_crop(face_crop_raw)
                face_crop = apply_clahe(face_crop)
                face_crop_rgb = _cv2.cvtColor(face_crop, _cv2.COLOR_BGR2RGB)

                try:
                    dynamic_thresh = compute_dynamic_threshold(face_crop_raw)
                except Exception as e:
                    print(f"[AIAS] Dynamic threshold error: {e}")
                    dynamic_thresh = RECOGNITION_THRESHOLD

                faces = _arcface.get(face_crop_rgb)
                if not faces:
                    faces = _arcface.get(frame)
                if not faces:
                    continue

                for face in faces:
                    try:
                        embedding = face.embedding
                        if embedding is None:
                            continue

                        norm = _np.linalg.norm(embedding)
                        if norm == 0:
                            continue
                        normed_emb = embedding / norm

                        # position key based on YOLO bbox center in frame coords
                        cx = (x1 + x2) / 2
                        cy = (y1 + y2) / 2
                        pos_key = f"{int(cx/50)}_{int(cy/50)}"

                        # accumulate embedding in buffer
                        if pos_key not in self._emb_buffer:
                            self._emb_buffer[pos_key] = []
                        self._emb_buffer[pos_key].append(normed_emb)

                        # wait until buffer is full before matching
                        if len(self._emb_buffer[pos_key]) < self._buffer_size:
                            continue

                        avg_emb = _np.mean(self._emb_buffer[pos_key], axis=0)
                        avg_emb = avg_emb / _np.linalg.norm(avg_emb)
                        self._emb_buffer[pos_key] = []  # reset for next window

                        best_match = None
                        best_score = 0.0
                        best_name  = None

                        for stu in self.enrolled_students:
                            db_emb = stu.get("Embedding", stu.get("FaceEmbedding", []))
                            if not db_emb:
                                continue
                            stored = _np.array(db_emb, dtype=_np.float32)
                            if stored.ndim != 1 or stored.shape[0] != avg_emb.shape[0]:
                                continue
                            stored_norm = _np.linalg.norm(stored)
                            if stored_norm < 1e-6:
                                continue
                            s_buf = float(_np.dot(avg_emb, stored / stored_norm))
                            stu_id = str(stu.get("StudentID", ""))

                            # AdaFace disabled in live capture — too slow on CPU
                            s_final = s_buf

                            if s_final > best_score:
                                best_score = s_final
                                best_match = stu.get("StudentID")
                                best_name  = stu.get("FullName", "")

                        print(f"[AIAS] Face {pos_key}: score={best_score:.3f} thresh={dynamic_thresh:.3f}")
                        # Instant recognition — no accumulation needed
                        INSTANT_THRESHOLD = 0.22

                        if best_match is not None and best_score >= INSTANT_THRESHOLD:
                            sid_str = str(best_match)
                            if sid_str in self._attendance:
                                continue
                            matched = True
                            print(f"[AIAS] Matched {sid_str}: score={best_score:.3f} ✅")
                        else:
                            matched = False

                        if matched:
                            sid = str(best_match)
                            recognized_this_frame[sid] = {
                                "name":       best_name,
                                "similarity": best_score,
                                "bbox":       [int(x1), int(y1), int(x2), int(y2)],
                            }
                            # Draw green box with student name on the recognition frame
                            _cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 220, 100), 2)
                            _label = str(best_name).split()[0] if best_name else ""
                            if _label:
                                _cv2.putText(
                                    frame, _label,
                                    (x1, max(y1 - 8, 12)),
                                    _cv2.FONT_HERSHEY_SIMPLEX, 0.55,
                                    (0, 220, 100), 2, _cv2.LINE_AA
                                )
                    except Exception as _fe:
                        print(f"[AIAS] face error: {_fe}")
                        continue

            except Exception as _be:
                print(f"[AIAS] box error: {_be}")
                continue

        return recognized_this_frame

    def stop(self):
        self._running = False
        self._score_accumulator.clear()
        self.wait()


def _augment_image(img_bgr):
    """
    Generate 8 augmented variants of a BGR face image.
    Fast, in-memory only — no disk writes.
    Returns a list of BGR images (the 8 augmented variants, NOT including the original).
    """
    import cv2
    import numpy as np
    import random

    variants = []
    h, w = img_bgr.shape[:2]

    # 1. Horizontal flip
    variants.append(cv2.flip(img_bgr, 1))

    # 2. Brightness +30
    bright = np.clip(img_bgr.astype(np.int32) + 30, 0, 255).astype(np.uint8)
    variants.append(bright)

    # 3. Brightness -30
    dark = np.clip(img_bgr.astype(np.int32) - 30, 0, 255).astype(np.uint8)
    variants.append(dark)

    # 4. Slight rotation +10 degrees
    M = cv2.getRotationMatrix2D((w / 2, h / 2), 10, 1.0)
    variants.append(cv2.warpAffine(img_bgr, M, (w, h)))

    # 5. Slight rotation -10 degrees
    M = cv2.getRotationMatrix2D((w / 2, h / 2), -10, 1.0)
    variants.append(cv2.warpAffine(img_bgr, M, (w, h)))

    # 6. Gaussian blur (simulate soft focus)
    variants.append(cv2.GaussianBlur(img_bgr, (3, 3), 0))

    # 7. Contrast adjustment (CLAHE already applied later, so light stretch here)
    lab = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    l = np.clip(l.astype(np.float32) * 1.2, 0, 255).astype(np.uint8)
    variants.append(cv2.cvtColor(cv2.merge([l, a, b]), cv2.COLOR_LAB2BGR))

    # 8. Small random crop + resize back
    crop_margin = max(4, int(min(h, w) * 0.08))
    y1 = random.randint(0, crop_margin)
    x1 = random.randint(0, crop_margin)
    y2 = h - random.randint(0, crop_margin)
    x2 = w - random.randint(0, crop_margin)
    if y2 > y1 + 10 and x2 > x1 + 10:
        cropped = img_bgr[y1:y2, x1:x2]
        variants.append(cv2.resize(cropped, (w, h)))
    else:
        variants.append(cv2.flip(img_bgr, 0))  # fallback: vertical flip

    return variants  # always exactly 8 items


class EmbeddingWorker(QThread):
    """Worker thread to extract face embedding without blocking the UI."""
    finished = pyqtSignal(list)   # embedding list
    error    = pyqtSignal(str)    # error message

    def __init__(self, image_paths):
        super().__init__()
        self.image_paths = image_paths if isinstance(image_paths, list) else [image_paths]

    def run(self):
        try:
            import cv2
            import numpy as np

            yolo, arcface, ok = _ensure_ai_models(det_size=(640, 640))
            if not ok:
                self.error.emit("AI models failed to load. Check that insightface and ultralytics are installed.")
                return

            embeddings = []
            for image_path in self.image_paths:
                img = cv2.imread(image_path)
                if img is None:
                    continue
                img = apply_clahe(img)

                results = yolo(img, verbose=False)
                boxes = results[0].boxes
                if len(boxes) == 0:
                    continue

                x1, y1, x2, y2 = boxes[0].xyxy[0].cpu().numpy().astype(int)
                if x2 <= x1 or y2 <= y1:
                    continue
                face_crop = img[y1:y2, x1:x2]
                face_crop = upscale_face_crop(face_crop, target_size=112)

                faces = arcface.get(face_crop)
                if not faces:
                    faces = arcface.get(img)
                if not faces:
                    continue

                raw = faces[0].embedding
                embeddings.append(raw / np.linalg.norm(raw))

                # ── Auto-augmentation (in-memory, no disk writes) ──
                # For each original photo, generate 8 variants and embed them
                aug_variants = _augment_image(face_crop)
                for aug_img in aug_variants:
                    try:
                        aug_clahe = apply_clahe(aug_img)
                        aug_faces = arcface.get(aug_clahe)
                        if not aug_faces:
                            aug_faces = arcface.get(aug_img)
                        if aug_faces:
                            aug_raw = aug_faces[0].embedding
                            embeddings.append(aug_raw / np.linalg.norm(aug_raw))
                    except Exception:
                        pass  # skip failed augmented variants silently

            if not embeddings:
                self.error.emit("No face detected in any of the selected photos.")
                return

            avg = np.mean(np.array(embeddings), axis=0)
            embedding = (avg / np.linalg.norm(avg)).tolist()
            self.finished.emit(embedding)

        except Exception as e:
            self.error.emit(str(e))


class BatchEmbedWorker(QThread):
    """Processes a root folder where each subfolder name = StudentID.
    Embeds all images per student and saves to MongoDB."""
    progress = pyqtSignal(str)          # status message per student
    finished = pyqtSignal(int, int)     # (processed, skipped)

    def __init__(self, root_folder, data_folder):
        super().__init__()
        self.root_folder = root_folder
        self.data_folder = data_folder  # where to copy images

    def run(self):
        import cv2
        import numpy as np
        import shutil

        yolo, arcface, ok = _ensure_ai_models(det_size=(640, 640))
        if not ok:
            self.progress.emit("ERROR: AI models failed to load.")
            self.finished.emit(0, 0)
            return

        processed = skipped = 0
        try:
            entries = [e for e in os.scandir(self.root_folder) if e.is_dir()]
        except Exception as e:
            self.progress.emit(f"ERROR reading folder: {e}")
            self.finished.emit(0, 0)
            return

        for entry in entries:
            sid = entry.name.strip()
            if not sid:
                continue

            student = db["Students"].find_one({"StudentID": sid}) if DB_CONNECTED else None
            if student is None:
                self.progress.emit(f"⚠ Skipped '{sid}' — not found in database")
                skipped += 1
                continue

            images = [
                f for f in os.listdir(entry.path)
                if f.lower().endswith((".jpg", ".jpeg", ".png"))
            ]
            if not images:
                self.progress.emit(f"⚠ Skipped '{sid}' — no images in folder")
                skipped += 1
                continue

            def _embed_one(img_file):
                try:
                    img_path = os.path.join(entry.path, img_file)
                    img = cv2.imread(img_path)
                    if img is None:
                        return [], None
                    img = apply_clahe(img)
                    results = yolo(img, verbose=False)
                    boxes = results[0].boxes
                    if len(boxes) == 0:
                        return [], None
                    x1, y1, x2, y2 = boxes[0].xyxy[0].cpu().numpy().astype(int)
                    face_crop = img[y1:y2, x1:x2]
                    face_crop = upscale_face_crop(face_crop, target_size=112)
                    faces = arcface.get(face_crop)
                    if not faces:
                        faces = arcface.get(img)
                    if not faces:
                        return [], None
                    raw = faces[0].embedding
                    embs = [raw / np.linalg.norm(raw)]

                    # ── Auto-augmentation (in-memory, no disk writes) ──
                    aug_variants = _augment_image(face_crop)
                    for aug_img in aug_variants:
                        try:
                            aug_clahe = apply_clahe(aug_img)
                            aug_faces = arcface.get(aug_clahe)
                            if not aug_faces:
                                aug_faces = arcface.get(aug_img)
                            if aug_faces:
                                aug_raw = aug_faces[0].embedding
                                embs.append(aug_raw / np.linalg.norm(aug_raw))
                        except Exception:
                            pass

                    # Copy image to Data folder
                    dest_dir = os.path.join(self.data_folder, sid)
                    os.makedirs(dest_dir, exist_ok=True)
                    dest = os.path.join(dest_dir, img_file)
                    if os.path.abspath(img_path) != os.path.abspath(dest):
                        try:
                            shutil.copy2(img_path, dest)
                        except Exception as e:
                            print(f"[AIAS] Copy error for {img_file}: {e}")
                    return embs, os.path.abspath(dest)
                except Exception:
                    return [], None

            with ThreadPoolExecutor(max_workers=4) as ex:
                results_parallel = list(ex.map(_embed_one, images))
            all_embeddings = [emb for r in results_parallel for emb in r[0]]
            dest_paths = [r[1] for r in results_parallel if r[1] is not None]

            if not all_embeddings:
                self.progress.emit(f"⚠ Skipped '{sid}' — no face detected in any image")
                skipped += 1
                continue

            mean_emb = np.mean(np.array(all_embeddings), axis=0)
            norm_emb = (mean_emb / np.linalg.norm(mean_emb)).tolist()

            if DB_CONNECTED:
                db["Students"].update_one(
                    {"StudentID": sid},
                    {"$set": {"FaceEmbedding": norm_emb, "ImagePaths": dest_paths}}
                )

            # Build AdaFace embedding for this student and save to MongoDB
            global _adaface_model, _adaface_db
            if _adaface_model is not None:
                try:
                    ada_embeddings = []
                    # Use original images + up to 4 augmented for AdaFace (faster)
                    orig_imgs = [f for f in images if not f.startswith('aug_')]
                    aug_imgs  = [f for f in images if f.startswith('aug_')][:4]
                    ada_imgs  = orig_imgs + aug_imgs
                    for img_file in ada_imgs:
                        img_ada = cv2.imread(os.path.join(entry.path, img_file))
                        if img_ada is not None:
                            ada_emb = _adaface_embedding(_adaface_model, img_ada)
                            if ada_emb is not None:
                                ada_embeddings.append(ada_emb)
                    if ada_embeddings:
                        avg_ada = np.mean(ada_embeddings, axis=0)
                        avg_ada = avg_ada / np.linalg.norm(avg_ada)
                        if DB_CONNECTED:
                            db["Students"].update_one(
                                {"StudentID": sid},
                                {"$set": {"AdaFaceEmbedding": avg_ada.tolist()}}
                            )
                        _adaface_db[str(sid)] = avg_ada
                        print(f"[AIAS] AdaFace saved: {sid} ({len(ada_embeddings)} images)")
                except Exception as e:
                    print(f"[AIAS] AdaFace batch save error: {e}")

            self.progress.emit(f"✔ '{sid}' — embedded from {len(all_embeddings)} image(s)")
            processed += 1

        self.finished.emit(processed, skipped)


class LiveSessionWindow(QWidget):
    def __init__(self, main_win, session_id=None, enrolled_students=None, course_data=None, camera_index=0):
        super().__init__()
        self.main_win          = main_win
        self.session_id        = session_id
        self.enrolled_students = enrolled_students or []
        self.course_data       = course_data or {}
        self._attendance       = {}
        self._session_start    = datetime.now()
        self._worker           = None
        self._interval_minutes = 5
        self._camera_index     = camera_index
        self._capture_running  = False
        self._session_ended    = False
        self.setWindowTitle("AIAS — Live Session")
        import os
        from PyQt5.QtGui import QIcon
        _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aias_icon.ico")
        if os.path.exists(_icon_path):
            self.setWindowIcon(QIcon(_icon_path))
        self.setMinimumSize(1100, 700)
        self.resize(1100, 700)
        self.setStyleSheet(f"background:{BG};")
        self._build()

    def _build(self):
        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        instr    = getattr(self.main_win, "instructor_data", {})
        fullname = instr.get("fullname", "Dr. Ahmed")
        parts    = fullname.split()
        initials = (parts[0][0] + parts[-1][0]).upper() if len(parts) >= 2 else fullname[:2].upper()
        sidebar, s_lay = make_sidebar_base(initials, fullname, "Instructor")

        self._cam_preview = QLabel("Camera feed")
        self._cam_preview.setFixedHeight(120)
        self._cam_preview.setAlignment(Qt.AlignCenter)
        self._cam_preview.setStyleSheet(
            f"background:#D1D5DB; border-radius:8px; border:none;"
            f"color:{TEXT_GRAY}; font-size:12px;"
        )
        s_lay.addWidget(self._cam_preview)
        s_lay.addSpacing(12)

        # ── Lecture duration ──────────────────────────
        lec_label = QLabel("Lecture duration")
        lec_label.setStyleSheet("font-size:12px; color: #888; font-weight:500;")

        lec_h_layout = QHBoxLayout()
        lec_h_layout.setSpacing(4)

        self._lec_hours = QLineEdit("1")
        self._lec_hours.setFixedWidth(45)
        self._lec_hours.setFixedHeight(36)
        self._lec_hours.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._lec_hours.setStyleSheet("""
            QLineEdit {
                background: #1a1a2e; color: white;
                border: 1px solid #2a2a4a; border-radius: 8px;
                font-size: 16px; font-weight: bold;
            }
        """)

        colon_lbl = QLabel(":")
        colon_lbl.setStyleSheet("color: white; font-size:18px; font-weight:bold;")
        colon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self._lec_mins = QLineEdit("30")
        self._lec_mins.setFixedWidth(45)
        self._lec_mins.setFixedHeight(36)
        self._lec_mins.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._lec_mins.setStyleSheet("""
            QLineEdit {
                background: #1a1a2e; color: white;
                border: 1px solid #2a2a4a; border-radius: 8px;
                font-size: 16px; font-weight: bold;
            }
        """)

        from PyQt5.QtGui import QIntValidator as _IVLec
        self._lec_hours.setValidator(_IVLec(0, 23))
        self._lec_mins.setValidator(_IVLec(0, 59))

        lec_h_layout.addWidget(self._lec_hours)
        lec_h_layout.addWidget(colon_lbl)
        lec_h_layout.addWidget(self._lec_mins)
        lec_h_layout.addStretch()

        lec_container = QVBoxLayout()
        lec_container.setSpacing(6)
        lec_container.addWidget(lec_label)
        lec_container.addLayout(lec_h_layout)

        s_lay.addLayout(lec_container)
        s_lay.addSpacing(12)

        # ── Capture interval ──────────────────────────
        interval_label = QLabel("Capture interval")
        interval_label.setStyleSheet("font-size:12px; color: #888; font-weight:500;")

        self._interval_input = QLineEdit("5")
        self._interval_input.setFixedHeight(38)
        self._interval_input.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._interval_input.setStyleSheet("""
            QLineEdit {
                background: #1a1a2e;
                color: white;
                border: 1px solid #2a2a4a;
                border-radius: 8px;
                font-size: 18px;
                font-weight: bold;
            }
            QLineEdit:focus {
                border: 1px solid #1B5E35;
            }
        """)
        from PyQt5.QtGui import QIntValidator
        self._interval_input.setValidator(QIntValidator(1, 9999))

        self._interval_unit = "M"
        unit_layout = QHBoxLayout()
        unit_layout.setSpacing(6)
        self._unit_buttons = {}
        for unit, label in [("S", "S"), ("M", "M"), ("H", "H")]:
            btn = QPushButton(label)
            btn.setFixedHeight(34)
            btn.setCheckable(True)
            btn.setChecked(unit == "M")
            btn.setStyleSheet("""
                QPushButton {
                    background: #1a1a2e;
                    color: #888;
                    border: 1px solid #2a2a4a;
                    border-radius: 8px;
                    font-size: 13px;
                    font-weight: bold;
                }
                QPushButton:checked {
                    background: #1B5E35;
                    color: white;
                    border: 1px solid #1B5E35;
                }
                QPushButton:hover:!checked {
                    background: #2a2a4a;
                    color: white;
                }
            """)
            self._unit_buttons[unit] = btn
            unit_layout.addWidget(btn)

        def _set_unit(u):
            self._interval_unit = u
            for k, b in self._unit_buttons.items():
                b.setChecked(k == u)
            _update_interval()

        def _update_interval():
            try:
                val = int(self._interval_input.text() or "5")
            except ValueError:
                val = 5
            multipliers = {"S": 1000, "M": 60000, "H": 3600000}
            ms = val * multipliers.get(self._interval_unit, 60000)
            if hasattr(self, '_worker') and self._worker:
                self._worker.set_interval(ms)

        for unit, btn in self._unit_buttons.items():
            btn.clicked.connect(lambda _, u=unit: _set_unit(u))
        self._interval_input.textChanged.connect(lambda _: _update_interval())

        self._btn_start_capture = QPushButton("▶  Start Capture")
        self._btn_start_capture.setFixedHeight(42)
        self._btn_start_capture.setStyleSheet("""
            QPushButton {
                background: #1B5E35;
                color: white;
                border: none;
                border-radius: 10px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2E7D4F;
            }
            QPushButton:pressed {
                background: #145228;
            }
        """)
        self._btn_start_capture.clicked.connect(self._start_capture_timer)

        interval_container = QVBoxLayout()
        interval_container.setSpacing(6)
        interval_container.addWidget(interval_label)
        interval_container.addWidget(self._interval_input)
        interval_container.addLayout(unit_layout)
        interval_container.addWidget(self._btn_start_capture)

        s_lay.addLayout(interval_container)
        self._start_capture(5)

        s_lay.addStretch()

        dark_btn = QPushButton("🌙  Dark Mode" if not IS_DARK_MODE else "☀  Light Mode")
        dark_btn.setStyleSheet(
            f"background:{'#2A3D35' if IS_DARK_MODE else PRIMARY_LIGHT}; "
            f"color:{'#4ecf8e' if IS_DARK_MODE else PRIMARY}; "
            "border-radius:6px; padding:8px 12px; font-size:12px; font-weight:600;"
        )
        dark_btn.setCursor(Qt.PointingHandCursor)
        dark_btn.clicked.connect(lambda: self._toggle_dark_mode(dark_btn))
        s_lay.addWidget(dark_btn)
        s_lay.addSpacing(4)

        end_btn = QPushButton("End session")
        end_btn.setFixedHeight(40)
        end_btn.setStyleSheet(
            "background:#DC2626; color:white; border-radius:6px;"
            "padding:8px 16px; font-size:13px; font-weight:700;"
        )
        end_btn.setCursor(Qt.PointingHandCursor)
        end_btn.clicked.connect(self._end_session)
        s_lay.addWidget(end_btn)

        main_w = QWidget()
        main_w.setStyleSheet(f"background:{BG};")
        m_lay = QVBoxLayout(main_w)
        m_lay.setContentsMargins(32, 28, 32, 24)
        m_lay.setSpacing(18)

        bar = QFrame()
        bar.setStyleSheet(
            f"QFrame {{ background:{PRIMARY_LIGHT}; border:1.5px solid {PRIMARY}; border-radius:8px; }}"
        )
        bar_lay = QVBoxLayout(bar)
        bar_lay.setContentsMargins(16, 8, 16, 8)
        bar_lay.setSpacing(4)

        self._lbl_lecture_timer = QLabel("Lecture ends in   --:--")
        self._lbl_lecture_timer.setStyleSheet(
            "font-size:22px; font-weight:bold; color:#1B5E35; background:transparent; border:none;"
        )
        self._lbl_lecture_timer.setAlignment(Qt.AlignCenter)

        self._lbl_capture_timer = QLabel("Next capture in   --:--")
        self._lbl_capture_timer.setStyleSheet(
            "font-size:14px; color:#888; background:transparent; border:none;"
        )
        self._lbl_capture_timer.setAlignment(Qt.AlignCenter)

        bar_lay.addWidget(self._lbl_lecture_timer)
        bar_lay.addWidget(self._lbl_capture_timer)
        m_lay.addWidget(bar)

        stats_row = QHBoxLayout()
        stats_row.setSpacing(16)

        def _make_live_card(title, color):
            frame = QFrame()
            frame.setStyleSheet(
                f"QFrame {{ background:{WHITE}; border:1px solid {BORDER}; border-radius:8px; }}"
            )
            frame.setFixedHeight(82)
            frame.setMinimumWidth(130)
            fl = QVBoxLayout(frame)
            fl.setContentsMargins(16, 12, 16, 12)
            fl.setSpacing(2)
            v = QLabel("0")
            v.setStyleSheet(
                f"font-size:26px; font-weight:800; color:{color}; border:none; background:transparent;"
            )
            t = QLabel(title)
            t.setStyleSheet(f"font-size:11px; color:{TEXT_GRAY}; border:none; background:transparent;")
            fl.addWidget(v)
            fl.addWidget(t)
            return frame, v

        present_card, self._present_val = _make_live_card("Present", "#22C55E")
        late_card,    self._late_val    = _make_live_card("Late",    "#F59E0B")
        absent_card,  self._absent_val  = _make_live_card("Absent",  "#EF4444")
        stats_row.addWidget(present_card)
        stats_row.addWidget(late_card)
        stats_row.addWidget(absent_card)
        stats_row.addStretch()
        m_lay.addLayout(stats_row)

        # Rec 3: persistent camera-unavailable banner (hidden until camera fails)
        self._cam_banner = QFrame()
        self._cam_banner.setStyleSheet(
            "QFrame { background:#FFFBEB; border:1px solid #FCD34D; border-radius:8px; }"
        )
        self._cam_banner.setFixedHeight(44)
        self._cam_banner.setVisible(False)
        _ban_lay = QHBoxLayout(self._cam_banner)
        _ban_lay.setContentsMargins(16, 0, 16, 0)
        _ban_lbl = QLabel("⚠  Camera unavailable — attendance must be entered manually.")
        _ban_lbl.setStyleSheet(
            "font-size:12px; font-weight:600; color:#92400E; background:transparent; border:none;"
        )
        _ban_lay.addWidget(_ban_lbl)
        m_lay.addWidget(self._cam_banner)

        self._live_tbl = make_table(
            ["Student ID", "Full Name", "First seen", "Last seen", "Status", ""],
            self.enrolled_students,
            col_widths={0: 140, 2: 100, 3: 100, 4: 130, 5: 70},
            stretch_col=1,
        )
        m_lay.addWidget(self._live_tbl)

        root.addWidget(sidebar)
        root.addWidget(main_w)

        self._remaining_seconds = self._interval_minutes * 60
        self._timer = QTimer()
        self._timer.timeout.connect(self._tick)
        self._timer.start(1000)

        self._lecture_remaining = 0
        self._lecture_timer_qt  = QTimer()
        self._lecture_timer_qt.setInterval(1000)
        self._lecture_timer_qt.timeout.connect(self._tick_lecture_timer)

        self._capture_countdown_remaining = 0
        self._capture_countdown_qt = QTimer()
        self._capture_countdown_qt.setInterval(1000)
        self._capture_countdown_qt.timeout.connect(self._tick_capture_timer)

        self._refresh_table()
        self._start_capture(5)

    def _toggle_dark_mode(self, btn):
        new_dark = not IS_DARK_MODE
        apply_theme(QApplication.instance(), dark=new_dark)
        btn.setText("☀  Light Mode" if new_dark else "🌙  Dark Mode")
        btn.setStyleSheet(
            f"QPushButton {{ background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px; "
            f"padding:8px 12px; font-size:12px; font-weight:600; }}"
            f"QPushButton:hover {{ background:{PRIMARY}; color:#ffffff; }}"
            f"QPushButton:pressed {{ background:{PRIMARY}; color:#ffffff; padding:9px 11px 7px 13px; }}"
        )
        for widget in QApplication.instance().topLevelWidgets():
            force_theme_on_all_widgets(widget)
            widget.setStyleSheet(f"background-color:{BG};")
            widget.update()
            widget.repaint()

    def _start_capture(self, interval_minutes):
        if self._worker:
            self._worker.stop()
        self._worker = CaptureWorker(interval_minutes, self.session_id, self.enrolled_students, camera_index=self._camera_index)
        self._worker.recognized.connect(self._on_recognized)
        self._worker.frame_ready.connect(self._update_preview)
        self._worker.camera_ok.connect(self._on_camera_status)
        self._worker.start()

        # ── Poll recognition queue from main thread every 200ms ──
        if not hasattr(self, '_recognition_poll_timer'):
            from PyQt5.QtCore import QTimer
            self._recognition_poll_timer = QTimer(self)
            self._recognition_poll_timer.timeout.connect(self._poll_recognition_queue)
        self._recognition_poll_timer.start(50)

    def _poll_recognition_queue(self):
        """Drain the recognition results queue from the main thread."""
        if not self._worker:
            return
        try:
            q = self._worker._recognition_results
            found = False
            while not q.empty():
                sids = q.get_nowait()
                self._on_recognized(sids)
                found = True
            if found:
                from PyQt5.QtWidgets import QApplication
                QApplication.processEvents()
        except Exception:
            pass

    def _tick(self):
        if self._remaining_seconds > 0:
            self._remaining_seconds -= 1
        else:
            self._remaining_seconds = self._interval_minutes * 60

    def _tick_lecture_timer(self):
        if self._lecture_remaining > 0:
            self._lecture_remaining -= 1
            h = self._lecture_remaining // 3600
            m = (self._lecture_remaining % 3600) // 60
            s = self._lecture_remaining % 60
            self._lbl_lecture_timer.setText(f"Lecture ends in   {h}:{m:02d}:{s:02d}")
        else:
            self._lecture_timer_qt.stop()
            self._lbl_lecture_timer.setText("Lecture ended")

    def _tick_capture_timer(self):
        if self._capture_countdown_remaining > 0:
            self._capture_countdown_remaining -= 1
            s_total = self._capture_countdown_remaining
            m = s_total // 60
            s = s_total % 60
            self._lbl_capture_timer.setText(f"Next capture in   {m}:{s:02d}")
        else:
            self._lbl_capture_timer.setText("Capturing...")
            # Reset countdown for next capture
            try:
                val = int(self._interval_input.text() or "5")
            except Exception:
                val = 5
            multipliers = {"S": 1, "M": 60, "H": 3600}
            self._capture_countdown_remaining = val * multipliers.get(
                getattr(self, "_interval_unit", "M"), 60
            )

    def _on_recognized(self, student_ids):
        now     = datetime.now()
        elapsed = (now - self._session_start).total_seconds() / 60

        for sid in student_ids:
            sid = str(sid)
            if sid not in self._attendance:
                status = "Present" if elapsed <= LATE_CUTOFF_MINUTES else "Late"
                self._attendance[sid] = {
                    "status":        status,
                    "first_seen":    now.strftime("%H:%M"),
                    "first_seen_dt": now,
                    "last_seen":     now.strftime("%H:%M"),
                    "last_seen_dt":  now,
                }
                if DB_CONNECTED:
                    try:
                        db["AttendanceLogs"].update_one(
                            {"SessionID": self.session_id, "StudentID": sid},
                            {"$set": {
                                "SessionID": self.session_id,
                                "StudentID": sid,
                                "Status":    status,
                                "FirstSeenAt": now,
                                "LastSeenAt":  now,
                            }},
                            upsert=True,
                        )
                    except Exception as e:
                        print(f"DB error: {e}")
            else:
                self._attendance[sid]["last_seen"]    = now.strftime("%H:%M")
                self._attendance[sid]["last_seen_dt"] = now
                if DB_CONNECTED:
                    try:
                        db["AttendanceLogs"].update_one(
                            {"SessionID": self.session_id, "StudentID": sid},
                            {"$set": {"LastSeenAt": now}},
                        )
                    except Exception as e:
                        print(f"DB error: {e}")

        for sid, data in self._attendance.items():
            if sid not in student_ids:
                last = data.get("last_seen_dt") or datetime.strptime(
                    data["last_seen"], "%H:%M"
                ).replace(year=now.year, month=now.month, day=now.day)
                if (now - last).total_seconds() / 60 >= EARLY_LEAVE_CUTOFF_MINUTES:
                    self._attendance[sid]["status"] = "Early Leave"

        self._refresh_table()

    def _update_preview(self, qimg, recognized):
        pixmap = QPixmap.fromImage(qimg)
        self._cam_preview.setPixmap(
            pixmap.scaled(
                self._cam_preview.width(), self._cam_preview.height(),
                Qt.KeepAspectRatio, Qt.SmoothTransformation,
            )
        )

    def _refresh_table(self):
        n = len(self.enrolled_students)
        self._live_tbl.setRowCount(n)
        n_present = n_late = n_absent = 0
        for r, stu in enumerate(self.enrolled_students):
            sid  = str(stu.get("StudentID", ""))
            name = stu.get("FullName", "")
            att    = self._attendance.get(sid, {})
            first  = att.get("first_seen", "—")
            last   = att.get("last_seen",  "—")
            status = att.get("status",     "Absent")
            if status in ("Present", "Early Leave"):
                n_present += 1
            elif status == "Late":
                n_late += 1
            elif status == "Absent":
                n_absent += 1
            self._live_tbl.setItem(r, 0, QTableWidgetItem(sid))
            self._live_tbl.setItem(r, 1, QTableWidgetItem(name))
            self._live_tbl.setItem(r, 2, QTableWidgetItem(first))
            self._live_tbl.setItem(r, 3, QTableWidgetItem(last))
            self._live_tbl.setCellWidget(r, 4, make_badge(status))
            edit_btn = QPushButton("Edit")
            edit_btn.setFixedHeight(28)
            edit_btn.setStyleSheet(
                f"background:#E6F1FB; color:#0C447C; border-radius:5px;"
                "font-size:11px; font-weight:600;"
            )
            edit_btn.setCursor(Qt.PointingHandCursor)
            edit_btn.clicked.connect(lambda _, s=sid: self._edit_attendance(s))
            self._live_tbl.setCellWidget(r, 5, edit_btn)
            self._live_tbl.setRowHeight(r, 44)
        self._present_val.setText(str(n_present))
        self._late_val.setText(str(n_late))
        self._absent_val.setText(str(n_absent))

    def _set_interval(self, mins):
        self._interval_minutes = mins
        self._remaining_seconds = mins * 60
        self._start_capture(mins)

    def _start_capture_timer(self):
        if self._capture_running:
            # Stop capture
            self._capture_running = False
            self._btn_start_capture.setText("▶  Start Capture")
            self._btn_start_capture.setStyleSheet("""
                QPushButton {
                    background: #1B5E35; color: white;
                    border: none; border-radius: 10px;
                    font-size: 13px; font-weight: bold;
                }
                QPushButton:hover { background: #2E7D4F; }
            """)
            if hasattr(self, '_worker') and self._worker:
                self._worker._capture_active = False
            self._lecture_timer_qt.stop()
            self._capture_countdown_qt.stop()
            self._lbl_lecture_timer.setText("Lecture ends in   --:--")
            self._lbl_capture_timer.setText("Next capture in   --:--")
        else:
            # Start capture
            self._capture_running = True
            self._btn_start_capture.setText("⏹  Stop Capture")
            self._btn_start_capture.setStyleSheet("""
                QPushButton {
                    background: #8B0000; color: white;
                    border: none; border-radius: 10px;
                    font-size: 13px; font-weight: bold;
                }
                QPushButton:hover { background: #A00000; }
            """)
            try:
                val = int(self._interval_input.text() or "5")
            except ValueError:
                val = 5
            multipliers = {"S": 1000, "M": 60000, "H": 3600000}
            ms = val * multipliers.get(self._interval_unit, 60000)
            if hasattr(self, '_worker') and self._worker:
                self._worker.set_interval(ms)
                self._worker.reset_timer()
                print(f"[AIAS] Capture timer started: every {val}{self._interval_unit}")
            # Start lecture timer
            try:
                lec_h = int(self._lec_hours.text() or "1")
                lec_m = int(self._lec_mins.text() or "30")
            except Exception:
                lec_h, lec_m = 1, 30
            self._lecture_remaining = lec_h * 3600 + lec_m * 60
            self._lecture_timer_qt.start()
            # Start capture countdown
            multipliers_s = {"S": 1, "M": 60, "H": 3600}
            self._capture_countdown_remaining = val * multipliers_s.get(self._interval_unit, 60)
            self._capture_countdown_qt.start()

    def _on_camera_status(self, ok):
        if not ok:
            self._cam_preview.setText("⚠ No camera\ndetected")
            self._cam_preview.setStyleSheet(
                "background:#FEE2E2; border-radius:8px; border:1px solid #FCA5A5;"
                "color:#DC2626; font-size:11px; font-weight:600;"
            )
            self._cam_banner.setVisible(True)

    def _edit_attendance(self, sid):
        name = next(
            (s.get("FullName", sid) for s in self.enrolled_students
             if str(s.get("StudentID", "")) == sid),
            sid,
        )
        current_status = self._attendance.get(sid, {}).get("status", "Absent")
        current_note   = self._attendance.get(sid, {}).get("note", "")

        dlg = QDialog(self)
        dlg.setWindowTitle(f"Edit Attendance — {name}")
        dlg.setFixedWidth(340)
        dlg.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(24, 20, 24, 20)
        lay.setSpacing(10)

        title_lbl = QLabel(f"Override status for:\n{name}")
        title_lbl.setStyleSheet(f"font-size:14px; font-weight:700; color:{TEXT_DARK};")
        title_lbl.setWordWrap(True)
        lay.addWidget(title_lbl)

        group = QButtonGroup(dlg)
        buttons = {}
        for status in ["Present", "Late", "Early Leave", "Absent"]:
            rb = QRadioButton(status)
            rb.setStyleSheet(f"font-size:13px; color:{TEXT_MED};")
            if status == current_status:
                rb.setChecked(True)
            group.addButton(rb)
            buttons[status] = rb
            lay.addWidget(rb)

        note_lbl = QLabel("Note (optional)")
        note_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        note_field = QLineEdit()
        note_field.setPlaceholderText("e.g. Medical excuse, arrived late by bus…")
        note_field.setFixedHeight(36)
        note_field.setText(current_note)
        lay.addWidget(note_lbl)
        lay.addWidget(note_field)

        btn_row = QHBoxLayout()
        cancel = QPushButton("Cancel")
        cancel.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:7px 14px;"
        )
        cancel.clicked.connect(dlg.reject)
        save = QPushButton("Save")
        save.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px;"
            "padding:7px 14px; font-weight:700;"
        )
        save.clicked.connect(dlg.accept)
        btn_row.addWidget(cancel)
        btn_row.addStretch()
        btn_row.addWidget(save)
        lay.addLayout(btn_row)

        if dlg.exec_() != QDialog.Accepted:
            return

        new_status = next(
            (s for s, rb in buttons.items() if rb.isChecked()), current_status
        )
        new_note = note_field.text().strip()
        if new_status == current_status and new_note == current_note:
            return

        now = datetime.now()
        if sid not in self._attendance:
            self._attendance[sid] = {
                "status":        new_status,
                "first_seen":    now.strftime("%H:%M"),
                "first_seen_dt": now,
                "last_seen":     now.strftime("%H:%M"),
                "last_seen_dt":  now,
                "note":          new_note,
            }
        else:
            self._attendance[sid]["status"] = new_status
            self._attendance[sid]["note"]   = new_note

        if DB_CONNECTED and self.session_id:
            try:
                db["AttendanceLogs"].update_one(
                    {"SessionID": self.session_id, "StudentID": sid},
                    {"$set": {"Status": new_status, "Note": new_note}},
                    upsert=True,
                )
            except Exception as e:
                print(f"[AIAS] Override error: {e}")

        self._refresh_table()

    def _end_session(self):
        if self._session_ended:
            return
        self._session_ended = True
        self._timer.stop()
        if self._worker:
            self._worker.stop()
        if hasattr(self, '_recognition_poll_timer'):
            self._recognition_poll_timer.stop()

        # Apply Early Leave to any Present/Late student not seen within the cutoff window
        _now = datetime.now()
        for _sid, _data in self._attendance.items():
            if _data.get("status") in ("Present", "Late"):
                _last = _data.get("last_seen_dt")
                if _last and (_now - _last).total_seconds() / 60 >= EARLY_LEAVE_CUTOFF_MINUTES:
                    self._attendance[_sid]["status"] = "Early Leave"

        if DB_CONNECTED and self.session_id:
            # Flush final statuses for all students (Early Leave may have just changed some)
            for sid, data in self._attendance.items():
                try:
                    db["AttendanceLogs"].update_one(
                        {"SessionID": self.session_id, "StudentID": sid},
                        {"$set": {"Status": data.get("status", "Absent"),
                                  "Note":   data.get("note", "")}},
                    )
                except Exception as e:
                    print(f"DB error: {e}")
            for stu in self.enrolled_students:
                sid = str(stu.get("StudentID", ""))
                if sid not in self._attendance:
                    try:
                        db["AttendanceLogs"].update_one(
                            {"SessionID": self.session_id, "StudentID": sid},
                            {"$set": {
                                "SessionID": self.session_id,
                                "StudentID": sid,
                                "Status":    "Absent",
                                "FirstSeenAt": None,
                                "LastSeenAt":  None,
                            }},
                            upsert=True,
                        )
                    except Exception as e:
                        print(f"DB error: {e}")
            try:
                db["Sessions"].update_one(
                    {"_id": self.session_id},
                    {"$set": {"EndTime": datetime.now(), "Status": "completed"}},
                )
            except Exception as e:
                print(f"DB error: {e}")

        # Auto-send email silently using per-instructor saved credentials
        _auto_email_ok = None
        if DB_CONNECTED:
            try:
                instr_username = getattr(self.main_win, "instructor_data", {}).get("username", "")
                instr_doc = db["Instructors"].find_one({"Username": instr_username})
                if instr_doc:
                    to_email = instr_doc.get("UniversityEmail", "")
                    if to_email:
                        _auto_email_ok = self._auto_send_report(to_email)
            except Exception as e:
                print(f"[AIAS] Auto-email error: {e}")
                _auto_email_ok = False

        self.hide()
        self._report = ReportWindow(
            self.main_win,
            session_id=self.session_id,
            attendance=self._attendance,
            enrolled_students=self.enrolled_students,
            course_data=self.course_data,
            auto_email_ok=_auto_email_ok,
        )
        self._report.show()

    def _auto_send_report(self, to_email):
        course_title = self.course_data.get("CourseTitle", "Attendance")
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        _success = False
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            ws.append(["Student ID", "Full Name", "First Seen", "Last Seen", "Status", "Note"])
            for stu in self.enrolled_students:
                sid = str(stu.get("StudentID", ""))
                rec = self._attendance.get(sid, {})
                first = rec.get("first_seen") or "—"
                last  = rec.get("last_seen")  or "—"
                ws.append([
                    sid,
                    stu.get("FullName", ""),
                    first,
                    last,
                    rec.get("status", "Absent"),
                    rec.get("note",   ""),
                ])
            wb.save(tmp.name)

            att_statuses = {str(s.get("StudentID","")): self._attendance.get(str(s.get("StudentID","")), {}).get("status", "Absent")
                            for s in self.enrolled_students}
            n_present = sum(1 for v in att_statuses.values() if v == "Present")
            n_late    = sum(1 for v in att_statuses.values() if v == "Late")
            n_early   = sum(1 for v in att_statuses.values() if v == "Early Leave")
            n_absent  = sum(1 for v in att_statuses.values() if v == "Absent")

            msg = MIMEMultipart()
            msg["From"]    = config.SMTP_FROM
            msg["To"]      = to_email
            msg["Subject"] = (
                f"Attendance Report — {course_title} — "
                f"{datetime.now().strftime('%Y-%m-%d')}"
            )
            body = (
                f"Attendance report for {course_title}.\n\n"
                f"Date: {datetime.now().strftime('%Y-%m-%d')}\n"
                f"Total: {len(self.enrolled_students)}  "
                f"Present: {n_present}  Late: {n_late}  "
                f"Early Leave: {n_early}  Absent: {n_absent}\n\n"
                "Sent automatically by AIAS."
            )
            msg.attach(MIMEText(body, "plain"))
            with open(tmp.name, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            safe = "".join(c for c in course_title if c.isalnum() or c in "-_ ")
            part.add_header(
                "Content-Disposition",
                f'attachment; filename="Attendance_{safe}.xlsx"',
            )
            msg.attach(part)

            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()
                server.login(config.SMTP_LOGIN, config.SMTP_PASSWORD)
                server.send_message(msg)
            print(f"[AIAS] Auto-report sent to {to_email}")
            _success = True
        except Exception as e:
            print(f"[AIAS] Auto-send failed: {e}")
        finally:
            try:
                os.unlink(tmp.name)
            except Exception as e:
                print(f"[AIAS] Auto-report temp file cleanup error: {e}")
        return _success

    def closeEvent(self, event):
        from PyQt5.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            self,
            "End Session",
            "Are you sure you want to end the attendance session?\nAll recorded attendance will be saved.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            # Stop worker thread cleanly before closing
            if self._worker is not None:
                try:
                    self._worker.stop()
                except Exception as e:
                    print(f"[AIAS] Session closeEvent worker stop error: {e}")
                try:
                    self._worker.quit()
                    self._worker.wait(2000)  # wait max 2 seconds
                except Exception as e:
                    print(f"[AIAS] Session closeEvent worker quit error: {e}")
                self._worker = None
            event.accept()
        else:
            event.ignore()


class ReportWindow(QWidget):
    def __init__(self, main_win, session_id=None, attendance=None, enrolled_students=None, course_data=None, auto_email_ok=None):
        super().__init__()
        self.main_win          = main_win
        self.session_id        = session_id
        self.attendance        = attendance or {}
        self.enrolled_students = enrolled_students or []
        self.course_data       = course_data or {}
        self.auto_email_ok     = auto_email_ok
        self.setWindowTitle("AIAS — Session Report")
        import os
        from PyQt5.QtGui import QIcon
        _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aias_icon.ico")
        if os.path.exists(_icon_path):
            self.setWindowIcon(QIcon(_icon_path))
        self.setMinimumSize(1100, 700)
        self.resize(1100, 700)
        self.setStyleSheet(f"background:{BG};")
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(48, 36, 48, 28)
        root.setSpacing(18)

        rows = []
        for stu in self.enrolled_students:
            sid  = str(stu.get("StudentID", ""))
            name = stu.get("FullName", "")
            att = self.attendance.get(sid, {})
            rows.append((
                sid, name,
                att.get("first_seen", "—"),
                att.get("last_seen",  "—"),
                att.get("status",     "Absent"),
                att.get("note",       ""),
            ))

        course_title = self.course_data.get("CourseTitle", "")
        sched_raw    = self.course_data.get("Schedule", "")
        if isinstance(sched_raw, dict):
            sched = f"{sched_raw.get('Day','')} {sched_raw.get('StartTime','')}–{sched_raw.get('EndTime','')}"
        else:
            sched = str(sched_raw)
        n_students = len(rows)

        title = QLabel("Session ended")
        title.setStyleSheet(f"font-size:24px; font-weight:800; color:{TEXT_DARK};")
        sub = QLabel(f"{course_title} · {sched} · {n_students} students")
        sub.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY};")
        root.addWidget(title)
        root.addWidget(sub)

        n_present     = sum(1 for _, _, _, _, s, _ in rows if s == "Present")
        n_late        = sum(1 for _, _, _, _, s, _ in rows if s == "Late")
        n_early_leave = sum(1 for _, _, _, _, s, _ in rows if s == "Early Leave")
        n_absent      = sum(1 for _, _, _, _, s, _ in rows if s == "Absent")
        stats_row = QHBoxLayout()
        stats_row.setSpacing(16)
        stats_row.addWidget(make_stat_card("Present",     n_present,     "#22C55E"))
        stats_row.addWidget(make_stat_card("Late",        n_late,        "#F59E0B"))
        stats_row.addWidget(make_stat_card("Early Leave", n_early_leave, "#3B82F6"))
        stats_row.addWidget(make_stat_card("Absent",      n_absent,      "#EF4444"))
        stats_row.addStretch()
        root.addLayout(stats_row)

        # Bug 4 / Rec 6: inline auto-email notice
        if self.auto_email_ok is not None:
            if self.auto_email_ok:
                _n_txt = "✅  Auto-email sent successfully."
                _n_bg, _n_bd, _n_fg = "#F0FDF4", "#86EFAC", "#166534"
            else:
                _n_txt = "⚠  Auto-email failed — use Send Email below to retry."
                _n_bg, _n_bd, _n_fg = "#FFFBEB", "#FCD34D", "#92400E"
            _notice = QLabel(_n_txt)
            _notice.setStyleSheet(
                f"background:{_n_bg}; border:1px solid {_n_bd}; border-radius:8px;"
                f"padding:10px 16px; color:{_n_fg}; font-size:12px; font-weight:600;"
            )
            root.addWidget(_notice)

        tbl = make_table(
            ["Student ID", "Full Name", "First seen", "Last seen", "Status", "Note"],
            rows,
            col_widths={0: 140, 2: 100, 3: 100, 4: 130},
            stretch_col=5,
        )
        for r, (sid, name, first, last, status, note) in enumerate(rows):
            tbl.setItem(r, 0, QTableWidgetItem(sid))
            tbl.setItem(r, 1, QTableWidgetItem(name))
            tbl.setItem(r, 2, QTableWidgetItem(first))
            tbl.setItem(r, 3, QTableWidgetItem(last))
            tbl.setCellWidget(r, 4, make_badge(status))
            tbl.setItem(r, 5, QTableWidgetItem(note))
            tbl.setRowHeight(r, 44)
        root.addWidget(tbl)

        actions_row = QHBoxLayout()
        actions_row.setSpacing(20)
        for card_title, desc, btn_txt in [
            ("Export report", "Download as PDF or Excel", "Export"),
            ("Send by email", "Send report to your university email", "Send email"),
        ]:
            card = QFrame()
            card.setStyleSheet(
                f"QFrame {{ background:{WHITE}; border:1px solid {BORDER}; border-radius:10px; }}"
            )
            card.setFixedHeight(110)
            c_lay = QVBoxLayout(card)
            c_lay.setContentsMargins(22, 16, 22, 16)
            c_lay.setSpacing(4)
            ct = QLabel(card_title)
            ct.setStyleSheet(
                f"font-size:14px; font-weight:700; color:{TEXT_DARK}; border:none; background:transparent;"
            )
            cd = QLabel(desc)
            cd.setStyleSheet(
                f"font-size:12px; color:{TEXT_GRAY}; border:none; background:transparent;"
            )
            cb = QPushButton(btn_txt)
            cb.setFixedWidth(110)
            cb.setFixedHeight(34)
            cb.setStyleSheet(
                f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
                f"font-size:12px; font-weight:600;"
            )
            cb.setCursor(Qt.PointingHandCursor)
            if btn_txt == "Export":
                cb.clicked.connect(self._export_report)
            elif btn_txt == "Send email":
                cb.clicked.connect(self._send_email)
            c_lay.addWidget(ct)
            c_lay.addWidget(cd)
            c_lay.addSpacing(4)
            c_lay.addWidget(cb, alignment=Qt.AlignLeft)
            actions_row.addWidget(card)

        pdf_btn = QPushButton("📄  Export PDF")
        pdf_btn.setStyleSheet(
            f"QPushButton {{ background:#DC2626; color:white; border-radius:6px; "
            f"padding:8px 16px; font-size:12px; font-weight:600; }}"
            f"QPushButton:hover {{ background:#B91C1C; color:white; }}"
        )
        pdf_btn.setCursor(Qt.PointingHandCursor)
        pdf_btn.clicked.connect(self._export_pdf)
        actions_row.addWidget(pdf_btn)
        actions_row.addStretch()
        root.addLayout(actions_row)

        back_row = QHBoxLayout()
        back_row.addStretch()
        back_btn = QPushButton("← Back to Home")
        back_btn.setFixedHeight(40)
        back_btn.setStyleSheet(
            f"background:{WHITE}; color:{PRIMARY}; border:1px solid {PRIMARY}; border-radius:6px;"
            "padding:8px 16px; font-size:13px; font-weight:600;"
        )
        back_btn.setCursor(Qt.PointingHandCursor)
        back_btn.clicked.connect(self._back_to_home)
        back_row.addWidget(back_btn)
        root.addLayout(back_row)

    def _export_report(self):
        course_title = self.course_data.get("CourseTitle", "Attendance") if self.course_data else "Attendance"
        default_name = f"Attendance_{course_title}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        path, selected_filter = QFileDialog.getSaveFileName(
            self, "Save Report", default_name,
            "Excel Files (*.xlsx);;PDF Files (*.pdf)"
        )
        if not path:
            return

        if selected_filter == "PDF Files (*.pdf)" or path.lower().endswith(".pdf"):
            self._export_pdf_qt(path, course_title)
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Report"

            green_fill = PatternFill("solid", start_color="1B5E35")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            center = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("A1:F1")
            ws["A1"] = f"Attendance Report — {course_title}"
            ws["A1"].font = Font(bold=True, size=14, color="1A1A2E")
            ws["A1"].alignment = center

            if self.course_data:
                sched = self.course_data.get("Schedule", "")
                if isinstance(sched, dict):
                    sched_str = f"{sched.get('Day','')} {sched.get('StartTime','')}–{sched.get('EndTime','')}"
                else:
                    sched_str = str(sched)
                ws.merge_cells("A2:F2")
                ws["A2"] = f"Date: {datetime.now().strftime('%Y-%m-%d')}  |  Schedule: {sched_str}"
                ws["A2"].alignment = center

            ws.append([])

            headers = ["Student ID", "Full Name", "First Seen", "Last Seen", "Status", "Note"]
            ws.append(headers)
            header_row = ws.max_row
            for col in range(1, 7):
                cell = ws.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = green_fill
                cell.alignment = center

            status_colors = {
                "Present":     "EAF3DE",
                "Late":        "FAEEDA",
                "Absent":      "FCEBEB",
                "Early Leave": "E6F1FB",
            }

            for stu in self.enrolled_students:
                sid = str(stu.get("StudentID", ""))
                name = stu.get("FullName", "")
                rec = self.attendance.get(sid, {})
                status = rec.get("status", "Absent")
                first_seen = rec.get("first_seen", "—")
                last_seen = rec.get("last_seen", "—")
                note = rec.get("note", "")

                ws.append([sid, name, first_seen, last_seen, status, note])
                row = ws.max_row
                status_fill = PatternFill("solid", start_color=status_colors.get(status, "FFFFFF"))
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = center
                    if col == 5:
                        cell.fill = status_fill

            ws.column_dimensions["A"].width = 16
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 14
            ws.column_dimensions["F"].width = 28

            wb.save(path)
            QMessageBox.information(self, "Export successful", f"Report saved to:\n{path}")

        except Exception as e:
            QMessageBox.critical(self, "Export failed", f"Error: {e}")

    def _export_pdf(self):
        from PyQt5.QtWidgets import QFileDialog, QMessageBox
        import datetime, os

        path, _ = QFileDialog.getSaveFileName(
            self, "Save PDF Report",
            f"AIAS_Report_{datetime.date.today()}.pdf",
            "PDF Files (*.pdf)"
        )
        if not path:
            return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                            Table, TableStyle, HRFlowable)
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

            doc = SimpleDocTemplate(
                path,
                pagesize=A4,
                rightMargin=2*cm,
                leftMargin=2*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )

            story = []
            styles = getSampleStyleSheet()

            # --- HEADER ---
            header_style = ParagraphStyle(
                'Header',
                fontSize=18,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#1B5E35'),
                alignment=TA_CENTER,
                spaceAfter=4
            )
            sub_style = ParagraphStyle(
                'Sub',
                fontSize=11,
                fontName='Helvetica',
                textColor=colors.HexColor('#6B7280'),
                alignment=TA_CENTER,
                spaceAfter=2
            )
            label_style = ParagraphStyle(
                'Label',
                fontSize=11,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#1A1A2E'),
                spaceAfter=4
            )
            normal_style = ParagraphStyle(
                'Normal2',
                fontSize=10,
                fontName='Helvetica',
                textColor=colors.HexColor('#374151'),
                spaceAfter=2
            )

            story.append(Paragraph("Qassim University", header_style))
            story.append(Paragraph("College of Computer Science", sub_style))
            story.append(Paragraph("AI Attendance System — Session Report", sub_style))
            story.append(Spacer(1, 0.3*cm))
            story.append(HRFlowable(width="100%", thickness=2,
                                    color=colors.HexColor('#1B5E35')))
            story.append(Spacer(1, 0.4*cm))

            # --- COURSE INFO ---
            import datetime as dt
            try:
                course_title    = self.course_data.get("CourseTitle", "N/A") if self.course_data else "N/A"
                section_id      = self.course_data.get("SectionID", "N/A") if self.course_data else "N/A"
                instructor_name = (self.main_win.instructor_data.get("fullname", "N/A")
                                   if self.main_win else "N/A")
                session_date    = str(dt.date.today())
                sched_raw       = self.course_data.get("Schedule", {}) if self.course_data else {}
                if isinstance(sched_raw, dict):
                    start_time = sched_raw.get("StartTime", "")
                    end_time   = sched_raw.get("EndTime", "")
                else:
                    start_time = end_time = str(sched_raw)
                records = []
                for stu in self.enrolled_students:
                    sid  = str(stu.get("StudentID", ""))
                    name = stu.get("FullName", "")
                    att  = self.attendance.get(sid, {})
                    records.append({
                        "StudentID":    sid,
                        "FullName":     name,
                        "Status":       att.get("status", "Absent"),
                        "RecognizedAt": att.get("first_seen", ""),
                    })
            except Exception:
                course_title = section_id = instructor_name = "N/A"
                session_date = str(dt.date.today())
                start_time = end_time = ""
                records = []

            info_data = [
                ["Course:", course_title,   "Section:", section_id],
                ["Instructor:", instructor_name, "Date:", session_date],
                ["Start Time:", start_time,  "End Time:", end_time],
            ]
            info_table = Table(info_data, colWidths=[3*cm, 6*cm, 3*cm, 5*cm])
            info_table.setStyle(TableStyle([
                ('FONTNAME',  (0,0), (-1,-1), 'Helvetica'),
                ('FONTNAME',  (0,0), (0,-1),  'Helvetica-Bold'),
                ('FONTNAME',  (2,0), (2,-1),  'Helvetica-Bold'),
                ('FONTSIZE',  (0,0), (-1,-1), 10),
                ('TEXTCOLOR', (0,0), (-1,-1), colors.HexColor('#1A1A2E')),
                ('ROWBACKGROUNDS', (0,0), (-1,-1),
                 [colors.HexColor('#F8F9FA'), colors.white]),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('TOPPADDING',    (0,0), (-1,-1), 6),
            ]))
            story.append(info_table)
            story.append(Spacer(1, 0.5*cm))

            # --- SUMMARY STATS ---
            present = sum(1 for r in records if r.get("Status") in ("Present", "Late"))
            absent  = sum(1 for r in records if r.get("Status") == "Absent")
            late    = sum(1 for r in records if r.get("Status") == "Late")
            total   = len(records)
            rate    = f"{present/total*100:.1f}%" if total > 0 else "0%"

            story.append(Paragraph("Session Summary", label_style))
            summary_data = [
                ["Total Students", "Present", "Late", "Absent", "Attendance Rate"],
                [str(total), str(present), str(late), str(absent), rate]
            ]
            summary_table = Table(summary_data, colWidths=[3.4*cm]*5)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND',   (0,0), (-1,0), colors.HexColor('#1B5E35')),
                ('TEXTCOLOR',    (0,0), (-1,0), colors.white),
                ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTNAME',     (0,1), (-1,1), 'Helvetica-Bold'),
                ('FONTSIZE',     (0,0), (-1,-1), 11),
                ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
                ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
                ('ROWHEIGHT',    (0,0), (-1,-1), 24),
                ('BACKGROUND',   (0,1), (-1,1), colors.HexColor('#F0FDF4')),
                ('TEXTCOLOR',    (0,1), (-1,1), colors.HexColor('#1B5E35')),
                ('GRID',         (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
                ('ROUNDEDCORNERS', [4]),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.5*cm))

            # --- ATTENDANCE TABLE ---
            story.append(Paragraph("Attendance Details", label_style))

            table_data = [["#", "Student ID", "Full Name", "Status", "Time"]]
            for i, rec in enumerate(records, 1):
                status = rec.get("Status", "")
                table_data.append([
                    str(i),
                    rec.get("StudentID", ""),
                    rec.get("FullName", ""),
                    status,
                    rec.get("RecognizedAt", "")
                ])

            att_table = Table(table_data,
                              colWidths=[1*cm, 3.5*cm, 6*cm, 2.5*cm, 4*cm])
            att_table.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#1B5E35')),
                ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
                ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',      (0,0), (-1,-1), 9),
                ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('ROWHEIGHT',     (0,0), (-1,-1), 20),
                ('ROWBACKGROUNDS',(0,1), (-1,-1),
                 [colors.white, colors.HexColor('#F9FAFB')]),
                ('GRID',          (0,0), (-1,-1), 0.5,
                 colors.HexColor('#E5E7EB')),
                ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
            ]))

            # Color status column
            for i, rec in enumerate(records, 1):
                status = rec.get("Status", "")
                color = (
                    colors.HexColor('#16A34A') if status == "Present" else
                    colors.HexColor('#F59E0B') if status == "Late" else
                    colors.HexColor('#DC2626')
                )
                att_table.setStyle(TableStyle([
                    ('TEXTCOLOR', (3, i), (3, i), color),
                    ('FONTNAME',  (3, i), (3, i), 'Helvetica-Bold'),
                ]))

            story.append(att_table)
            story.append(Spacer(1, 1*cm))

            # --- FOOTER ---
            footer_style = ParagraphStyle(
                'Footer',
                fontSize=8,
                fontName='Helvetica',
                textColor=colors.HexColor('#9CA3AF'),
                alignment=TA_CENTER
            )
            story.append(HRFlowable(width="100%", thickness=0.5,
                                    color=colors.HexColor('#E5E7EB')))
            story.append(Spacer(1, 0.2*cm))
            story.append(Paragraph(
                f"Generated by AIAS — AI Attendance System  ·  "
                f"{dt.datetime.now().strftime('%Y-%m-%d %H:%M')}  ·  "
                f"Qassim University",
                footer_style
            ))

            doc.build(story)
            QMessageBox.information(self, "Success",
                                    f"PDF report saved to:\n{path}")

        except ImportError:
            QMessageBox.warning(self, "Missing Library",
                                "reportlab is required.\nRun: pip install reportlab")
        except Exception as e:
            QMessageBox.warning(self, "Export Failed", f"Error:\n{e}")

    def _export_pdf_qt(self, path, course_title):
        try:
            from PyQt5.QtPrintSupport import QPrinter
            from PyQt5.QtGui import QTextDocument

            sched_raw = self.course_data.get("Schedule", "") if self.course_data else ""
            if isinstance(sched_raw, dict):
                sched_str = f"{sched_raw.get('Day','')} {sched_raw.get('StartTime','')}–{sched_raw.get('EndTime','')}"
            else:
                sched_str = str(sched_raw)

            status_colors_html = {
                "Present":     "#EAF3DE", "Late":        "#FAEEDA",
                "Absent":      "#FCEBEB", "Early Leave":  "#E6F1FB",
            }

            rows_html = ""
            for stu in self.enrolled_students:
                sid  = str(stu.get("StudentID", ""))
                name = stu.get("FullName", "")
                rec  = self.attendance.get(sid, {})
                status     = rec.get("status", "Absent")
                first_seen = rec.get("first_seen", "—")
                last_seen  = rec.get("last_seen",  "—")
                note       = rec.get("note", "")
                bg = status_colors_html.get(status, "#FFFFFF")
                rows_html += (
                    f"<tr>"
                    f"<td>{sid}</td><td>{name}</td>"
                    f"<td>{first_seen}</td><td>{last_seen}</td>"
                    f"<td style='background:{bg};'>{status}</td>"
                    f"<td>{note}</td>"
                    f"</tr>"
                )

            html = f"""
            <html><body style='font-family:Arial,sans-serif; font-size:12px;'>
            <h2 style='color:#1A1A2E;'>Attendance Report — {course_title}</h2>
            <p style='color:#6B7280;'>Date: {datetime.now().strftime('%Y-%m-%d')} &nbsp;|&nbsp; Schedule: {sched_str}</p>
            <table border='1' cellspacing='0' cellpadding='6' width='100%'
                   style='border-collapse:collapse; border-color:#E5E7EB;'>
              <thead>
                <tr style='background:#1B5E35; color:white;'>
                  <th>Student ID</th><th>Full Name</th>
                  <th>First Seen</th><th>Last Seen</th><th>Status</th><th>Note</th>
                </tr>
              </thead>
              <tbody>{rows_html}</tbody>
            </table>
            </body></html>
            """

            doc = QTextDocument()
            doc.setHtml(html)

            printer = QPrinter()
            printer.setOutputFileName(path)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setPageMargins(15, 15, 15, 15, QPrinter.Millimeter)
            doc.print_(printer)

            QMessageBox.information(self, "Export successful", f"PDF saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "PDF Export failed", f"Error: {e}")

    def _send_email(self):
        to_email = ""
        if DB_CONNECTED:
            try:
                username = self.main_win.instructor_data.get("username", "")
                instr = db["Instructors"].find_one({"Username": username})
                if instr:
                    to_email = instr.get("UniversityEmail", "")
            except Exception as e:
                print(f"[AIAS] Email credentials load error: {e}")

        if to_email:
            self._do_send_email(to_email)
            return

        # First-time setup dialog (only asks for destination email)
        dlg = QDialog(self)
        dlg.setWindowTitle("Email Setup")
        dlg.setFixedWidth(420)
        dlg.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        t = QLabel("Email Setup")
        t.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(t)
        sub = QLabel("Enter the university email address to send reports to.")
        sub.setStyleSheet(f"font-size:12px; color:{TEXT_GRAY};")
        sub.setWordWrap(True)
        lay.addWidget(sub)

        to_lbl = QLabel("Send reports to:")
        to_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        to_field = QLineEdit()
        to_field.setPlaceholderText("name@university.edu")
        to_field.setFixedHeight(38)
        to_field.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:6px; padding:6px 10px; background:{WHITE}; color:{TEXT_DARK};"
        )
        lay.addWidget(to_lbl)
        lay.addWidget(to_field)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.clicked.connect(dlg.reject)
        send_btn = QPushButton("Save & Send")
        send_btn.setStyleSheet(
            "background:#1B5E35; color:white; border-radius:6px; padding:8px 20px; font-weight:700;"
        )
        send_btn.clicked.connect(dlg.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(send_btn)
        lay.addLayout(btn_row)

        if dlg.exec_() != QDialog.Accepted:
            return

        to_email = to_field.text().strip()
        if not to_email:
            QMessageBox.warning(self, "Missing info", "Please enter a destination email address.")
            return

        if DB_CONNECTED:
            try:
                username = self.main_win.instructor_data.get("username", "")
                db["Instructors"].update_one(
                    {"Username": username},
                    {"$set": {"UniversityEmail": to_email}},
                )
            except Exception as e:
                print(f"[AIAS] Failed to save email credentials: {e}")

        self._do_send_email(to_email)

    def _do_send_email(self, to_email):
        course_title = self.course_data.get("CourseTitle", "Attendance") if self.course_data else "Attendance"
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Report"

            green_fill  = PatternFill("solid", start_color="1B5E35")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            center      = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("A1:F1")
            ws["A1"] = f"Attendance Report — {course_title}"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = center

            ws.append([])
            ws.append(["Student ID", "Full Name", "First Seen", "Last Seen", "Status", "Note"])
            for col in range(1, 7):
                cell = ws.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = green_fill
                cell.alignment = center

            status_colors = {
                "Present": "EAF3DE", "Late": "FAEEDA",
                "Absent": "FCEBEB", "Early Leave": "E6F1FB",
            }
            for stu in self.enrolled_students:
                sid  = str(stu.get("StudentID", ""))
                rec  = self.attendance.get(sid, {})
                status     = rec.get("status",     "Absent")
                first_seen = rec.get("first_seen", "—")
                last_seen  = rec.get("last_seen",  "—")
                note       = rec.get("note",       "")
                ws.append([sid, stu.get("FullName", ""), first_seen, last_seen, status, note])
                ws.cell(row=ws.max_row, column=5).fill = PatternFill(
                    "solid", start_color=status_colors.get(status, "FFFFFF")
                )
            for col, w in zip("ABCDEF", [16, 28, 14, 14, 14, 28]):
                ws.column_dimensions[col].width = w
            wb.save(tmp.name)

            n_present    = sum(1 for r in self.attendance.values() if r.get("status") == "Present")
            n_late       = sum(1 for r in self.attendance.values() if r.get("status") == "Late")
            n_early      = sum(1 for r in self.attendance.values() if r.get("status") == "Early Leave")
            n_absent     = (
                sum(1 for r in self.attendance.values() if r.get("status") == "Absent")
                + (len(self.enrolled_students) - len(self.attendance))
            )

            msg = MIMEMultipart()
            msg["From"]    = config.SMTP_FROM
            msg["To"]      = to_email
            msg["Subject"] = (
                f"Attendance Report — {course_title} — "
                f"{datetime.now().strftime('%Y-%m-%d')}"
            )
            body = (
                f"Dear Instructor,\n\n"
                f"Please find attached the attendance report for {course_title}.\n\n"
                f"Date: {datetime.now().strftime('%Y-%m-%d')}\n"
                f"Total students: {len(self.enrolled_students)}\n"
                f"Present: {n_present}  Late: {n_late}  "
                f"Early Leave: {n_early}  Absent: {n_absent}\n\n"
                f"Best regards,\nAIAS — AI Attendance System"
            )
            msg.attach(MIMEText(body, "plain"))

            with open(tmp.name, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            safe = "".join(c for c in course_title if c.isalnum() or c in "-_ ")
            part.add_header(
                "Content-Disposition", f'attachment; filename="Attendance_{safe}.xlsx"'
            )
            msg.attach(part)

            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()
                server.login(config.SMTP_LOGIN, config.SMTP_PASSWORD)
                server.send_message(msg)

            QMessageBox.information(self, "Email sent", f"Report sent successfully to:\n{to_email}")

        except Exception as e:
            QMessageBox.critical(
                self, "Email failed",
                f"Error: {str(e)}\n\nMake sure you are using a valid Brevo SMTP Key.\n"
                "To update your settings, use Email Settings in the sidebar.",
            )
        finally:
            try:
                os.unlink(tmp.name)
            except Exception as _e:
                print(f"[AIAS] error: {_e}")

    def _back_to_home(self):
        self.hide()
        self.main_win.show()


class AdminPanelWindow(QWidget):
    def __init__(self, embedded=False):
        super().__init__()
        self._embedded = embedded
        self._embedding_worker = None
        self._pending_student  = None
        self._pending_dest     = None
        if not embedded:
            self.setWindowTitle("AIAS — Admin Panel")
            import os
            from PyQt5.QtGui import QIcon
            _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aias_icon.ico")
            if os.path.exists(_icon_path):
                self.setWindowIcon(QIcon(_icon_path))
            self.setMinimumSize(1100, 700)
            self.resize(1100, 700)
        self.setStyleSheet(f"background:{BG};")
        self._build()

    def _build(self):
        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        sidebar, s_lay = make_sidebar_base("AD", "Admin", "System administrator")

        nav_lbl = QLabel("NAVIGATION")
        nav_lbl.setStyleSheet(
            f"font-size:10px; font-weight:700; color:{TEXT_GRAY}; letter-spacing:1px;"
            f"background:transparent; border:none;"
        )
        s_lay.addWidget(nav_lbl)
        s_lay.addSpacing(4)

        self._nav_btns = []
        for i, label in enumerate(["Students", "Courses", "Instructors", "Settings"]):
            btn = QPushButton(label)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setCheckable(False)
            btn.clicked.connect(lambda _, idx=i: self._switch_page(idx))
            self._nav_btns.append(btn)
            s_lay.addWidget(btn)

        self._apply_nav(0)

        s_lay.addStretch()

        dark_btn = QPushButton("🌙  Dark Mode" if not IS_DARK_MODE else "☀  Light Mode")
        dark_btn.setStyleSheet(
            f"background:{'#2A3D35' if IS_DARK_MODE else PRIMARY_LIGHT}; "
            f"color:{'#4ecf8e' if IS_DARK_MODE else PRIMARY}; "
            "border-radius:6px; padding:8px 12px; font-size:12px; font-weight:600;"
        )
        dark_btn.setCursor(Qt.PointingHandCursor)
        dark_btn.clicked.connect(lambda: self._toggle_dark_mode(dark_btn))
        s_lay.addWidget(dark_btn)
        s_lay.addSpacing(4)

        logout_btn = QPushButton("Logout")
        logout_btn.setToolTip("Sign out and return to login screen")
        logout_btn.setStyleSheet(
            "QPushButton { background:#DC2626; color:white; border-radius:6px; "
            "padding:8px 12px; font-size:12px; font-weight:700; }"
            "QPushButton:hover { background:#B91C1C; color:white; }"
            "QPushButton:pressed { background:#991B1B; color:white; }"
        )
        logout_btn.setCursor(Qt.PointingHandCursor)
        logout_btn.clicked.connect(self._logout)
        s_lay.addWidget(logout_btn)

        self._stack = QStackedWidget()
        self._stack.addWidget(self._build_students_page())
        self._stack.addWidget(self._build_courses_page())
        self._stack.addWidget(self._build_instructors_page())
        self._stack.addWidget(self._build_settings_page())

        root.addWidget(sidebar)
        root.addWidget(self._stack)

    def _toggle_dark_mode(self, btn):
        new_dark = not IS_DARK_MODE
        apply_theme(QApplication.instance(), dark=new_dark)
        btn.setText("☀  Light Mode" if new_dark else "🌙  Dark Mode")
        btn.setStyleSheet(
            f"QPushButton {{ background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px; "
            f"padding:8px 12px; font-size:12px; font-weight:600; }}"
            f"QPushButton:hover {{ background:{PRIMARY}; color:#ffffff; }}"
            f"QPushButton:pressed {{ background:{PRIMARY}; color:#ffffff; padding:9px 11px 7px 13px; }}"
        )
        for widget in QApplication.instance().topLevelWidgets():
            force_theme_on_all_widgets(widget)
            widget.setStyleSheet(f"background-color:{BG};")
            widget.update()
            widget.repaint()

    def _apply_nav(self, active_idx):
        for i, btn in enumerate(self._nav_btns):
            if i == active_idx:
                btn.setStyleSheet(
                    f"background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px;"
                    f"padding:8px 12px; text-align:left; font-weight:600;"
                )
            else:
                btn.setStyleSheet(
                    f"background:transparent; color:{TEXT_MED}; border-radius:6px;"
                    f"padding:8px 12px; text-align:left; font-weight:400;"
                )

    def _switch_page(self, idx):
        self._stack.setCurrentIndex(idx)
        self._apply_nav(idx)

    def _refresh_stack_page(self, idx, build_fn):
        old = self._stack.widget(idx)
        new = build_fn()
        self._stack.insertWidget(idx, new)
        if old is not None:
            self._stack.removeWidget(old)
            old.deleteLater()
        self._stack.setCurrentIndex(idx)

    def _page_header(self, title, subtitle, show_add_student=False, show_add_course=False):
        hdr = QHBoxLayout()
        col = QVBoxLayout()
        col.setSpacing(2)
        t = QLabel(title)
        t.setStyleSheet(f"font-size:20px; font-weight:800; color:{TEXT_DARK};")
        s = QLabel(subtitle)
        s.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY};")
        col.addWidget(t)
        col.addWidget(s)
        hdr.addLayout(col)
        hdr.addStretch()
        if show_add_student:
            action_btn = QPushButton("+ Add Student")
            action_btn.clicked.connect(self._add_student)
        elif show_add_course:
            action_btn = QPushButton("+ Add Course")
            action_btn.clicked.connect(self._add_course)
        else:
            action_btn = QPushButton("+ Import Excel")
            action_btn.clicked.connect(self._import_excel)
        action_btn.setFixedHeight(36)
        action_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px; padding:6px 16px;"
        )
        action_btn.setCursor(Qt.PointingHandCursor)
        hdr.addWidget(action_btn)
        return hdr

    def _add_student(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Student")
        dialog.setFixedWidth(440)
        dialog.setStyleSheet(f"background:{BG};")

        lay = QVBoxLayout(dialog)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        title_lbl = QLabel("Add New Student")
        title_lbl.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title_lbl)

        id_lbl = QLabel("Student ID")
        id_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        id_field = QLineEdit()
        id_field.setPlaceholderText("e.g. 431109377")
        id_field.setFixedHeight(38)
        lay.addWidget(id_lbl)
        lay.addWidget(id_field)

        name_lbl = QLabel("Full Name")
        name_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        name_field = QLineEdit()
        name_field.setPlaceholderText("e.g. Mohammed Al-Harbi")
        name_field.setFixedHeight(38)
        lay.addWidget(name_lbl)
        lay.addWidget(name_field)

        courses_lbl = QLabel("Enroll in Courses")
        courses_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        lay.addWidget(courses_lbl)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(160)
        scroll.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:6px; background:{WHITE};"
        )
        scroll_widget = QWidget()
        scroll_widget.setStyleSheet(f"background:{WHITE};")
        scroll_lay = QVBoxLayout(scroll_widget)
        scroll_lay.setContentsMargins(10, 8, 10, 8)
        scroll_lay.setSpacing(6)

        checkboxes = []
        courses = []
        if DB_CONNECTED:
            try:
                courses = list(db["Courses"].find({}))
            except Exception as _e:
                print(f"[AIAS] error: {_e}")

        for course in courses:
            section_id = str(course.get("SectionID", ""))
            course_title = course.get("CourseTitle", "")
            cb = QCheckBox(f"{section_id} — {course_title}")
            cb.setStyleSheet(f"color:{TEXT_DARK}; font-size:12px; background:transparent;")
            checkboxes.append((cb, course))
            scroll_lay.addWidget(cb)

        if not courses:
            no_lbl = QLabel("No courses found")
            no_lbl.setStyleSheet(f"color:{TEXT_GRAY}; font-size:12px; background:transparent;")
            scroll_lay.addWidget(no_lbl)

        scroll_lay.addStretch()
        scroll.setWidget(scroll_widget)
        lay.addWidget(scroll)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(dialog.reject)
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:8px 20px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(dialog.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        if dialog.exec_() != QDialog.Accepted:
            return

        student_id = id_field.text().strip()
        full_name = name_field.text().strip()

        if not student_id or not full_name:
            QMessageBox.warning(self, "Missing Fields", "Please fill in Student ID and Full Name.")
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        try:
            if db["Students"].find_one({"StudentID": student_id}):
                QMessageBox.warning(self, "Duplicate ID", f"Student ID '{student_id}' already exists.")
                return
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        selected_sections = [
            str(course.get("SectionID", ""))
            for cb, course in checkboxes
            if cb.isChecked()
        ]

        try:
            db["Students"].insert_one({
                "StudentID": student_id,
                "FullName": full_name,
                "EnrolledSections": selected_sections,
            })
        except Exception as e:
            QMessageBox.critical(self, "DB Error", f"Could not insert student:\n{e}")
            return

        for cb, course in checkboxes:
            if cb.isChecked():
                try:
                    db["Courses"].update_one(
                        {"_id": course["_id"]},
                        {"$addToSet": {"EnrolledStudents": student_id}},
                    )
                except Exception as e:
                    print(f"DB error updating course: {e}")

        QMessageBox.information(self, "Success", f"Student '{full_name}' added successfully.")

        self._refresh_stack_page(0, self._build_students_page)

    def _build_students_page(self):
        page = QWidget()
        page.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(page)
        lay.setContentsMargins(32, 28, 32, 24)
        lay.setSpacing(18)

        hdr_row = QHBoxLayout()
        hdr_row.addLayout(self._page_header("Students", "Manage student photos and embeddings", show_add_student=True))
        batch_btn = QPushButton("⚡ Batch Embed Folder")
        batch_btn.setFixedHeight(36)
        batch_btn.setStyleSheet(
            f"background:#6366F1; color:white; border-radius:6px; padding:6px 16px; font-weight:600;"
        )
        batch_btn.setCursor(Qt.PointingHandCursor)
        batch_btn.clicked.connect(self._batch_embed_folder)
        hdr_row.addWidget(batch_btn)
        lay.addLayout(hdr_row)

        students = []
        if DB_CONNECTED:
            try:
                students = list(db["Students"].find({}))
            except Exception as e:
                print(f"[AIAS] Students page load error: {e}")

        total = len(students)
        ready = sum(1 for s in students if s.get("FaceEmbedding"))
        missing = total - ready

        sr = QHBoxLayout()
        sr.setSpacing(16)
        sr.addWidget(make_stat_card("Total", total, PRIMARY))
        sr.addWidget(make_stat_card("Photo uploaded", ready, "#22C55E"))
        sr.addWidget(make_stat_card("Missing photo", missing, "#EF4444"))
        sr.addStretch()
        lay.addLayout(sr)

        # Build course → enrolled student IDs map for filtering
        course_enrolled = {}   # {"CS101 — Title": ["id1", "id2", ...]}
        if DB_CONNECTED:
            try:
                for c in db["Courses"].find({}):
                    key = f"{c.get('SectionID', '')} — {c.get('CourseTitle', '')}"
                    course_enrolled[key] = [str(x) for x in c.get("EnrolledStudents", [])]
            except Exception as e:
                print(f"[AIAS] Students page course filter load error: {e}")

        combo = QComboBox()
        combo.setFixedWidth(240)
        combo.setFixedHeight(34)
        combo.addItem("All Students")
        for key in course_enrolled:
            combo.addItem(key)
        if combo.count() == 1 and not course_enrolled:
            combo.addItem("No courses found")
        lay.addWidget(combo)

        tbl = make_table(
            ["Student ID", "Full Name", "Face", "Status", "Action", "Delete"],
            students,
            col_widths={0: 140, 2: 80, 3: 120, 4: 180, 5: 100},
            stretch_col=1,
        )
        for r, stu in enumerate(students):
            has_photo = bool(stu.get("FaceEmbedding"))
            n_photos  = len(stu.get("ImagePaths", []))
            status = "Ready" if has_photo else "No photo"
            tbl.setItem(r, 0, QTableWidgetItem(str(stu.get("StudentID", ""))))
            tbl.setItem(r, 1, QTableWidgetItem(stu.get("FullName", "")))

            ph = QLabel((f"✓ {n_photos}") if has_photo else "✗")
            ph.setAlignment(Qt.AlignCenter)
            ph.setStyleSheet(
                f"color:{'#22C55E' if has_photo else '#EF4444'}; font-size:14px; font-weight:700;"
                f"background:transparent; border:none;"
            )
            tbl.setCellWidget(r, 2, ph)
            tbl.setCellWidget(r, 3, make_badge(status))

            action_lbl = "View" if has_photo else "Upload"
            action_btn = QPushButton(action_lbl)
            action_btn.setToolTip("Upload or view student photo for face recognition")
            action_btn.setCursor(Qt.PointingHandCursor)
            if has_photo:
                action_btn.setStyleSheet(
                    "background:#EAF3DE; color:#27500A; border-radius:5px;"
                    "padding:4px 10px; font-size:12px; font-weight:600;"
                )
                action_btn.clicked.connect(lambda _, s=stu: self._view_photo(s))
            else:
                action_btn.setStyleSheet(
                    f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:5px;"
                    "padding:4px 10px; font-size:12px;"
                )
                action_btn.clicked.connect(lambda _, s=stu: self._upload_photo(s))
            edit_btn = QPushButton("Edit")
            edit_btn.setToolTip("Edit student information")
            edit_btn.setCursor(Qt.PointingHandCursor)
            edit_btn.setStyleSheet(
                f"background:#E6F1FB; color:#0C447C; border-radius:5px;"
                "padding:4px 10px; font-size:12px; font-weight:600;"
            )
            edit_btn.clicked.connect(lambda _, s=stu: self._edit_student(s))
            wrapper = QWidget()
            wrapper.setStyleSheet("background:transparent;")
            wl = QHBoxLayout(wrapper)
            wl.setContentsMargins(6, 4, 6, 4)
            wl.setSpacing(6)
            wl.addWidget(action_btn, alignment=Qt.AlignCenter)
            wl.addWidget(edit_btn, alignment=Qt.AlignCenter)
            tbl.setCellWidget(r, 4, wrapper)

            del_btn = QPushButton("Delete")
            del_btn.setToolTip("Remove this student permanently")
            del_btn.setCursor(Qt.PointingHandCursor)
            del_btn.setStyleSheet(
                "background:#FCEBEB; color:#791F1F; border-radius:5px;"
                "padding:4px 12px; font-size:12px; font-weight:600;"
            )
            sid = str(stu.get("StudentID", ""))
            fname = stu.get("FullName", "")
            del_btn.clicked.connect(lambda _, s=sid, n=fname: self._delete_student(s, n))
            del_wrapper = QWidget()
            del_wrapper.setStyleSheet("background:transparent;")
            dl = QHBoxLayout(del_wrapper)
            dl.setContentsMargins(6, 4, 6, 4)
            dl.addWidget(del_btn, alignment=Qt.AlignCenter)
            tbl.setCellWidget(r, 5, del_wrapper)
            tbl.setRowHeight(r, 44)

        def _filter_students(index):
            selected = combo.currentText()
            if selected == "All Students" or selected == "No courses found":
                for row in range(tbl.rowCount()):
                    tbl.setRowHidden(row, False)
            else:
                allowed = set(course_enrolled.get(selected, []))
                for row in range(tbl.rowCount()):
                    sid = tbl.item(row, 0).text() if tbl.item(row, 0) else ""
                    tbl.setRowHidden(row, sid not in allowed)

        combo.currentIndexChanged.connect(_filter_students)

        search_row = QHBoxLayout()
        search_field = QLineEdit()
        search_field.setPlaceholderText("🔍  Search by name or student ID...")
        search_field.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:8px; padding:8px 14px; "
            f"font-size:13px; background:{WHITE}; color:{TEXT_DARK};"
        )
        search_field.setFixedHeight(38)
        search_row.addWidget(search_field)
        search_row.addStretch()
        lay.addLayout(search_row)

        def filter_students(text):
            text = text.lower().strip()
            for row in range(tbl.rowCount()):
                match = False
                for col in [0, 1]:
                    item = tbl.item(row, col)
                    if item and text in item.text().lower():
                        match = True
                        break
                tbl.setRowHidden(row, not match if text else False)

        search_field.textChanged.connect(filter_students)

        lay.addWidget(tbl)
        return page

    def _edit_student(self, student):
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Student")
        dialog.setFixedWidth(440)
        dialog.setStyleSheet(f"background:{BG};")

        lay = QVBoxLayout(dialog)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        title_lbl = QLabel("Edit Student")
        title_lbl.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title_lbl)

        name_lbl = QLabel("Full Name")
        name_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        name_field = QLineEdit()
        name_field.setText(student.get("FullName", ""))
        name_field.setFixedHeight(38)
        lay.addWidget(name_lbl)
        lay.addWidget(name_field)

        courses_lbl = QLabel("Enroll in Courses")
        courses_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        lay.addWidget(courses_lbl)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(160)
        scroll.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:6px; background:{WHITE};"
        )
        scroll_widget = QWidget()
        scroll_widget.setStyleSheet(f"background:{WHITE};")
        scroll_lay = QVBoxLayout(scroll_widget)
        scroll_lay.setContentsMargins(10, 8, 10, 8)
        scroll_lay.setSpacing(6)

        enrolled = [str(e) for e in student.get("EnrolledSections", [])]
        checkboxes = []
        courses = []
        if DB_CONNECTED:
            try:
                courses = list(db["Courses"].find({}))
            except Exception as e:
                print(f"[AIAS] Edit student courses load error: {e}")

        for course in courses:
            section_id = str(course.get("SectionID", ""))
            course_title = course.get("CourseTitle", "")
            cb = QCheckBox(f"{section_id} — {course_title}")
            cb.setStyleSheet(f"color:{TEXT_DARK}; font-size:12px; background:transparent;")
            cb.setChecked(section_id in enrolled)
            checkboxes.append((cb, course))
            scroll_lay.addWidget(cb)

        if not courses:
            no_lbl = QLabel("No courses found")
            no_lbl.setStyleSheet(f"color:{TEXT_GRAY}; font-size:12px; background:transparent;")
            scroll_lay.addWidget(no_lbl)

        scroll_lay.addStretch()
        scroll.setWidget(scroll_widget)
        lay.addWidget(scroll)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet(f"color:{BORDER};")
        lay.addWidget(sep)

        photo_lbl = QLabel("Add More Photos (optional)")
        photo_lbl.setStyleSheet(f"font-size:12px; color:{TEXT_MED};")
        lay.addWidget(photo_lbl)

        selected_paths = []

        photo_row = QHBoxLayout()
        select_btn = QPushButton("📷 Select Photos")
        select_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:6px 14px; font-weight:600;"
        )
        select_btn.setCursor(Qt.PointingHandCursor)
        count_lbl = QLabel("0 photos selected")
        count_lbl.setStyleSheet(f"font-size:12px; color:{TEXT_GRAY};")

        def _select_photos():
            paths, _ = QFileDialog.getOpenFileNames(
                dialog, "Select Photos (multiple allowed)", "", "Images (*.jpg *.jpeg *.png)"
            )
            if paths:
                selected_paths.clear()
                selected_paths.extend(paths)
                count_lbl.setText(f"{len(paths)} photo(s) selected")

        select_btn.clicked.connect(_select_photos)
        photo_row.addWidget(select_btn)
        photo_row.addWidget(count_lbl)
        photo_row.addStretch()
        lay.addLayout(photo_row)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(dialog.reject)
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:8px 20px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(dialog.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        if dialog.exec_() != QDialog.Accepted:
            return

        new_name = name_field.text().strip()
        if not new_name:
            QMessageBox.warning(self, "Missing Fields", "Full Name cannot be empty.")
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        student_id = str(student.get("StudentID", ""))
        selected_sections = [str(course.get("SectionID", "")) for cb, course in checkboxes if cb.isChecked()]

        try:
            db["Students"].update_one(
                {"StudentID": student_id},
                {"$set": {"FullName": new_name, "EnrolledSections": selected_sections}},
            )
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        try:
            db["Courses"].update_many({}, {"$pull": {"EnrolledStudents": student_id}})
            for cb, course in checkboxes:
                if cb.isChecked():
                    db["Courses"].update_one(
                        {"_id": course["_id"]},
                        {"$addToSet": {"EnrolledStudents": student_id}},
                    )
        except Exception as e:
            print(f"DB error syncing courses: {e}")

        if selected_paths:
            import cv2
            import numpy as np
            import shutil

            yolo, arcface, ai_ok = _ensure_ai_models(det_size=(640, 640))
            if ai_ok:
                new_embeddings = []
                folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", student_id)
                os.makedirs(folder, exist_ok=True)

                for photo_path in selected_paths:
                    img = cv2.imread(photo_path)
                    if img is None:
                        continue
                    img = apply_clahe(img)
                    results = yolo(img, verbose=False)
                    boxes_det = results[0].boxes
                    if len(boxes_det) == 0:
                        continue
                    x1, y1, x2, y2 = boxes_det[0].xyxy[0].cpu().numpy().astype(int)
                    if x2 <= x1 or y2 <= y1:
                        continue
                    face_crop = img[y1:y2, x1:x2]
                    face_crop = upscale_face_crop(face_crop, target_size=112)
                    faces = arcface.get(face_crop)
                    if not faces:
                        faces = arcface.get(img)
                    if not faces:
                        continue
                    raw = faces[0].embedding
                    new_embeddings.append(raw / np.linalg.norm(raw))

                if not new_embeddings:
                    QMessageBox.warning(
                        self, "No Face Detected",
                        "No face detected in selected photos. Photos not saved."
                    )
                else:
                    existing_files = [
                        f for f in os.listdir(folder)
                        if f.lower().endswith((".jpg", ".jpeg", ".png"))
                    ]
                    start_idx = len(existing_files) + 1
                    for i, photo_path in enumerate(selected_paths):
                        ext = os.path.splitext(photo_path)[1]
                        dest = os.path.join(folder, f"img{start_idx + i}{ext}")
                        shutil.copy2(photo_path, dest)

                    existing_emb = student.get("FaceEmbedding", [])
                    all_embeddings = list(new_embeddings)
                    if existing_emb:
                        arr = np.array(existing_emb, dtype=np.float32)
                        arr_norm = np.linalg.norm(arr)
                        if arr_norm > 0:
                            all_embeddings.insert(0, arr / arr_norm)

                    avg = np.mean(np.array(all_embeddings), axis=0)
                    avg = avg / np.linalg.norm(avg)

                    all_image_paths = [
                        os.path.abspath(os.path.join(folder, f))
                        for f in os.listdir(folder)
                        if f.lower().endswith((".jpg", ".jpeg", ".png"))
                    ]

                    try:
                        db["Students"].update_one(
                            {"StudentID": student_id},
                            {"$set": {
                                "FaceEmbedding": avg.tolist(),
                                "ImagePaths":    all_image_paths,
                            }},
                        )
                    except Exception as e:
                        print(f"[AIAS] Edit student embedding update error: {e}")

        self._refresh_stack_page(0, self._build_students_page)

    def _delete_student(self, student_id, full_name):
        import shutil
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete {full_name}?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        try:
            db["Students"].delete_one({"StudentID": student_id})
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        try:
            db["Courses"].update_many({}, {"$pull": {"EnrolledStudents": student_id}})
        except Exception as e:
            print(f"DB error removing student from courses: {e}")

        folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", student_id)
        if os.path.isdir(folder):
            try:
                shutil.rmtree(folder)
            except Exception as e:
                print(f"Could not delete photo folder: {e}")

        self._refresh_stack_page(0, self._build_students_page)

    def _build_courses_page(self):
        page = QWidget()
        page.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(page)
        lay.setContentsMargins(32, 28, 32, 24)
        lay.setSpacing(18)

        lay.addLayout(self._page_header("Courses", "All sections and enrolled students", show_add_course=True))

        courses = []
        all_students = []
        if DB_CONNECTED:
            try:
                courses = list(db["Courses"].find({}))
                all_students = list(db["Students"].find({}))
            except Exception as e:
                print(f"[AIAS] Courses page load error: {e}")

        total_courses = len(courses)
        total_students = sum(len(c.get("EnrolledStudents", [])) for c in courses)
        enrolled_ids = {sid for c in courses for sid in c.get("EnrolledStudents", [])}
        ready_count = sum(
            1 for s in all_students
            if s.get("StudentID") in enrolled_ids and s.get("FaceEmbedding")
        )

        sr = QHBoxLayout()
        sr.setSpacing(16)
        sr.addWidget(make_stat_card("Total courses", total_courses, PRIMARY))
        sr.addWidget(make_stat_card("Total students", total_students))
        sr.addWidget(make_stat_card("Ready", ready_count, "#22C55E"))
        sr.addStretch()
        lay.addLayout(sr)

        search_field_c = QLineEdit()
        search_field_c.setPlaceholderText("🔍  Search by course title or section ID...")
        search_field_c.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:8px; padding:8px 14px; "
            f"font-size:13px; background:{WHITE}; color:{TEXT_DARK};"
        )
        search_field_c.setFixedHeight(38)
        lay.addWidget(search_field_c)

        course_widgets = []
        cards_row = QHBoxLayout()
        cards_row.setSpacing(16)
        cards_row.setContentsMargins(0, 0, 0, 0)

        for course in courses:
            section_id = str(course.get("SectionID", ""))
            title = course.get("CourseTitle", "")
            sched_raw = course.get("Schedule", "")
            if isinstance(sched_raw, dict):
                schedule = f"{sched_raw.get('Day', '')} {sched_raw.get('StartTime', '')}–{sched_raw.get('EndTime', '')}"
            else:
                schedule = str(sched_raw)
            enrolled = course.get("EnrolledStudents", [])
            n_enrolled = len(enrolled)
            n_ready = sum(
                1 for s in all_students
                if s.get("StudentID") in enrolled and s.get("FaceEmbedding")
            )
            progress = int(n_ready / n_enrolled * 100) if n_enrolled else 0

            ccard = QFrame()
            ccard.setFixedWidth(340)
            ccard.setStyleSheet(
                f"QFrame {{ background:{WHITE}; border:1px solid {BORDER}; border-radius:10px; }}"
            )
            cc = QVBoxLayout(ccard)
            cc.setContentsMargins(20, 16, 20, 16)
            cc.setSpacing(8)

            badge_lbl = QLabel(f"Section {section_id}")
            badge_lbl.setFixedHeight(24)
            badge_lbl.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            badge_lbl.setStyleSheet(
                f"background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:4px;"
                f"padding:2px 10px; font-size:11px; font-weight:600;"
            )
            cc.addWidget(badge_lbl, alignment=Qt.AlignLeft)

            title_lbl = QLabel(title)
            title_lbl.setStyleSheet(
                f"font-size:15px; font-weight:700; color:{TEXT_DARK};"
                f"background:transparent; border:none;"
            )
            cc.addWidget(title_lbl)

            for icon_txt, detail in [
                ("🗓", schedule),
                ("👥", f"{n_enrolled} students"),
            ]:
                row = QHBoxLayout()
                row.setSpacing(6)
                il = QLabel(icon_txt)
                il.setStyleSheet("font-size:14px; background:transparent; border:none;")
                dl = QLabel(detail)
                dl.setStyleSheet(
                    f"font-size:12px; color:{TEXT_MED}; background:transparent; border:none;"
                )
                row.addWidget(il)
                row.addWidget(dl)
                row.addStretch()
                cc.addLayout(row)

            prog_lbl = QLabel(f"{n_ready}/{n_enrolled} photos uploaded")
            prog_lbl.setStyleSheet(
                f"font-size:11px; color:{TEXT_GRAY}; background:transparent; border:none;"
            )
            prog = QProgressBar()
            prog.setValue(progress)
            prog.setTextVisible(False)
            prog.setFixedHeight(8)
            cc.addWidget(prog_lbl)
            cc.addWidget(prog)

            del_row = QHBoxLayout()
            del_row.addStretch()
            edit_btn = QPushButton("Edit")
            edit_btn.setCursor(Qt.PointingHandCursor)
            edit_btn.setFixedHeight(30)
            edit_btn.setStyleSheet(
                "background:#E6F1FB; color:#0C447C; border-radius:5px;"
                "padding:4px 14px; font-size:12px; font-weight:600;"
            )
            edit_btn.clicked.connect(lambda _, c=course: self._edit_course(c))
            del_row.addWidget(edit_btn)
            del_row.addSpacing(6)
            del_btn = QPushButton("Delete")
            del_btn.setCursor(Qt.PointingHandCursor)
            del_btn.setFixedHeight(30)
            del_btn.setStyleSheet(
                "background:#FCEBEB; color:#791F1F; border-radius:5px;"
                "padding:4px 14px; font-size:12px; font-weight:600;"
            )
            del_btn.clicked.connect(lambda _, sid=section_id, t=title: self._delete_course(sid, t))
            del_row.addWidget(del_btn)
            cc.addLayout(del_row)

            cards_row.addWidget(ccard)
            course_widgets.append((section_id, title, ccard))

        def filter_courses(text):
            text = text.lower().strip()
            for sid, ttl, widget in course_widgets:
                if not text:
                    widget.setVisible(True)
                else:
                    widget.setVisible(text in sid.lower() or text in ttl.lower())

        search_field_c.textChanged.connect(filter_courses)

        cards_row.addStretch()
        lay.addLayout(cards_row)
        lay.addStretch()
        return page

    def _edit_course(self, course):
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Course")
        dialog.setFixedWidth(460)
        dialog.setStyleSheet(f"background:{BG};")

        lay = QVBoxLayout(dialog)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        title_lbl = QLabel("Edit Course")
        title_lbl.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title_lbl)

        sched_raw = course.get("Schedule", {})
        if isinstance(sched_raw, dict):
            cur_day = sched_raw.get("Day", "")
            cur_start = sched_raw.get("StartTime", "")
            cur_end = sched_raw.get("EndTime", "")
        else:
            cur_day = cur_start = cur_end = ""

        fields = {}
        for label, key, value, placeholder in [
            ("Course Title", "course_title", course.get("CourseTitle", ""), "e.g. Data Science"),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
            field = QLineEdit()
            field.setText(value)
            field.setPlaceholderText(placeholder)
            field.setFixedHeight(38)
            fields[key] = field
            lay.addWidget(lbl)
            lay.addWidget(field)

        days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
        times = [f"{h:02d}:{m:02d}" for h in range(8, 23) for m in [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55]]
        combos = {}
        for label, key, options, current in [
            ("Day", "day", days, cur_day),
            ("Start Time", "start_time", times, cur_start),
            ("End Time", "end_time", times, cur_end),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
            cb = QComboBox()
            cb.addItems(options)
            cb.setFixedHeight(38)
            idx = cb.findText(current)
            if idx >= 0:
                cb.setCurrentIndex(idx)
            combos[key] = cb
            lay.addWidget(lbl)
            lay.addWidget(cb)

        instr_lbl = QLabel("Instructor")
        instr_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        lay.addWidget(instr_lbl)
        instr_combo = QComboBox()
        instr_combo.setFixedHeight(38)
        instructors = []
        if DB_CONNECTED:
            try:
                instructors = list(db["Instructors"].find({}))
            except Exception as e:
                print(f"[AIAS] Edit course instructors load error: {e}")
        current_instr = course.get("InstructorUsername", "")
        for instr in instructors:
            instr_combo.addItem(instr.get("Username", ""))
        if not instructors:
            instr_combo.addItem("(No instructors)")
        idx = instr_combo.findText(current_instr)
        if idx >= 0:
            instr_combo.setCurrentIndex(idx)
        lay.addWidget(instr_combo)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(dialog.reject)
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:8px 20px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(dialog.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        if dialog.exec_() != QDialog.Accepted:
            return

        new_title = fields["course_title"].text().strip()
        new_day = combos["day"].currentText()
        new_start = combos["start_time"].currentText()
        new_end = combos["end_time"].currentText()
        new_instr = instr_combo.currentText()

        if not new_title or new_instr == "(No instructors)":
            QMessageBox.warning(self, "Missing Fields", "Please fill in all fields.")
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        section_id = str(course.get("SectionID", ""))
        try:
            db["Courses"].update_one(
                {"SectionID": section_id},
                {"$set": {
                    "CourseTitle": new_title,
                    "InstructorUsername": new_instr,
                    "Schedule": {"Day": new_day, "StartTime": new_start, "EndTime": new_end},
                }},
            )
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        self._refresh_stack_page(1, self._build_courses_page)

    def _delete_course(self, section_id, course_title):
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete {course_title}?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        try:
            db["Courses"].delete_one({"SectionID": section_id})
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        try:
            db["Instructors"].update_many({}, {"$pull": {"AssignedSections": section_id}})
        except Exception as e:
            print(f"DB error removing section from instructors: {e}")

        try:
            db["Students"].update_many({}, {"$pull": {"EnrolledSections": section_id}})
        except Exception as e:
            print(f"DB error removing section from students: {e}")

        self._refresh_stack_page(1, self._build_courses_page)

    def _upload_photo(self, student):
        import shutil

        sid = str(student.get("StudentID", ""))

        paths, _ = QFileDialog.getOpenFileNames(
            self, f"Select Photos (multiple allowed) for {student.get('FullName', '')}",
            "", "Images (*.jpg *.jpeg *.png)"
        )
        if not paths:
            return

        try:
            folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", sid)
            os.makedirs(folder, exist_ok=True)

            existing = [
                f for f in os.listdir(folder)
                if f.lower().endswith((".jpg", ".jpeg", ".png"))
            ]
            start_idx = len(existing) + 1

            for i, path in enumerate(paths):
                ext = os.path.splitext(path)[1]
                dest = os.path.join(folder, f"img{start_idx + i}{ext}")
                shutil.copy2(path, dest)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not save photos:\n{e}")
            return

        all_folder_paths = [
            os.path.join(folder, f)
            for f in os.listdir(folder)
            if f.lower().endswith((".jpg", ".jpeg", ".png"))
        ]

        self._pending_student = student
        self._pending_folder  = folder

        self._embedding_worker = EmbeddingWorker(all_folder_paths)
        self._embedding_worker.finished.connect(self._on_embedding_done)
        self._embedding_worker.error.connect(self._on_embedding_error)
        self._embedding_worker.start()

        QMessageBox.information(
            self, "Processing...",
            f"{len(paths)} photo(s) saved for {student.get('FullName', '')}.\n"
            "Generating augmented variants and extracting embeddings...\n"
            "You'll see a confirmation when done."
        )

    def _on_embedding_done(self, embedding):
        student = self._pending_student
        folder  = self._pending_folder
        sid     = str(student.get("StudentID", ""))

        all_image_paths = [
            os.path.abspath(os.path.join(folder, f))
            for f in os.listdir(folder)
            if f.lower().endswith((".jpg", ".jpeg", ".png"))
        ] if os.path.isdir(folder) else []

        if DB_CONNECTED:
            try:
                db["Students"].update_one(
                    {"StudentID": sid},
                    {"$set": {
                        "ImagePaths":    all_image_paths,
                        "FaceEmbedding": embedding,
                    }}
                )
            except Exception as e:
                QMessageBox.critical(self, "DB Error", f"Could not save embedding:\n{e}")
                return

        # Also compute and save AdaFace embedding
        if _adaface_model is not None:
            try:
                import cv2 as _cv2_ada
                import numpy as _np_ada
                ada_embeddings = []
                for photo_path in all_image_paths:
                    img = _cv2_ada.imread(photo_path)
                    if img is None:
                        continue
                    ada_emb = _adaface_embedding(_adaface_model, img)
                    if ada_emb is not None:
                        ada_embeddings.append(ada_emb)
                if ada_embeddings:
                    avg_ada = _np_ada.mean(ada_embeddings, axis=0)
                    avg_ada = avg_ada / _np_ada.linalg.norm(avg_ada)
                    if DB_CONNECTED:
                        db["Students"].update_one(
                            {"StudentID": sid},
                            {"$set": {"AdaFaceEmbedding": avg_ada.tolist()}}
                        )
                    _adaface_db[str(sid)] = avg_ada
                    print(f"[AIAS] AdaFace saved for {sid}")
            except Exception as e:
                print(f"[AIAS] AdaFace save error: {e}")

        QMessageBox.information(
            self, "✅ Done",
            f"Embedding saved successfully for {student.get('FullName', '')}!"
        )

        self._refresh_stack_page(0, self._build_students_page)

        self._pending_student = None
        self._pending_folder  = None

    def _on_embedding_error(self, error_msg):
        QMessageBox.warning(self, "Embedding Failed", error_msg)
        self._pending_student = None
        self._pending_dest    = None

    def _view_photo(self, student):
        image_paths = student.get("ImagePaths", [])
        if not image_paths:
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Photo — {student.get('FullName', '')}")
        dialog.setFixedSize(320, 360)
        lay = QVBoxLayout(dialog)

        lbl = QLabel()
        pixmap = QPixmap(image_paths[0])
        if not pixmap.isNull():
            lbl.setPixmap(pixmap.scaled(280, 280, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        else:
            lbl.setText("Image not found")
        lbl.setAlignment(Qt.AlignCenter)

        name_lbl = QLabel(student.get("FullName", ""))
        name_lbl.setAlignment(Qt.AlignCenter)
        name_lbl.setStyleSheet("font-size:13px; font-weight:600;")

        lay.addWidget(lbl)
        lay.addWidget(name_lbl)
        dialog.exec_()

    def _add_course(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Course")
        dialog.setFixedWidth(460)
        dialog.setStyleSheet(f"background:{BG};")

        lay = QVBoxLayout(dialog)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        title_lbl = QLabel("Add New Course")
        title_lbl.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title_lbl)

        fields = {}
        for label, key, placeholder in [
            ("Section ID", "section_id", "e.g. 102"),
            ("Course Title", "course_title", "e.g. Data Science"),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
            field = QLineEdit()
            field.setPlaceholderText(placeholder)
            field.setFixedHeight(38)
            fields[key] = field
            lay.addWidget(lbl)
            lay.addWidget(field)

        days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
        times = [f"{h:02d}:{m:02d}" for h in range(8, 23) for m in [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55]]
        combos = {}
        for label, key, options in [
            ("Day", "day", days),
            ("Start Time", "start_time", times),
            ("End Time", "end_time", times),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
            cb = QComboBox()
            cb.addItems(options)
            cb.setFixedHeight(38)
            combos[key] = cb
            lay.addWidget(lbl)
            lay.addWidget(cb)

        instr_lbl = QLabel("Instructor")
        instr_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        lay.addWidget(instr_lbl)
        instr_combo = QComboBox()
        instr_combo.setFixedHeight(38)
        instructors = []
        instr_id_map = {}
        if DB_CONNECTED:
            try:
                instructors = list(db["Instructors"].find({}))
            except Exception as _e:
                print(f"[AIAS] error: {_e}")
        for instr in instructors:
            username = instr.get("Username", "")
            instr_combo.addItem(username)
            instr_id_map[username] = instr.get("_id")
        if not instructors:
            instr_combo.addItem("(No instructors)")
        lay.addWidget(instr_combo)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(dialog.reject)
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:8px 20px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(dialog.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        if dialog.exec_() != QDialog.Accepted:
            return

        section_id = fields["section_id"].text().strip()
        course_title = fields["course_title"].text().strip()
        day = combos["day"].currentText()
        start_time = combos["start_time"].currentText()
        end_time = combos["end_time"].currentText()
        selected_instructor = instr_combo.currentText()

        if not all([section_id, course_title]) or selected_instructor == "(No instructors)":
            QMessageBox.warning(self, "Missing Fields", "Please fill in all fields.")
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        try:
            if db["Courses"].find_one({"SectionID": section_id}):
                QMessageBox.warning(self, "Duplicate Section ID", f"Section ID '{section_id}' already exists.")
                return
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        try:
            db["Courses"].insert_one({
                "SectionID": section_id,
                "CourseTitle": course_title,
                "InstructorUsername": selected_instructor,
                "InstructorID": instr_id_map.get(selected_instructor),
                "Schedule": {"Day": day, "StartTime": start_time, "EndTime": end_time},
                "EnrolledStudents": [],
            })
        except Exception as e:
            QMessageBox.critical(self, "DB Error", f"Could not insert course:\n{e}")
            return

        try:
            db["Instructors"].update_one(
                {"Username": selected_instructor},
                {"$addToSet": {"AssignedSections": section_id}},
            )
        except Exception as e:
            print(f"DB error updating instructor: {e}")

        QMessageBox.information(self, "Success", f"Course '{course_title}' added successfully.")
        self._refresh_stack_page(1, self._build_courses_page)

    def _build_instructors_page(self):
        page = QWidget()
        page.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(page)
        lay.setContentsMargins(32, 28, 32, 24)
        lay.setSpacing(18)

        hdr = QHBoxLayout()
        col = QVBoxLayout()
        col.setSpacing(2)
        t = QLabel("Instructors")
        t.setStyleSheet(f"font-size:20px; font-weight:800; color:{TEXT_DARK};")
        s = QLabel("Manage instructor accounts and section assignments")
        s.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY};")
        col.addWidget(t)
        col.addWidget(s)
        hdr.addLayout(col)
        hdr.addStretch()
        add_btn = QPushButton("+ Add Instructor")
        add_btn.setFixedHeight(36)
        add_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px; padding:6px 16px;"
        )
        add_btn.setCursor(Qt.PointingHandCursor)
        add_btn.clicked.connect(self._add_instructor)
        hdr.addWidget(add_btn)
        lay.addLayout(hdr)

        instructors = []
        if DB_CONNECTED:
            try:
                instructors = list(db["Instructors"].find({}))
            except Exception as _e:
                print(f"[AIAS] error: {_e}")

        total = len(instructors)
        active_sections = len({
            sec
            for instr in instructors
            for sec in instr.get("AssignedSections", [])
        })

        sr = QHBoxLayout()
        sr.setSpacing(16)
        sr.addWidget(make_stat_card("Total instructors", total, PRIMARY))
        sr.addWidget(make_stat_card("Active sections", active_sections))
        sr.addStretch()
        lay.addLayout(sr)

        tbl = make_table(
            ["Username", "Full Name", "Email", "Sections", "Action"],
            instructors,
            col_widths={0: 140, 2: 200, 3: 100, 4: 180},
            stretch_col=1,
        )
        for r, instr in enumerate(instructors):
            username = instr.get("Username", "")
            fullname = instr.get("FullName", "")
            email = instr.get("UniversityEmail", "")
            sections = ", ".join(str(sec) for sec in instr.get("AssignedSections", []))

            tbl.setItem(r, 0, QTableWidgetItem(username))
            tbl.setItem(r, 1, QTableWidgetItem(fullname))
            tbl.setItem(r, 2, QTableWidgetItem(email))
            tbl.setItem(r, 3, QTableWidgetItem(sections))

            edit_btn = QPushButton("Edit")
            edit_btn.setCursor(Qt.PointingHandCursor)
            edit_btn.setStyleSheet(
                f"background:#E6F1FB; color:#0C447C; border-radius:5px;"
                "padding:4px 10px; font-size:12px; font-weight:600;"
            )
            edit_btn.clicked.connect(lambda _, i=instr: self._edit_instructor(i))
            del_btn = QPushButton("Delete")
            del_btn.setCursor(Qt.PointingHandCursor)
            del_btn.setStyleSheet(
                "background:#FCEBEB; color:#791F1F; border-radius:5px;"
                "padding:4px 10px; font-size:12px; font-weight:600;"
            )
            del_btn.clicked.connect(lambda _, u=username: self._delete_instructor(u))
            wrapper = QWidget()
            wrapper.setStyleSheet("background:transparent;")
            wl = QHBoxLayout(wrapper)
            wl.setContentsMargins(6, 4, 6, 4)
            wl.setSpacing(6)
            wl.addWidget(edit_btn, alignment=Qt.AlignCenter)
            wl.addWidget(del_btn, alignment=Qt.AlignCenter)
            tbl.setCellWidget(r, 4, wrapper)
            tbl.setRowHeight(r, 44)

        search_row = QHBoxLayout()
        search_field_i = QLineEdit()
        search_field_i.setPlaceholderText("🔍  Search by username or full name...")
        search_field_i.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:8px; padding:8px 14px; "
            f"font-size:13px; background:{WHITE}; color:{TEXT_DARK};"
        )
        search_field_i.setFixedHeight(38)
        search_row.addWidget(search_field_i)
        search_row.addStretch()

        def filter_instructors(text):
            text = text.lower().strip()
            for row in range(tbl.rowCount()):
                match = False
                for col in [0, 1]:
                    item = tbl.item(row, col)
                    if item and text in item.text().lower():
                        match = True
                        break
                tbl.setRowHidden(row, not match if text else False)

        search_field_i.textChanged.connect(filter_instructors)

        lay.addLayout(search_row)
        lay.addWidget(tbl)
        return page

    def _add_instructor(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Instructor")
        dialog.setFixedWidth(460)
        dialog.setStyleSheet(f"background:{BG};")

        lay = QVBoxLayout(dialog)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        title_lbl = QLabel("Add New Instructor")
        title_lbl.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title_lbl)

        fields = {}
        for label, key, placeholder, is_password in [
            ("Username", "username", "e.g. dr_ali", False),
            ("Full Name", "fullname", "e.g. Dr. Ali Al-Harbi", False),
            ("Password", "password", "Enter password", True),
            ("University Email", "email", "e.g. ali@qu.edu.sa", False),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
            field = QLineEdit()
            field.setPlaceholderText(placeholder)
            field.setFixedHeight(38)
            if is_password:
                field.setEchoMode(QLineEdit.Password)
            fields[key] = field
            lay.addWidget(lbl)
            lay.addWidget(field)

        sections_lbl = QLabel("Assign Sections")
        sections_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        lay.addWidget(sections_lbl)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(140)
        scroll.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:6px; background:{WHITE};"
        )
        scroll_widget = QWidget()
        scroll_widget.setStyleSheet(f"background:{WHITE};")
        scroll_lay = QVBoxLayout(scroll_widget)
        scroll_lay.setContentsMargins(10, 8, 10, 8)
        scroll_lay.setSpacing(6)

        checkboxes = []
        courses = []
        if DB_CONNECTED:
            try:
                courses = list(db["Courses"].find({}))
            except Exception as _e:
                print(f"[AIAS] error: {_e}")

        for course in courses:
            section_id = str(course.get("SectionID", ""))
            course_title = course.get("CourseTitle", "")
            cb = QCheckBox(f"{section_id} — {course_title}")
            cb.setStyleSheet(f"color:{TEXT_DARK}; font-size:12px; background:transparent;")
            checkboxes.append((cb, section_id))
            scroll_lay.addWidget(cb)

        if not courses:
            no_lbl = QLabel("No courses found")
            no_lbl.setStyleSheet(f"color:{TEXT_GRAY}; font-size:12px; background:transparent;")
            scroll_lay.addWidget(no_lbl)

        scroll_lay.addStretch()
        scroll.setWidget(scroll_widget)
        lay.addWidget(scroll)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(dialog.reject)
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:8px 20px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(dialog.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        if dialog.exec_() != QDialog.Accepted:
            return

        username = fields["username"].text().strip()
        fullname = fields["fullname"].text().strip()
        password = fields["password"].text()
        email = fields["email"].text().strip()

        if not username or not fullname or not password or not email:
            QMessageBox.warning(self, "Missing Fields", "Please fill in all fields.")
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        try:
            if db["Instructors"].find_one({"Username": username}):
                QMessageBox.warning(self, "Duplicate Username", f"Username '{username}' already exists.")
                return
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        hashed = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt())
        selected_sections = [sid for cb, sid in checkboxes if cb.isChecked()]

        try:
            db["Instructors"].insert_one({
                "Username": username,
                "FullName": fullname,
                "Password": hashed,
                "UniversityEmail": email,
                "AssignedSections": selected_sections,
            })
        except Exception as e:
            QMessageBox.critical(self, "DB Error", f"Could not insert instructor:\n{e}")
            return

        for sid in selected_sections:
            try:
                db["Courses"].update_one(
                    {"SectionID": sid},
                    {"$set": {"InstructorUsername": username}},
                )
            except Exception as e:
                print(f"DB error updating course: {e}")

        QMessageBox.information(self, "Success", f"Instructor '{fullname}' added successfully.")
        self._refresh_stack_page(2, self._build_instructors_page)

    def _edit_instructor(self, instructor):
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Instructor")
        dialog.setFixedWidth(460)
        dialog.setStyleSheet(f"background:{BG};")

        lay = QVBoxLayout(dialog)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        title_lbl = QLabel("Edit Instructor")
        title_lbl.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title_lbl)

        fields = {}
        for label, key, value, is_password, placeholder in [
            ("Full Name", "fullname", instructor.get("FullName", ""), False, ""),
            ("University Email", "email", instructor.get("UniversityEmail", ""), False, ""),
            ("New Password", "password", "", True, "Leave blank to keep current password"),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
            field = QLineEdit()
            field.setText(value)
            field.setFixedHeight(38)
            if is_password:
                field.setEchoMode(QLineEdit.Password)
            if placeholder:
                field.setPlaceholderText(placeholder)
            fields[key] = field
            lay.addWidget(lbl)
            lay.addWidget(field)

        sections_lbl = QLabel("Assigned Sections")
        sections_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        lay.addWidget(sections_lbl)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(140)
        scroll.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:6px; background:{WHITE};"
        )
        scroll_widget = QWidget()
        scroll_widget.setStyleSheet(f"background:{WHITE};")
        scroll_lay = QVBoxLayout(scroll_widget)
        scroll_lay.setContentsMargins(10, 8, 10, 8)
        scroll_lay.setSpacing(6)

        assigned = [str(s) for s in instructor.get("AssignedSections", [])]
        checkboxes = []
        courses = []
        if DB_CONNECTED:
            try:
                courses = list(db["Courses"].find({}))
            except Exception as _e:
                print(f"[AIAS] error: {_e}")

        for course in courses:
            section_id = str(course.get("SectionID", ""))
            course_title = course.get("CourseTitle", "")
            cb = QCheckBox(f"{section_id} — {course_title}")
            cb.setStyleSheet(f"color:{TEXT_DARK}; font-size:12px; background:transparent;")
            cb.setChecked(section_id in assigned)
            checkboxes.append((cb, section_id))
            scroll_lay.addWidget(cb)

        if not courses:
            no_lbl = QLabel("No courses found")
            no_lbl.setStyleSheet(f"color:{TEXT_GRAY}; font-size:12px; background:transparent;")
            scroll_lay.addWidget(no_lbl)

        scroll_lay.addStretch()
        scroll.setWidget(scroll_widget)
        lay.addWidget(scroll)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(dialog.reject)
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:8px 20px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(dialog.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        if dialog.exec_() != QDialog.Accepted:
            return

        new_name = fields["fullname"].text().strip()
        new_email = fields["email"].text().strip()
        new_password = fields["password"].text()

        if not new_name or not new_email:
            QMessageBox.warning(self, "Missing Fields", "Full Name and Email cannot be empty.")
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        username = instructor.get("Username", "")
        new_sections = [sid for cb, sid in checkboxes if cb.isChecked()]
        update_fields = {
            "FullName": new_name,
            "UniversityEmail": new_email,
            "AssignedSections": new_sections,
        }
        if new_password:
            update_fields["Password"] = bcrypt.hashpw(new_password.encode("utf-8"), bcrypt.gensalt())

        try:
            db["Instructors"].update_one({"Username": username}, {"$set": update_fields})
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        try:
            db["Courses"].update_many(
                {"InstructorUsername": username, "SectionID": {"$nin": new_sections}},
                {"$unset": {"InstructorUsername": ""}},
            )
            for sid in new_sections:
                db["Courses"].update_one(
                    {"SectionID": sid},
                    {"$set": {"InstructorUsername": username}},
                )
        except Exception as e:
            print(f"DB error syncing courses: {e}")

        self._refresh_stack_page(2, self._build_instructors_page)

    def _delete_instructor(self, username):
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete instructor '{username}'?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        try:
            db["Instructors"].delete_one({"Username": username})
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        try:
            db["Courses"].update_many(
                {"InstructorUsername": username},
                {"$unset": {"InstructorUsername": ""}},
            )
        except Exception as e:
            print(f"DB error cleaning instructor from courses: {e}")

        self._refresh_stack_page(2, self._build_instructors_page)

    def _import_excel(self):
        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        path, _ = QFileDialog.getOpenFileName(
            self, "Import Students from Excel", "", "Excel Files (*.xlsx *.xls)"
        )
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active

            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

            # Accept flexible column names
            def _col(names):
                for n in names:
                    for i, h in enumerate(headers):
                        if n in h:
                            return i
                return None

            id_col   = _col(["student id", "studentid", "id", "student_id"])
            name_col = _col(["full name", "fullname", "name", "full_name"])
            sec_col  = _col(["section", "course", "enrolled"])

            if id_col is None or name_col is None:
                QMessageBox.critical(
                    self, "Import Failed",
                    "Could not find required columns.\n"
                    "Excel must have: 'Student ID' and 'Full Name' columns.\n"
                    f"Found columns: {', '.join(headers)}"
                )
                return

            inserted = skipped = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                sid  = str(row[id_col]).strip()   if row[id_col]   is not None else ""
                name = str(row[name_col]).strip()  if row[name_col] is not None else ""
                sec  = str(row[sec_col]).strip()   if sec_col is not None and row[sec_col] is not None else ""

                if not sid or not name or sid == "None":
                    continue

                if db["Students"].find_one({"StudentID": sid}):
                    skipped += 1
                    continue

                doc = {"StudentID": sid, "FullName": name, "EnrolledSections": []}
                if sec:
                    doc["EnrolledSections"] = [sec]
                    db["Courses"].update_one({"SectionID": sec}, {"$addToSet": {"EnrolledStudents": sid}})

                db["Students"].insert_one(doc)
                inserted += 1

            wb.close()
            QMessageBox.information(
                self, "Import Complete",
                f"Import finished.\n✔ Inserted: {inserted}\n⚠ Skipped (duplicate): {skipped}"
            )

            self._refresh_stack_page(0, self._build_students_page)

        except Exception as e:
            QMessageBox.critical(self, "Import Failed", f"Error reading Excel file:\n{e}")

    def _batch_embed_folder(self):
        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        root_folder = QFileDialog.getExistingDirectory(
            self, "Select Root Folder (subfolders named by Student ID)"
        )
        if not root_folder:
            return

        data_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data")
        os.makedirs(data_folder, exist_ok=True)

        # Progress dialog
        progress_dialog = QDialog(self)
        progress_dialog.setWindowTitle("Batch Embedding")
        progress_dialog.setFixedSize(500, 380)
        progress_dialog.setStyleSheet(f"background:{BG};")
        pd_lay = QVBoxLayout(progress_dialog)
        pd_lay.setContentsMargins(24, 20, 24, 20)
        pd_lay.setSpacing(10)

        pd_title = QLabel("Processing student folders…")
        pd_title.setStyleSheet(f"font-size:15px; font-weight:700; color:{TEXT_DARK};")
        pd_lay.addWidget(pd_title)

        log_widget = QTableWidget(0, 1)
        log_widget.horizontalHeader().setVisible(False)
        log_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        log_widget.verticalHeader().setVisible(False)
        log_widget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        log_widget.setShowGrid(False)
        log_widget.setStyleSheet(f"background:{WHITE}; border:1px solid {BORDER}; border-radius:6px;")
        pd_lay.addWidget(log_widget)

        close_btn = QPushButton("Close")
        close_btn.setEnabled(False)
        close_btn.setFixedHeight(36)
        close_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; padding:6px 20px; font-weight:600;"
        )
        close_btn.clicked.connect(progress_dialog.accept)
        pd_lay.addWidget(close_btn, alignment=Qt.AlignRight)

        self._batch_worker = BatchEmbedWorker(root_folder, data_folder)

        def _on_progress(msg):
            r = log_widget.rowCount()
            log_widget.insertRow(r)
            item = QTableWidgetItem(msg)
            item.setFlags(Qt.ItemIsEnabled)
            log_widget.setItem(r, 0, item)
            log_widget.scrollToBottom()

        def _on_finished(processed, skipped):
            pd_title.setText(f"Done — ✔ {processed} embedded, ⚠ {skipped} skipped")
            close_btn.setEnabled(True)
            self._refresh_stack_page(0, self._build_students_page)

        self._batch_worker.progress.connect(_on_progress)
        self._batch_worker.finished.connect(_on_finished)
        self._batch_worker.start()
        progress_dialog.exec_()

    def _build_settings_page(self):
        page = QWidget()
        page.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(page)
        lay.setContentsMargins(40, 32, 40, 32)
        lay.setSpacing(20)

        late_cutoff  = LATE_CUTOFF_MINUTES
        early_cutoff = EARLY_LEAVE_CUTOFF_MINUTES
        threshold    = RECOGNITION_THRESHOLD
        if DB_CONNECTED:
            try:
                doc = db["Settings"].find_one({"_id": "global"})
                if doc:
                    late_cutoff  = doc.get("LateCutoffMinutes",       10)
                    early_cutoff = doc.get("EarlyLeaveCutoffMinutes", 15)
                    threshold    = doc.get("RecognitionThreshold",    0.55)
            except Exception as e:
                print(f"[AIAS] Settings page load error: {e}")

        title = QLabel("System Settings")
        title.setStyleSheet(f"font-size:20px; font-weight:800; color:{TEXT_DARK};")
        sub = QLabel("Changes take effect immediately and persist across restarts.")
        sub.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY};")
        lay.addWidget(title)
        lay.addWidget(sub)

        fields = {}
        for label, key, val, desc in [
            ("Late Arrival Cutoff (minutes)", "late", str(late_cutoff),
             "Students arriving after this many minutes are marked Late"),
            ("Early Leave Cutoff (minutes)", "early", str(early_cutoff),
             "Students not seen for this many minutes are marked Early Leave"),
            ("Face Recognition Threshold (0.0–1.0)", "threshold", str(threshold),
             "Minimum cosine similarity to identify a student (higher = stricter)"),
        ]:
            card = QFrame()
            card.setStyleSheet(
                f"QFrame{{background:{WHITE};border:1px solid {BORDER};border-radius:8px;}}"
            )
            cl = QVBoxLayout(card)
            cl.setContentsMargins(20, 16, 20, 16)
            cl.setSpacing(6)
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:13px; font-weight:700; color:{TEXT_DARK};")
            desc_lbl = QLabel(desc)
            desc_lbl.setStyleSheet(f"font-size:11px; color:{TEXT_GRAY};")
            field = QLineEdit(val)
            field.setFixedHeight(38)
            field.setFixedWidth(160)
            fields[key] = field
            cl.addWidget(lbl)
            cl.addWidget(desc_lbl)
            cl.addWidget(field)
            lay.addWidget(card)

        save_btn = QPushButton("Save Settings")
        save_btn.setFixedHeight(40)
        save_btn.setFixedWidth(160)
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(lambda: self._save_settings(fields))
        lay.addWidget(save_btn)
        lay.addStretch()
        return page

    def _save_settings(self, fields):
        try:
            late   = int(fields["late"].text().strip())
            early  = int(fields["early"].text().strip())
            thresh = float(fields["threshold"].text().strip())
        except ValueError:
            QMessageBox.warning(self, "Invalid Input", "Please enter valid numeric values.")
            return
        if not (0.0 < thresh <= 1.0):
            QMessageBox.warning(self, "Invalid Threshold", "Threshold must be between 0.0 (exclusive) and 1.0 (inclusive).")
            return
        if late < 1 or early < 1:
            QMessageBox.warning(self, "Invalid Value", "Cutoff values must be at least 1 minute.")
            return

        global LATE_CUTOFF_MINUTES, EARLY_LEAVE_CUTOFF_MINUTES, RECOGNITION_THRESHOLD
        LATE_CUTOFF_MINUTES        = late
        EARLY_LEAVE_CUTOFF_MINUTES = early
        RECOGNITION_THRESHOLD      = thresh

        if DB_CONNECTED:
            try:
                db["Settings"].update_one(
                    {"_id": "global"},
                    {"$set": {
                        "LateCutoffMinutes":       late,
                        "EarlyLeaveCutoffMinutes": early,
                        "RecognitionThreshold":    thresh,
                    }},
                    upsert=True,
                )
                QMessageBox.information(self, "Saved", "Settings saved and applied immediately.")
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))
        else:
            QMessageBox.information(self, "Saved", "Settings applied for this session (no DB).")

    def _logout(self):
        for widget in QApplication.instance().topLevelWidgets():
            if isinstance(widget, AppWindow):
                widget._login_page.username_field.clear()
                widget._login_page.password_field.clear()
                widget._stack.setCurrentWidget(widget._login_page)
                widget.setStyleSheet("background:#0a0f0a;")
                break
        self.deleteLater()

    def keyPressEvent(self, event):
        from PyQt5.QtCore import Qt
        if event.key() == Qt.Key_Escape:
            pass
        else:
            super().keyPressEvent(event)


class SessionHistoryWindow(QWidget):
    def __init__(self, main_win):
        super().__init__()
        self.main_win = main_win
        self.setWindowTitle("AIAS — Session History")
        import os
        from PyQt5.QtGui import QIcon
        _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aias_icon.ico")
        if os.path.exists(_icon_path):
            self.setWindowIcon(QIcon(_icon_path))
        self.setMinimumSize(1100, 700)
        self.resize(1100, 700)
        self.setStyleSheet(f"background:{BG};")
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(48, 36, 48, 28)
        root.setSpacing(18)

        title = QLabel("Session History")
        title.setStyleSheet(f"font-size:24px; font-weight:800; color:{TEXT_DARK};")
        sub = QLabel("Past completed sessions for your courses")
        sub.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY};")

        header_row = QHBoxLayout()
        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        title_col.addWidget(title)
        title_col.addWidget(sub)
        header_row.addLayout(title_col)
        header_row.addStretch()

        dark_btn = QPushButton("🌙  Dark Mode" if not IS_DARK_MODE else "☀  Light Mode")
        dark_btn.setStyleSheet(
            f"background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px; "
            "padding:6px 14px; font-size:12px; font-weight:600;"
        )
        dark_btn.setCursor(Qt.PointingHandCursor)
        dark_btn.clicked.connect(lambda: _toggle_dark(dark_btn))

        def _toggle_dark(btn):
            new_dark = not IS_DARK_MODE
            apply_theme(QApplication.instance(), dark=new_dark)
            btn.setText("☀  Light Mode" if new_dark else "🌙  Dark Mode")
            btn.setStyleSheet(
                f"background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px; "
                "padding:6px 14px; font-size:12px; font-weight:600;"
            )
            for widget in QApplication.instance().topLevelWidgets():
                force_theme_on_all_widgets(widget)
                widget.update()

        header_row.addWidget(dark_btn)
        root.addLayout(header_row)

        sessions = []
        if DB_CONNECTED:
            try:
                instr_username = self.main_win.instructor_data.get("username", "")
                raw = list(
                    db["Sessions"].find(
                        {"InstructorID": instr_username, "Status": "completed"}
                    ).sort("StartTime", -1).limit(100)
                )

                # Bug 7: batch-fetch all courses and all logs in 2 queries instead of N+1
                section_ids = list({s.get("SectionID", "") for s in raw if s.get("SectionID")})
                course_map = {}
                if section_ids:
                    for c in db["Courses"].find({"SectionID": {"$in": section_ids}}):
                        _sec = c.get("SectionID", "")
                        course_map[_sec] = f"{_sec} — {c.get('CourseTitle', '')}"

                sess_ids = [s["_id"] for s in raw]
                log_counts = {}   # session_id -> {status: count}
                sess_logs  = {}   # session_id -> [log docs]
                if sess_ids:
                    for log in db["AttendanceLogs"].find({"SessionID": {"$in": sess_ids}}):
                        _s = log.get("SessionID")
                        if _s not in log_counts:
                            log_counts[_s] = {"Present": 0, "Late": 0, "Early Leave": 0, "Absent": 0}
                            sess_logs[_s]  = []
                        _st = log.get("Status", "Absent")
                        if _st in log_counts[_s]:
                            log_counts[_s][_st] += 1
                        sess_logs[_s].append(log)

                for s in raw:
                    _doc_id    = s.get("_id")
                    section_id = s.get("SectionID", "")
                    start_time = s.get("StartTime")
                    date_str   = start_time.strftime("%Y-%m-%d") if start_time else "—"
                    time_str   = start_time.strftime("%H:%M")    if start_time else "—"
                    course_title = course_map.get(section_id, section_id)
                    counts = log_counts.get(_doc_id, {"Present": 0, "Late": 0, "Early Leave": 0, "Absent": 0})
                    logs   = sess_logs.get(_doc_id, [])
                    sessions.append((
                        date_str, time_str, course_title,
                        counts["Present"], counts["Late"], counts["Early Leave"], counts["Absent"],
                        logs,
                    ))
            except Exception as e:
                print(f"[AIAS] History load error: {e}")

        cols = ["Date", "Time", "Course", "Present", "Late", "Early Leave", "Absent", ""]
        tbl  = make_table(cols, [], col_widths={0: 110, 1: 75, 7: 80}, stretch_col=2)
        tbl.setRowCount(len(sessions))

        for r, (date_s, time_s, course_t, n_p, n_l, n_e, n_a, logs) in enumerate(sessions):
            for col, val in enumerate([date_s, time_s, course_t, str(n_p), str(n_l), str(n_e), str(n_a)]):
                item = QTableWidgetItem(val)
                item.setFlags(Qt.ItemIsEnabled)
                tbl.setItem(r, col, item)
            view_btn = QPushButton("View")
            view_btn.setFixedHeight(30)
            view_btn.setStyleSheet(
                f"background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:5px;"
                "font-size:11px; font-weight:600;"
            )
            view_btn.setCursor(Qt.PointingHandCursor)
            view_btn.clicked.connect(
                lambda _, c=course_t, d=date_s, t=time_s, ls=logs: self._view_detail(c, d, t, ls)
            )
            tbl.setCellWidget(r, 7, view_btn)
            tbl.setRowHeight(r, 44)

        root.addWidget(tbl)

        back_row = QHBoxLayout()
        back_row.addStretch()
        back_btn = QPushButton("← Back")
        back_btn.setFixedHeight(40)
        back_btn.setStyleSheet(
            f"background:{WHITE}; color:{PRIMARY}; border:1px solid {PRIMARY}; border-radius:6px;"
            "padding:8px 16px; font-size:13px; font-weight:600;"
        )
        back_btn.setCursor(Qt.PointingHandCursor)
        back_btn.clicked.connect(self._go_back)
        back_row.addWidget(back_btn)
        root.addLayout(back_row)

    def _view_detail(self, course_title, date_str, time_str, logs):
        session_id = logs[0].get("SessionID") if logs else None

        dlg = QDialog(self)
        dlg.setWindowTitle(f"Session Detail — {course_title}")
        dlg.setMinimumSize(900, 500)
        dlg.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        hdr = QLabel(f"{course_title}  ·  {date_str}  {time_str}")
        hdr.setStyleSheet(f"font-size:15px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(hdr)

        rows_data = []
        for log in logs:
            sid = str(log.get("StudentID", ""))
            fullname = sid
            if DB_CONNECTED:
                try:
                    stu = db["Students"].find_one({"StudentID": sid})
                    if stu:
                        fullname = stu.get("FullName", sid)
                except Exception as e:
                    print(f"[AIAS] Student lookup error for {sid}: {e}")
            first_dt = log.get("FirstSeenAt")
            last_dt  = log.get("LastSeenAt")
            rows_data.append([
                sid, fullname,
                first_dt.strftime("%H:%M") if first_dt else "—",
                last_dt.strftime("%H:%M")  if last_dt  else "—",
                log.get("Status", "—"),
                log.get("Note",   ""),
            ])

        det_tbl = make_table(
            ["Student ID", "Full Name", "First Seen", "Last Seen", "Status", "Note", ""],
            rows_data,
            col_widths={0: 120, 2: 90, 3: 90, 4: 110, 5: 150, 6: 60},
            stretch_col=1,
        )
        det_tbl.setRowCount(len(rows_data))

        def _fill_row(r, row):
            det_tbl.setItem(r, 0, QTableWidgetItem(row[0]))
            det_tbl.setItem(r, 1, QTableWidgetItem(row[1]))
            det_tbl.setItem(r, 2, QTableWidgetItem(row[2]))
            det_tbl.setItem(r, 3, QTableWidgetItem(row[3]))
            det_tbl.setCellWidget(r, 4, make_badge(row[4]))
            det_tbl.setItem(r, 5, QTableWidgetItem(row[5]))
            edit_btn = QPushButton("Edit")
            edit_btn.setFixedHeight(28)
            edit_btn.setStyleSheet(
                "background:#E6F1FB; color:#0C447C; border-radius:5px;"
                "font-size:11px; font-weight:600;"
            )
            edit_btn.setCursor(Qt.PointingHandCursor)
            edit_btn.clicked.connect(
                lambda _, ri=r: self._edit_log_entry(session_id, ri, rows_data, det_tbl, _fill_row)
            )
            det_tbl.setCellWidget(r, 6, edit_btn)
            det_tbl.setRowHeight(r, 40)

        for r, row in enumerate(rows_data):
            _fill_row(r, row)
        lay.addWidget(det_tbl)

        btn_row = QHBoxLayout()
        export_btn = QPushButton("Export Excel")
        export_btn.setStyleSheet(
            f"background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px;"
            "padding:8px 16px; font-weight:600;"
        )
        export_btn.setCursor(Qt.PointingHandCursor)
        export_btn.clicked.connect(lambda: self._export_session(course_title, date_str, rows_data))
        close_btn = QPushButton("Close")
        close_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; padding:8px 20px; font-weight:600;"
        )
        close_btn.clicked.connect(dlg.accept)
        btn_row.addWidget(export_btn)
        btn_row.addStretch()
        btn_row.addWidget(close_btn)
        lay.addLayout(btn_row)
        dlg.exec_()

    def _edit_log_entry(self, session_id, row_idx, rows_data, tbl, fill_row_fn):
        row = rows_data[row_idx]
        sid, name, current_status, current_note = row[0], row[1], row[4], row[5]

        dlg = QDialog(self)
        dlg.setWindowTitle(f"Edit — {name}")
        dlg.setFixedWidth(340)
        dlg.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(24, 20, 24, 20)
        lay.setSpacing(10)

        title_lbl = QLabel(f"Edit attendance for:\n{name}")
        title_lbl.setStyleSheet(f"font-size:14px; font-weight:700; color:{TEXT_DARK};")
        title_lbl.setWordWrap(True)
        lay.addWidget(title_lbl)

        group = QButtonGroup(dlg)
        buttons = {}
        for status in ["Present", "Late", "Early Leave", "Absent"]:
            rb = QRadioButton(status)
            rb.setStyleSheet(f"font-size:13px; color:{TEXT_MED};")
            if status == current_status:
                rb.setChecked(True)
            group.addButton(rb)
            buttons[status] = rb
            lay.addWidget(rb)

        note_lbl = QLabel("Note (optional)")
        note_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        note_field = QLineEdit()
        note_field.setPlaceholderText("e.g. Medical excuse, arrived late by bus…")
        note_field.setFixedHeight(36)
        note_field.setText(current_note)
        lay.addWidget(note_lbl)
        lay.addWidget(note_field)

        btn_row = QHBoxLayout()
        cancel = QPushButton("Cancel")
        cancel.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:7px 14px;"
        )
        cancel.clicked.connect(dlg.reject)
        save = QPushButton("Save")
        save.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px;"
            "padding:7px 14px; font-weight:700;"
        )
        save.clicked.connect(dlg.accept)
        btn_row.addWidget(cancel)
        btn_row.addStretch()
        btn_row.addWidget(save)
        lay.addLayout(btn_row)

        if dlg.exec_() != QDialog.Accepted:
            return

        new_status = next((s for s, rb in buttons.items() if rb.isChecked()), current_status)
        new_note   = note_field.text().strip()
        if new_status == current_status and new_note == current_note:
            return

        if DB_CONNECTED and session_id is not None:
            try:
                db["AttendanceLogs"].update_one(
                    {"SessionID": session_id, "StudentID": sid},
                    {"$set": {"Status": new_status, "Note": new_note}},
                )
            except Exception as e:
                QMessageBox.critical(self, "DB Error", f"Could not save:\n{e}")
                return

        rows_data[row_idx][4] = new_status
        rows_data[row_idx][5] = new_note
        fill_row_fn(row_idx, rows_data[row_idx])

    def _export_session(self, course_title, date_str, rows_data):
        safe_course = "".join(c for c in course_title if c.isalnum() or c in " _-").strip()
        default_name = f"Attendance_{safe_course}_{date_str}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Attendance", default_name, "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance"

            green_fill  = PatternFill("solid", start_color="1B5E35")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            center      = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("A1:F1")
            ws["A1"] = f"Attendance Report — {course_title}  |  {date_str}"
            ws["A1"].font      = Font(bold=True, size=13, color="1A1A2E")
            ws["A1"].alignment = center

            headers = ["Student ID", "Full Name", "First Seen", "Last Seen", "Status", "Note"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=h)
                cell.fill      = green_fill
                cell.font      = header_font
                cell.alignment = center

            for row_idx, (sid, name, fs, ls, status, note) in enumerate(rows_data, 3):
                ws.cell(row=row_idx, column=1, value=sid)
                ws.cell(row=row_idx, column=2, value=name)
                ws.cell(row=row_idx, column=3, value=fs)
                ws.cell(row=row_idx, column=4, value=ls)
                ws.cell(row=row_idx, column=5, value=status)
                ws.cell(row=row_idx, column=6, value=note)

            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            wb.save(path)
            QMessageBox.information(self, "Exported", f"Saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    def _go_back(self):
        self.hide()
        self.main_win.show()


class AnalyticsWindow(QWidget):
    def __init__(self, main_win):
        super().__init__()
        self.main_win = main_win
        self.setWindowTitle("AIAS — Attendance Analytics")
        import os
        from PyQt5.QtGui import QIcon
        _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aias_icon.ico")
        if os.path.exists(_icon_path):
            self.setWindowIcon(QIcon(_icon_path))
        self.setMinimumSize(1100, 700)
        self.resize(1100, 700)
        self.setStyleSheet(f"background:{BG};")
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(48, 36, 48, 28)
        root.setSpacing(16)

        title = QLabel("Attendance Analytics")
        title.setStyleSheet(f"font-size:24px; font-weight:800; color:{TEXT_DARK};")
        sub = QLabel("Per-student attendance rates across all completed sessions")
        sub.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY};")

        header_row = QHBoxLayout()
        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        title_col.addWidget(title)
        title_col.addWidget(sub)
        header_row.addLayout(title_col)
        header_row.addStretch()

        dark_btn2 = QPushButton("🌙  Dark Mode" if not IS_DARK_MODE else "☀  Light Mode")
        dark_btn2.setStyleSheet(
            f"background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px; "
            "padding:6px 14px; font-size:12px; font-weight:600;"
        )
        dark_btn2.setCursor(Qt.PointingHandCursor)
        dark_btn2.clicked.connect(lambda: _toggle_dark2(dark_btn2))

        def _toggle_dark2(btn):
            new_dark = not IS_DARK_MODE
            apply_theme(QApplication.instance(), dark=new_dark)
            btn.setText("☀  Light Mode" if new_dark else "🌙  Dark Mode")
            btn.setStyleSheet(
                f"background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px; "
                "padding:6px 14px; font-size:12px; font-weight:600;"
            )
            for widget in QApplication.instance().topLevelWidgets():
                force_theme_on_all_widgets(widget)
                widget.update()

        header_row.addWidget(dark_btn2)

        export_btn = QPushButton("📥  Export to Excel")
        export_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; "
            "padding:8px 16px; font-size:12px; font-weight:600;"
        )
        export_btn.setCursor(Qt.PointingHandCursor)
        export_btn.clicked.connect(lambda: self._export_analytics())
        header_row.addWidget(export_btn)
        root.addLayout(header_row)

        courses_data = getattr(self.main_win, "_courses_data", [])
        ctrl_row = QHBoxLayout()
        course_lbl = QLabel("Course:")
        course_lbl.setStyleSheet(f"font-size:13px; font-weight:600; color:{TEXT_MED};")
        self._course_combo = QComboBox()
        self._course_combo.setFixedHeight(36)
        self._course_combo.setMinimumWidth(300)
        for c in courses_data:
            self._course_combo.addItem(
                f"{c.get('SectionID', '')} — {c.get('CourseTitle', '')}",
                c
            )
        ctrl_row.addWidget(course_lbl)
        ctrl_row.addSpacing(8)
        ctrl_row.addWidget(self._course_combo)
        ctrl_row.addStretch()
        root.addLayout(ctrl_row)

        def _make_updatable(label_text, color=TEXT_DARK):
            f = QFrame()
            f.setStyleSheet(
                f"QFrame{{background:{WHITE};border:1px solid {BORDER};border-radius:8px;}}"
            )
            f.setFixedHeight(82)
            f.setMinimumWidth(130)
            fl = QVBoxLayout(f)
            fl.setContentsMargins(16, 12, 16, 12)
            fl.setSpacing(2)
            vl = QLabel("—")
            vl.setStyleSheet(
                f"font-size:26px; font-weight:800; color:{color};"
                "border:none; background:transparent;"
            )
            tl = QLabel(label_text)
            tl.setStyleSheet(
                f"font-size:11px; color:{TEXT_GRAY}; border:none; background:transparent;"
            )
            fl.addWidget(vl)
            fl.addWidget(tl)
            return f, vl

        sessions_card, self._lbl_sessions = _make_updatable("Total Sessions",   PRIMARY)
        avg_card,      self._lbl_avg      = _make_updatable("Avg Attendance %", "#22C55E")
        atrisk_card,   self._lbl_atrisk   = _make_updatable("At Risk (<60%)",   "#EF4444")

        stats_row = QHBoxLayout()
        stats_row.setSpacing(16)
        stats_row.addWidget(sessions_card)
        stats_row.addWidget(avg_card)
        stats_row.addWidget(atrisk_card)
        stats_row.addStretch()
        root.addLayout(stats_row)

        self._analytics_rows = []
        self._tbl = make_table(
            ["Student ID", "Full Name", "Attended", "Total Sessions", "Attendance Rate"],
            [],
            col_widths={0: 140, 2: 100, 3: 130, 4: 160},
            stretch_col=1,
        )
        self._tbl.setCursor(Qt.PointingHandCursor)
        self._tbl.cellClicked.connect(self._on_analytics_row_click)
        root.addWidget(self._tbl)

        hint = QLabel("Click any row to see per-session breakdown.")
        hint.setStyleSheet(f"font-size:11px; color:{TEXT_GRAY};")
        root.addWidget(hint)

        back_row = QHBoxLayout()
        back_row.addStretch()
        back_btn = QPushButton("← Back")
        back_btn.setFixedHeight(40)
        back_btn.setStyleSheet(
            f"background:{WHITE}; color:{PRIMARY}; border:1px solid {PRIMARY}; border-radius:6px;"
            "padding:8px 16px; font-size:13px; font-weight:600;"
        )
        back_btn.setCursor(Qt.PointingHandCursor)
        back_btn.clicked.connect(self._go_back)
        back_row.addWidget(back_btn)
        root.addLayout(back_row)

        self._course_combo.currentIndexChanged.connect(self._load_analytics)
        if courses_data:
            self._load_analytics(0)

    def _load_analytics(self, _idx=0):
        course = self._course_combo.currentData()
        if not course:
            return

        section_id   = str(course.get("SectionID", ""))
        enrolled_ids = [str(s) for s in course.get("EnrolledStudents", [])]

        sessions = []
        if DB_CONNECTED:
            try:
                sessions = list(db["Sessions"].find(
                    {"SectionID": section_id, "Status": "completed"}
                ).sort("StartTime", 1))
            except Exception as e:
                print(f"[AIAS] Analytics session load: {e}")

        n_sessions = len(sessions)
        self._lbl_sessions.setText(str(n_sessions))

        # Build per-student attendance count and per-session breakdown
        attended        = {sid: 0 for sid in enrolled_ids}
        student_sessions = {sid: [] for sid in enrolled_ids}

        if DB_CONNECTED and sessions:
            sess_ids   = [s["_id"] for s in sessions]
            sess_dates = {s["_id"]: s.get("StartTime") for s in sessions}
            try:
                for log in db["AttendanceLogs"].find({"SessionID": {"$in": sess_ids}}):
                    _s = str(log.get("StudentID", ""))
                    if _s in student_sessions:
                        _dt = sess_dates.get(log.get("SessionID"))
                        student_sessions[_s].append({
                            "date":   _dt,
                            "status": log.get("Status", "Absent"),
                        })
                        if log.get("Status") in ("Present", "Late"):
                            attended[_s] += 1
            except Exception as e:
                print(f"[AIAS] Analytics log error: {e}")

        name_map = {}
        if DB_CONNECTED and enrolled_ids:
            try:
                for stu in db["Students"].find({"StudentID": {"$in": enrolled_ids}}):
                    name_map[str(stu.get("StudentID", ""))] = stu.get("FullName", "")
            except Exception as e:
                print(f"[AIAS] Analytics student name load error: {e}")

        rows = [
            (sid, name_map.get(sid, sid), attended.get(sid, 0), n_sessions,
             (attended.get(sid, 0) / n_sessions * 100) if n_sessions > 0 else 0.0,
             student_sessions.get(sid, []))
            for sid in enrolled_ids
        ]
        self._analytics_rows = rows

        self._tbl.setRowCount(len(rows))
        for r, (sid, name, cnt, total, rate, _sess) in enumerate(rows):
            self._tbl.setItem(r, 0, QTableWidgetItem(sid))
            self._tbl.setItem(r, 1, QTableWidgetItem(name))
            self._tbl.setItem(r, 2, QTableWidgetItem(str(cnt)))
            self._tbl.setItem(r, 3, QTableWidgetItem(str(total)))

            if rate >= 80:
                bg, fg = "#EAF3DE", "#27500A"
            elif rate >= 60:
                bg, fg = "#FAEEDA", "#633806"
            else:
                bg, fg = "#FCEBEB", "#791F1F"

            rate_lbl = QLabel(f"{rate:.0f}%")
            rate_lbl.setAlignment(Qt.AlignCenter)
            rate_lbl.setFixedHeight(26)
            rate_lbl.setStyleSheet(
                f"background:{bg}; color:{fg}; border-radius:4px;"
                "padding:2px 10px; font-size:12px; font-weight:700;"
            )
            outer = QWidget()
            outer.setStyleSheet("background:transparent;")
            ol = QHBoxLayout(outer)
            ol.setContentsMargins(8, 4, 8, 4)
            ol.addWidget(rate_lbl)
            self._tbl.setCellWidget(r, 4, outer)
            self._tbl.setRowHeight(r, 44)

        avg     = (sum(r[4] for r in rows) / len(rows)) if rows else 0.0
        at_risk = sum(1 for r in rows if r[4] < 60)
        self._lbl_avg.setText(f"{avg:.0f}%")
        self._lbl_atrisk.setText(str(at_risk))

    def _on_analytics_row_click(self, row, _col):
        if row < len(self._analytics_rows):
            sid, name, _cnt, _total, _rate, sess_data = self._analytics_rows[row]
            self._show_student_sessions(sid, name, sess_data)

    def _show_student_sessions(self, sid, name, sess_data):
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Session History — {name}")
        dlg.setMinimumSize(480, 400)
        dlg.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(24, 20, 24, 20)
        lay.setSpacing(12)

        hdr = QLabel(f"{name}  ({sid})")
        hdr.setStyleSheet(f"font-size:15px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(hdr)

        if not sess_data:
            empty = QLabel("No attendance records found.")
            empty.setAlignment(Qt.AlignCenter)
            empty.setStyleSheet(f"color:{TEXT_GRAY}; font-size:13px;")
            lay.addWidget(empty)
        else:
            det_tbl = make_table(["Date", "Status"], sess_data, col_widths={0: 160}, stretch_col=1)
            det_tbl.setRowCount(len(sess_data))
            for r, entry in enumerate(sess_data):
                dt = entry.get("date")
                det_tbl.setItem(r, 0, QTableWidgetItem(
                    dt.strftime("%Y-%m-%d %H:%M") if dt else "—"
                ))
                det_tbl.setCellWidget(r, 1, make_badge(entry.get("status", "—")))
                det_tbl.setRowHeight(r, 40)
            lay.addWidget(det_tbl)

        close_btn = QPushButton("Close")
        close_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; padding:8px 20px; font-weight:600;"
        )
        close_btn.clicked.connect(dlg.accept)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_row.addWidget(close_btn)
        lay.addLayout(btn_row)
        dlg.exec_()

    def _export_analytics(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from PyQt5.QtWidgets import QFileDialog, QMessageBox
        import datetime

        path, _ = QFileDialog.getSaveFileName(
            self, "Save Analytics Report",
            f"AIAS_Analytics_{datetime.date.today()}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Analytics"

        # Header row
        headers = ["Student ID", "Full Name", "Sessions Present", "Total Sessions", "Attendance %", "Status"]
        header_fill = PatternFill("solid", fgColor="1B5E35")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows — read from the analytics table widget
        tbl = self._tbl
        for row in range(tbl.rowCount()):
            for col in range(tbl.columnCount()):
                item = tbl.item(row, col)
                val = item.text() if item else ""
                ws.cell(row=row+2, column=col+1, value=val)

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        try:
            wb.save(path)
            QMessageBox.information(self, "Export Successful", f"Analytics exported to:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Export Failed", f"Could not save file:\n{e}")

    def _go_back(self):
        self.hide()
        self.main_win.show()


if __name__ == "__main__":
    import socket
    try:
        _lock_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        _lock_socket.bind(('localhost', 47200))
    except OSError:
        print("[AIAS] Another instance is already running. Exiting.")
        sys.exit(0)

    # Bug 3 / Rec 1: mark any stuck active sessions as completed on startup
    if DB_CONNECTED:
        try:
            _cleanup = db["Sessions"].update_many(
                {"Status": "active"},
                {"$set": {"Status": "completed", "EndTime": datetime.now()}},
            )
            if _cleanup.modified_count:
                print(f"[AIAS] Cleaned up {_cleanup.modified_count} stuck active session(s).")
        except Exception as _ce:
            print(f"[AIAS] Session cleanup error: {_ce}")

    # Bug 1 / Rec 4: ensure admin account exists in MongoDB
    _ensure_admin_exists()

    app = QApplication(sys.argv)
    apply_theme(app, dark=False)

    import os
    from PyQt5.QtGui import QIcon
    _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aias_icon.ico")
    if os.path.exists(_icon_path):
        app.setWindowIcon(QIcon(_icon_path))

    # Load AI models BEFORE showing any window
    # Use a blocking approach with event loop processing so UI stays responsive
    _models_loaded = [False]

    def _on_models_done():
        _models_loaded[0] = True

    _AI_LOADER_THREAD = AIModelLoader()
    _AI_LOADER_THREAD.finished.connect(_on_models_done)
    _AI_LOADER_THREAD.start()

    # Wait for models to finish — process events so app doesn't freeze
    while not _models_loaded[0]:
        app.processEvents()
        QThread.msleep(50)

    # Models ready — now show app with splash
    app_win = AppWindow()
    app_win.show()

    sys.exit(app.exec_())