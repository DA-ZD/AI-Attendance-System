# ui_admin.py
# Contains: AdminPanelWindow

import os
import bcrypt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QFrame, QDialog, QLineEdit,
    QFileDialog, QMessageBox, QScrollArea, QCheckBox, QComboBox,
    QProgressBar, QStackedWidget, QAbstractItemView, QHeaderView,
    QSizePolicy, QApplication, QRadioButton, QButtonGroup,
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap, QColor

import config
from config import (
    BG, WHITE, PRIMARY, PRIMARY_LIGHT, BORDER,
    TEXT_DARK, TEXT_MED, TEXT_GRAY,
    DB_CONNECTED, db,
    LATE_CUTOFF_MINUTES, EARLY_LEAVE_CUTOFF_MINUTES, RECOGNITION_THRESHOLD,
)
from ai_models import (
    _ensure_ai_models, apply_clahe, upscale_face_crop,
)
from ui_theme import (
    make_stat_card, make_badge, make_sidebar_base, make_table,
)
from ui_session import EmbeddingWorker, BatchEmbedWorker, ImageAttendanceWorker


class AdminPanelWindow(QWidget):
    def __init__(self, embedded=False):
        super().__init__()
        self._embedded = embedded
        self._embedding_worker = None
        self._pending_student  = None
        self._pending_dest     = None
        if not embedded:
            self.setWindowTitle("AIAS — Admin Panel")
            import os
            from PyQt5.QtGui import QIcon
            _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "aias_icon.ico")
            if os.path.exists(_icon_path):
                self.setWindowIcon(QIcon(_icon_path))
            self.setMinimumSize(1100, 700)
            self.resize(1100, 700)
        self.setStyleSheet(f"background:{BG};")
        self._build()

    def _build(self):
        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        sidebar, s_lay = make_sidebar_base("AD", "Admin", "System administrator")

        nav_lbl = QLabel("NAVIGATION")
        nav_lbl.setStyleSheet(
            f"font-size:10px; font-weight:700; color:{TEXT_GRAY}; letter-spacing:1px;"
            f"background:transparent; border:none;"
        )
        s_lay.addWidget(nav_lbl)
        s_lay.addSpacing(4)

        self._nav_btns = []
        for i, label in enumerate(["Students", "Courses", "Instructors", "Settings"]):
            btn = QPushButton(label)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setCheckable(False)
            btn.clicked.connect(lambda _, idx=i: self._switch_page(idx))
            self._nav_btns.append(btn)
            s_lay.addWidget(btn)

        self._apply_nav(0)

        s_lay.addStretch()

        logout_btn = QPushButton("Logout")
        logout_btn.setToolTip("Sign out and return to login screen")
        logout_btn.setStyleSheet(
            "QPushButton { background:#DC2626; color:white; border-radius:6px; "
            "padding:8px 12px; font-size:12px; font-weight:700; }"
            "QPushButton:hover { background:#B91C1C; color:white; }"
            "QPushButton:pressed { background:#991B1B; color:white; }"
        )
        logout_btn.setCursor(Qt.PointingHandCursor)
        logout_btn.clicked.connect(self._logout)
        s_lay.addWidget(logout_btn)

        self._stack = QStackedWidget()
        self._stack.addWidget(self._build_students_page())
        self._stack.addWidget(self._build_courses_page())
        self._stack.addWidget(self._build_instructors_page())
        self._stack.addWidget(self._build_settings_page())

        root.addWidget(sidebar)
        root.addWidget(self._stack)

    def _apply_nav(self, active_idx):
        for i, btn in enumerate(self._nav_btns):
            if i == active_idx:
                btn.setStyleSheet(
                    f"background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:6px;"
                    f"padding:8px 12px; text-align:left; font-weight:600;"
                )
            else:
                btn.setStyleSheet(
                    f"background:transparent; color:{TEXT_MED}; border-radius:6px;"
                    f"padding:8px 12px; text-align:left; font-weight:400;"
                )

    def _switch_page(self, idx):
        self._stack.setCurrentIndex(idx)
        self._apply_nav(idx)

    def _refresh_stack_page(self, idx, build_fn):
        old = self._stack.widget(idx)
        new = build_fn()
        self._stack.insertWidget(idx, new)
        if old is not None:
            self._stack.removeWidget(old)
            old.deleteLater()
        self._stack.setCurrentIndex(idx)

    def _page_header(self, title, subtitle, show_add_student=False, show_add_course=False):
        hdr = QHBoxLayout()
        col = QVBoxLayout()
        col.setSpacing(2)
        t = QLabel(title)
        t.setStyleSheet(f"font-size:20px; font-weight:800; color:{TEXT_DARK};")
        s = QLabel(subtitle)
        s.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY};")
        col.addWidget(t)
        col.addWidget(s)
        hdr.addLayout(col)
        hdr.addStretch()
        if show_add_student:
            action_btn = QPushButton("+ Add Student")
            action_btn.clicked.connect(self._add_student)
        elif show_add_course:
            action_btn = QPushButton("+ Add Course")
            action_btn.clicked.connect(self._add_course)
        else:
            action_btn = QPushButton("+ Import Excel")
            action_btn.clicked.connect(self._import_excel)
        action_btn.setFixedHeight(36)
        action_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px; padding:6px 16px;"
        )
        action_btn.setCursor(Qt.PointingHandCursor)
        hdr.addWidget(action_btn)
        return hdr

    def _add_student(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Student")
        dialog.setFixedWidth(440)
        dialog.setStyleSheet(f"background:{BG};")

        lay = QVBoxLayout(dialog)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        title_lbl = QLabel("Add New Student")
        title_lbl.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title_lbl)

        id_lbl = QLabel("Student ID")
        id_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        id_field = QLineEdit()
        id_field.setPlaceholderText("e.g. 431109377")
        id_field.setFixedHeight(38)
        lay.addWidget(id_lbl)
        lay.addWidget(id_field)

        name_lbl = QLabel("Full Name")
        name_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        name_field = QLineEdit()
        name_field.setPlaceholderText("e.g. Mohammed Al-Harbi")
        name_field.setFixedHeight(38)
        lay.addWidget(name_lbl)
        lay.addWidget(name_field)

        courses_lbl = QLabel("Enroll in Courses")
        courses_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        lay.addWidget(courses_lbl)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(160)
        scroll.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:6px; background:{WHITE};"
        )
        scroll_widget = QWidget()
        scroll_widget.setStyleSheet(f"background:{WHITE};")
        scroll_lay = QVBoxLayout(scroll_widget)
        scroll_lay.setContentsMargins(10, 8, 10, 8)
        scroll_lay.setSpacing(6)

        checkboxes = []
        courses = []
        if DB_CONNECTED:
            try:
                courses = list(db["Courses"].find({}))
            except Exception as _e:
                print(f"[AIAS] error: {_e}")

        for course in courses:
            section_id = str(course.get("SectionID", ""))
            course_title = course.get("CourseTitle", "")
            cb = QCheckBox(f"{section_id} — {course_title}")
            cb.setStyleSheet(f"color:{TEXT_DARK}; font-size:12px; background:transparent;")
            checkboxes.append((cb, course))
            scroll_lay.addWidget(cb)

        if not courses:
            no_lbl = QLabel("No courses found")
            no_lbl.setStyleSheet(f"color:{TEXT_GRAY}; font-size:12px; background:transparent;")
            scroll_lay.addWidget(no_lbl)

        scroll_lay.addStretch()
        scroll.setWidget(scroll_widget)
        lay.addWidget(scroll)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(dialog.reject)
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:8px 20px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(dialog.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        if dialog.exec_() != QDialog.Accepted:
            return

        student_id = id_field.text().strip()
        full_name = name_field.text().strip()

        if not student_id or not full_name:
            QMessageBox.warning(self, "Missing Fields", "Please fill in Student ID and Full Name.")
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        try:
            if db["Students"].find_one({"StudentID": student_id}):
                QMessageBox.warning(self, "Duplicate ID", f"Student ID '{student_id}' already exists.")
                return
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        selected_sections = [
            str(course.get("SectionID", ""))
            for cb, course in checkboxes
            if cb.isChecked()
        ]

        try:
            db["Students"].insert_one({
                "StudentID": student_id,
                "FullName": full_name,
                "EnrolledSections": selected_sections,
            })
        except Exception as e:
            QMessageBox.critical(self, "DB Error", f"Could not insert student:\n{e}")
            return

        for cb, course in checkboxes:
            if cb.isChecked():
                try:
                    db["Courses"].update_one(
                        {"_id": course["_id"]},
                        {"$addToSet": {"EnrolledStudents": student_id}},
                    )
                except Exception as e:
                    print(f"DB error updating course: {e}")

        QMessageBox.information(self, "Success", f"Student '{full_name}' added successfully.")

        self._refresh_stack_page(0, self._build_students_page)

    def _build_students_page(self):
        page = QWidget()
        page.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(page)
        lay.setContentsMargins(32, 28, 32, 24)
        lay.setSpacing(18)

        hdr_row = QHBoxLayout()
        hdr_row.addLayout(self._page_header("Students", "Manage student photos and embeddings", show_add_student=True))
        batch_btn = QPushButton("⚡ Batch Embed Folder")
        batch_btn.setFixedHeight(36)
        batch_btn.setStyleSheet(
            f"background:#6366F1; color:white; border-radius:6px; padding:6px 16px; font-weight:600;"
        )
        batch_btn.setCursor(Qt.PointingHandCursor)
        batch_btn.clicked.connect(self._batch_embed_folder)
        hdr_row.addWidget(batch_btn)
        lay.addLayout(hdr_row)

        students = []
        if DB_CONNECTED:
            try:
                students = list(db["Students"].find({}))
            except Exception as e:
                print(f"[AIAS] Students page load error: {e}")

        total = len(students)
        ready = sum(1 for s in students if s.get("FaceEmbedding"))
        missing = total - ready

        sr = QHBoxLayout()
        sr.setSpacing(16)
        sr.addWidget(make_stat_card("Total", total, PRIMARY))
        sr.addWidget(make_stat_card("Photo uploaded", ready, "#22C55E"))
        sr.addWidget(make_stat_card("Missing photo", missing, "#EF4444"))
        sr.addStretch()
        lay.addLayout(sr)

        # Build course → enrolled student IDs map for filtering
        course_enrolled = {}   # {"CS101 — Title": ["id1", "id2", ...]}
        if DB_CONNECTED:
            try:
                for c in db["Courses"].find({}):
                    key = f"{c.get('SectionID', '')} — {c.get('CourseTitle', '')}"
                    course_enrolled[key] = [str(x) for x in c.get("EnrolledStudents", [])]
            except Exception as e:
                print(f"[AIAS] Students page course filter load error: {e}")

        combo = QComboBox()
        combo.setFixedWidth(240)
        combo.setFixedHeight(34)
        combo.addItem("All Students")
        for key in course_enrolled:
            combo.addItem(key)
        if combo.count() == 1 and not course_enrolled:
            combo.addItem("No courses found")
        lay.addWidget(combo)

        tbl = make_table(
            ["Student ID", "Full Name", "Face", "Status", "Action", "Delete"],
            students,
            col_widths={0: 140, 2: 80, 3: 120, 4: 180, 5: 100},
            stretch_col=1,
        )
        for r, stu in enumerate(students):
            has_photo = bool(stu.get("FaceEmbedding"))
            n_photos  = len(stu.get("ImagePaths", []))
            status = "Ready" if has_photo else "No photo"
            tbl.setItem(r, 0, QTableWidgetItem(str(stu.get("StudentID", ""))))
            tbl.setItem(r, 1, QTableWidgetItem(stu.get("FullName", "")))

            ph = QLabel((f"✓ {n_photos}") if has_photo else "✗")
            ph.setAlignment(Qt.AlignCenter)
            ph.setStyleSheet(
                f"color:{'#22C55E' if has_photo else '#EF4444'}; font-size:14px; font-weight:700;"
                f"background:transparent; border:none;"
            )
            tbl.setCellWidget(r, 2, ph)
            tbl.setCellWidget(r, 3, make_badge(status))

            action_lbl = "View" if has_photo else "Upload"
            action_btn = QPushButton(action_lbl)
            action_btn.setToolTip("Upload or view student photo for face recognition")
            action_btn.setCursor(Qt.PointingHandCursor)
            if has_photo:
                action_btn.setStyleSheet(
                    "background:#EAF3DE; color:#27500A; border-radius:5px;"
                    "padding:4px 10px; font-size:12px; font-weight:600;"
                )
                action_btn.clicked.connect(lambda _, s=stu: self._view_photo(s))
            else:
                action_btn.setStyleSheet(
                    f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:5px;"
                    "padding:4px 10px; font-size:12px;"
                )
                action_btn.clicked.connect(lambda _, s=stu: self._upload_photo(s))
            edit_btn = QPushButton("Edit")
            edit_btn.setToolTip("Edit student information")
            edit_btn.setCursor(Qt.PointingHandCursor)
            edit_btn.setStyleSheet(
                f"background:#E6F1FB; color:#0C447C; border-radius:5px;"
                "padding:4px 10px; font-size:12px; font-weight:600;"
            )
            edit_btn.clicked.connect(lambda _, s=stu: self._edit_student(s))
            wrapper = QWidget()
            wrapper.setStyleSheet("background:transparent;")
            wl = QHBoxLayout(wrapper)
            wl.setContentsMargins(6, 4, 6, 4)
            wl.setSpacing(6)
            wl.addWidget(action_btn, alignment=Qt.AlignCenter)
            wl.addWidget(edit_btn, alignment=Qt.AlignCenter)
            tbl.setCellWidget(r, 4, wrapper)

            del_btn = QPushButton("Delete")
            del_btn.setToolTip("Remove this student permanently")
            del_btn.setCursor(Qt.PointingHandCursor)
            del_btn.setStyleSheet(
                "background:#FCEBEB; color:#791F1F; border-radius:5px;"
                "padding:4px 12px; font-size:12px; font-weight:600;"
            )
            sid = str(stu.get("StudentID", ""))
            fname = stu.get("FullName", "")
            del_btn.clicked.connect(lambda _, s=sid, n=fname: self._delete_student(s, n))
            del_wrapper = QWidget()
            del_wrapper.setStyleSheet("background:transparent;")
            dl = QHBoxLayout(del_wrapper)
            dl.setContentsMargins(6, 4, 6, 4)
            dl.addWidget(del_btn, alignment=Qt.AlignCenter)
            tbl.setCellWidget(r, 5, del_wrapper)
            tbl.setRowHeight(r, 44)

        def _filter_students(index):
            selected = combo.currentText()
            if selected == "All Students" or selected == "No courses found":
                for row in range(tbl.rowCount()):
                    tbl.setRowHidden(row, False)
            else:
                allowed = set(course_enrolled.get(selected, []))
                for row in range(tbl.rowCount()):
                    sid = tbl.item(row, 0).text() if tbl.item(row, 0) else ""
                    tbl.setRowHidden(row, sid not in allowed)

        combo.currentIndexChanged.connect(_filter_students)

        search_row = QHBoxLayout()
        search_field = QLineEdit()
        search_field.setPlaceholderText("🔍  Search by name or student ID...")
        search_field.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:8px; padding:8px 14px; "
            f"font-size:13px; background:{WHITE}; color:{TEXT_DARK};"
        )
        search_field.setFixedHeight(38)
        search_row.addWidget(search_field)
        search_row.addStretch()
        lay.addLayout(search_row)

        def filter_students(text):
            text = text.lower().strip()
            for row in range(tbl.rowCount()):
                match = False
                for col in [0, 1]:
                    item = tbl.item(row, col)
                    if item and text in item.text().lower():
                        match = True
                        break
                tbl.setRowHidden(row, not match if text else False)

        search_field.textChanged.connect(filter_students)

        lay.addWidget(tbl)
        return page

    def _edit_student(self, student):
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Student")
        dialog.setFixedWidth(440)
        dialog.setStyleSheet(f"background:{BG};")

        lay = QVBoxLayout(dialog)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        title_lbl = QLabel("Edit Student")
        title_lbl.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title_lbl)

        name_lbl = QLabel("Full Name")
        name_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        name_field = QLineEdit()
        name_field.setText(student.get("FullName", ""))
        name_field.setFixedHeight(38)
        lay.addWidget(name_lbl)
        lay.addWidget(name_field)

        courses_lbl = QLabel("Enroll in Courses")
        courses_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        lay.addWidget(courses_lbl)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(160)
        scroll.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:6px; background:{WHITE};"
        )
        scroll_widget = QWidget()
        scroll_widget.setStyleSheet(f"background:{WHITE};")
        scroll_lay = QVBoxLayout(scroll_widget)
        scroll_lay.setContentsMargins(10, 8, 10, 8)
        scroll_lay.setSpacing(6)

        enrolled = [str(e) for e in student.get("EnrolledSections", [])]
        checkboxes = []
        courses = []
        if DB_CONNECTED:
            try:
                courses = list(db["Courses"].find({}))
            except Exception as e:
                print(f"[AIAS] Edit student courses load error: {e}")

        for course in courses:
            section_id = str(course.get("SectionID", ""))
            course_title = course.get("CourseTitle", "")
            cb = QCheckBox(f"{section_id} — {course_title}")
            cb.setStyleSheet(f"color:{TEXT_DARK}; font-size:12px; background:transparent;")
            cb.setChecked(section_id in enrolled)
            checkboxes.append((cb, course))
            scroll_lay.addWidget(cb)

        if not courses:
            no_lbl = QLabel("No courses found")
            no_lbl.setStyleSheet(f"color:{TEXT_GRAY}; font-size:12px; background:transparent;")
            scroll_lay.addWidget(no_lbl)

        scroll_lay.addStretch()
        scroll.setWidget(scroll_widget)
        lay.addWidget(scroll)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet(f"color:{BORDER};")
        lay.addWidget(sep)

        photo_lbl = QLabel("Add More Photos (optional)")
        photo_lbl.setStyleSheet(f"font-size:12px; color:{TEXT_MED};")
        lay.addWidget(photo_lbl)

        selected_paths = []

        photo_row = QHBoxLayout()
        select_btn = QPushButton("📷 Select Photos")
        select_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:6px 14px; font-weight:600;"
        )
        select_btn.setCursor(Qt.PointingHandCursor)
        count_lbl = QLabel("0 photos selected")
        count_lbl.setStyleSheet(f"font-size:12px; color:{TEXT_GRAY};")

        def _select_photos():
            paths, _ = QFileDialog.getOpenFileNames(
                dialog, "Select Photos (multiple allowed)", "", "Images (*.jpg *.jpeg *.png)"
            )
            if paths:
                selected_paths.clear()
                selected_paths.extend(paths)
                count_lbl.setText(f"{len(paths)} photo(s) selected")

        select_btn.clicked.connect(_select_photos)
        photo_row.addWidget(select_btn)
        photo_row.addWidget(count_lbl)
        photo_row.addStretch()
        lay.addLayout(photo_row)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(dialog.reject)
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:8px 20px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(dialog.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        if dialog.exec_() != QDialog.Accepted:
            return

        new_name = name_field.text().strip()
        if not new_name:
            QMessageBox.warning(self, "Missing Fields", "Full Name cannot be empty.")
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        student_id = str(student.get("StudentID", ""))
        selected_sections = [str(course.get("SectionID", "")) for cb, course in checkboxes if cb.isChecked()]

        try:
            db["Students"].update_one(
                {"StudentID": student_id},
                {"$set": {"FullName": new_name, "EnrolledSections": selected_sections}},
            )
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        try:
            db["Courses"].update_many({}, {"$pull": {"EnrolledStudents": student_id}})
            for cb, course in checkboxes:
                if cb.isChecked():
                    db["Courses"].update_one(
                        {"_id": course["_id"]},
                        {"$addToSet": {"EnrolledStudents": student_id}},
                    )
        except Exception as e:
            print(f"DB error syncing courses: {e}")

        if selected_paths:
            import cv2
            import numpy as np
            import shutil

            yolo, arcface, ai_ok = _ensure_ai_models(det_size=(640, 640))
            if ai_ok:
                new_embeddings = []
                folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", student_id)
                os.makedirs(folder, exist_ok=True)

                for photo_path in selected_paths:
                    img = cv2.imread(photo_path)
                    if img is None:
                        continue
                    img = apply_clahe(img)
                    results = yolo(img, verbose=False)
                    boxes_det = results[0].boxes
                    if len(boxes_det) == 0:
                        continue
                    x1, y1, x2, y2 = boxes_det[0].xyxy[0].cpu().numpy().astype(int)
                    if x2 <= x1 or y2 <= y1:
                        continue
                    face_crop = img[y1:y2, x1:x2]
                    face_crop = upscale_face_crop(face_crop, target_size=112)
                    faces = arcface.get(face_crop)
                    if not faces:
                        faces = arcface.get(img)
                    if not faces:
                        continue
                    raw = faces[0].embedding
                    new_embeddings.append(raw / np.linalg.norm(raw))

                if not new_embeddings:
                    QMessageBox.warning(
                        self, "No Face Detected",
                        "No face detected in selected photos. Photos not saved."
                    )
                else:
                    existing_files = [
                        f for f in os.listdir(folder)
                        if f.lower().endswith((".jpg", ".jpeg", ".png"))
                    ]
                    start_idx = len(existing_files) + 1
                    for i, photo_path in enumerate(selected_paths):
                        ext = os.path.splitext(photo_path)[1]
                        dest = os.path.join(folder, f"img{start_idx + i}{ext}")
                        shutil.copy2(photo_path, dest)

                    existing_emb = student.get("FaceEmbedding", [])
                    all_embeddings = list(new_embeddings)
                    if existing_emb:
                        arr = np.array(existing_emb, dtype=np.float32)
                        arr_norm = np.linalg.norm(arr)
                        if arr_norm > 0:
                            all_embeddings.insert(0, arr / arr_norm)

                    avg = np.mean(np.array(all_embeddings), axis=0)
                    avg = avg / np.linalg.norm(avg)

                    all_image_paths = [
                        os.path.abspath(os.path.join(folder, f))
                        for f in os.listdir(folder)
                        if f.lower().endswith((".jpg", ".jpeg", ".png"))
                    ]

                    try:
                        db["Students"].update_one(
                            {"StudentID": student_id},
                            {"$set": {
                                "FaceEmbedding": avg.tolist(),
                                "ImagePaths":    all_image_paths,
                            }},
                        )
                    except Exception as e:
                        print(f"[AIAS] Edit student embedding update error: {e}")

        self._refresh_stack_page(0, self._build_students_page)

    def _delete_student(self, student_id, full_name):
        import shutil
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete {full_name}?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        try:
            db["Students"].delete_one({"StudentID": student_id})
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        try:
            db["Courses"].update_many({}, {"$pull": {"EnrolledStudents": student_id}})
        except Exception as e:
            print(f"DB error removing student from courses: {e}")

        folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", student_id)
        if os.path.isdir(folder):
            try:
                shutil.rmtree(folder)
            except Exception as e:
                print(f"Could not delete photo folder: {e}")

        self._refresh_stack_page(0, self._build_students_page)

    def _build_courses_page(self):
        page = QWidget()
        page.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(page)
        lay.setContentsMargins(32, 28, 32, 24)
        lay.setSpacing(18)

        lay.addLayout(self._page_header("Courses", "All sections and enrolled students", show_add_course=True))

        courses = []
        all_students = []
        if DB_CONNECTED:
            try:
                courses = list(db["Courses"].find({}))
                all_students = list(db["Students"].find({}))
            except Exception as e:
                print(f"[AIAS] Courses page load error: {e}")

        total_courses = len(courses)
        total_students = sum(len(c.get("EnrolledStudents", [])) for c in courses)
        enrolled_ids = {sid for c in courses for sid in c.get("EnrolledStudents", [])}
        ready_count = sum(
            1 for s in all_students
            if s.get("StudentID") in enrolled_ids and s.get("FaceEmbedding")
        )

        sr = QHBoxLayout()
        sr.setSpacing(16)
        sr.addWidget(make_stat_card("Total courses", total_courses, PRIMARY))
        sr.addWidget(make_stat_card("Total students", total_students))
        sr.addWidget(make_stat_card("Ready", ready_count, "#22C55E"))
        sr.addStretch()
        lay.addLayout(sr)

        search_field_c = QLineEdit()
        search_field_c.setPlaceholderText("🔍  Search by course title or section ID...")
        search_field_c.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:8px; padding:8px 14px; "
            f"font-size:13px; background:{WHITE}; color:{TEXT_DARK};"
        )
        search_field_c.setFixedHeight(38)
        lay.addWidget(search_field_c)

        course_widgets = []
        cards_row = QHBoxLayout()
        cards_row.setSpacing(16)
        cards_row.setContentsMargins(0, 0, 0, 0)

        for course in courses:
            section_id = str(course.get("SectionID", ""))
            title = course.get("CourseTitle", "")
            sched_raw = course.get("Schedule", "")
            if isinstance(sched_raw, dict):
                schedule = f"{sched_raw.get('Day', '')} {sched_raw.get('StartTime', '')}–{sched_raw.get('EndTime', '')}"
            else:
                schedule = str(sched_raw)
            enrolled = course.get("EnrolledStudents", [])
            n_enrolled = len(enrolled)
            n_ready = sum(
                1 for s in all_students
                if s.get("StudentID") in enrolled and s.get("FaceEmbedding")
            )
            progress = int(n_ready / n_enrolled * 100) if n_enrolled else 0

            ccard = QFrame()
            ccard.setFixedWidth(340)
            ccard.setStyleSheet(
                f"QFrame {{ background:{WHITE}; border:1px solid {BORDER}; border-radius:10px; }}"
            )
            cc = QVBoxLayout(ccard)
            cc.setContentsMargins(20, 16, 20, 16)
            cc.setSpacing(8)

            badge_lbl = QLabel(f"Section {section_id}")
            badge_lbl.setFixedHeight(24)
            badge_lbl.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            badge_lbl.setStyleSheet(
                f"background:{PRIMARY_LIGHT}; color:{PRIMARY}; border-radius:4px;"
                f"padding:2px 10px; font-size:11px; font-weight:600;"
            )
            cc.addWidget(badge_lbl, alignment=Qt.AlignLeft)

            title_lbl = QLabel(title)
            title_lbl.setStyleSheet(
                f"font-size:15px; font-weight:700; color:{TEXT_DARK};"
                f"background:transparent; border:none;"
            )
            cc.addWidget(title_lbl)

            for icon_txt, detail in [
                ("🗓", schedule),
                ("👥", f"{n_enrolled} students"),
            ]:
                row = QHBoxLayout()
                row.setSpacing(6)
                il = QLabel(icon_txt)
                il.setStyleSheet("font-size:14px; background:transparent; border:none;")
                dl = QLabel(detail)
                dl.setStyleSheet(
                    f"font-size:12px; color:{TEXT_MED}; background:transparent; border:none;"
                )
                row.addWidget(il)
                row.addWidget(dl)
                row.addStretch()
                cc.addLayout(row)

            prog_lbl = QLabel(f"{n_ready}/{n_enrolled} photos uploaded")
            prog_lbl.setStyleSheet(
                f"font-size:11px; color:{TEXT_GRAY}; background:transparent; border:none;"
            )
            prog = QProgressBar()
            prog.setValue(progress)
            prog.setTextVisible(False)
            prog.setFixedHeight(8)
            cc.addWidget(prog_lbl)
            cc.addWidget(prog)

            del_row = QHBoxLayout()
            del_row.addStretch()
            img_att_btn = QPushButton("Attendance from Image")
            img_att_btn.setCursor(Qt.PointingHandCursor)
            img_att_btn.setFixedHeight(30)
            img_att_btn.setStyleSheet(
                "background:#E6F7EF; color:#1B5E35; border-radius:5px;"
                "padding:4px 10px; font-size:12px; font-weight:600;"
            )
            img_att_btn.clicked.connect(lambda _, c=course: self._attendance_from_image(c))
            del_row.addWidget(img_att_btn)
            del_row.addSpacing(6)
            edit_btn = QPushButton("Edit")
            edit_btn.setCursor(Qt.PointingHandCursor)
            edit_btn.setFixedHeight(30)
            edit_btn.setStyleSheet(
                "background:#E6F1FB; color:#0C447C; border-radius:5px;"
                "padding:4px 14px; font-size:12px; font-weight:600;"
            )
            edit_btn.clicked.connect(lambda _, c=course: self._edit_course(c))
            del_row.addWidget(edit_btn)
            del_row.addSpacing(6)
            del_btn = QPushButton("Delete")
            del_btn.setCursor(Qt.PointingHandCursor)
            del_btn.setFixedHeight(30)
            del_btn.setStyleSheet(
                "background:#FCEBEB; color:#791F1F; border-radius:5px;"
                "padding:4px 14px; font-size:12px; font-weight:600;"
            )
            del_btn.clicked.connect(lambda _, sid=section_id, t=title: self._delete_course(sid, t))
            del_row.addWidget(del_btn)
            cc.addLayout(del_row)

            cards_row.addWidget(ccard)
            course_widgets.append((section_id, title, ccard))

        def filter_courses(text):
            text = text.lower().strip()
            for sid, ttl, widget in course_widgets:
                if not text:
                    widget.setVisible(True)
                else:
                    widget.setVisible(text in sid.lower() or text in ttl.lower())

        search_field_c.textChanged.connect(filter_courses)

        cards_row.addStretch()
        lay.addLayout(cards_row)
        lay.addStretch()
        return page

    def _edit_course(self, course):
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Course")
        dialog.setFixedWidth(460)
        dialog.setStyleSheet(f"background:{BG};")

        lay = QVBoxLayout(dialog)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        title_lbl = QLabel("Edit Course")
        title_lbl.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title_lbl)

        sched_raw = course.get("Schedule", {})
        if isinstance(sched_raw, dict):
            cur_day = sched_raw.get("Day", "")
            cur_start = sched_raw.get("StartTime", "")
            cur_end = sched_raw.get("EndTime", "")
        else:
            cur_day = cur_start = cur_end = ""

        fields = {}
        for label, key, value, placeholder in [
            ("Course Title", "course_title", course.get("CourseTitle", ""), "e.g. Data Science"),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
            field = QLineEdit()
            field.setText(value)
            field.setPlaceholderText(placeholder)
            field.setFixedHeight(38)
            fields[key] = field
            lay.addWidget(lbl)
            lay.addWidget(field)

        days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
        times = [f"{h:02d}:{m:02d}" for h in range(8, 23) for m in [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55]]
        combos = {}
        for label, key, options, current in [
            ("Day", "day", days, cur_day),
            ("Start Time", "start_time", times, cur_start),
            ("End Time", "end_time", times, cur_end),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
            cb = QComboBox()
            cb.addItems(options)
            cb.setFixedHeight(38)
            idx = cb.findText(current)
            if idx >= 0:
                cb.setCurrentIndex(idx)
            combos[key] = cb
            lay.addWidget(lbl)
            lay.addWidget(cb)

        instr_lbl = QLabel("Instructor")
        instr_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        lay.addWidget(instr_lbl)
        instr_combo = QComboBox()
        instr_combo.setFixedHeight(38)
        instructors = []
        if DB_CONNECTED:
            try:
                instructors = list(db["Instructors"].find({"Username": {"$ne": "admin"}}))
            except Exception as e:
                print(f"[AIAS] Edit course instructors load error: {e}")
        current_instr = course.get("InstructorUsername", "")
        for instr in instructors:
            instr_combo.addItem(instr.get("Username", ""))
        if not instructors:
            instr_combo.addItem("(No instructors)")
        idx = instr_combo.findText(current_instr)
        if idx >= 0:
            instr_combo.setCurrentIndex(idx)
        lay.addWidget(instr_combo)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(dialog.reject)
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:8px 20px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(dialog.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        if dialog.exec_() != QDialog.Accepted:
            return

        new_title = fields["course_title"].text().strip()
        new_day = combos["day"].currentText()
        new_start = combos["start_time"].currentText()
        new_end = combos["end_time"].currentText()
        new_instr = instr_combo.currentText()

        if not new_title or new_instr == "(No instructors)":
            QMessageBox.warning(self, "Missing Fields", "Please fill in all fields.")
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        section_id = str(course.get("SectionID", ""))
        try:
            db["Courses"].update_one(
                {"SectionID": section_id},
                {"$set": {
                    "CourseTitle": new_title,
                    "InstructorUsername": new_instr,
                    "Schedule": {"Day": new_day, "StartTime": new_start, "EndTime": new_end},
                }},
            )
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        self._refresh_stack_page(1, self._build_courses_page)

    def _delete_course(self, section_id, course_title):
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete {course_title}?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        try:
            db["Courses"].delete_one({"SectionID": section_id})
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        try:
            db["Instructors"].update_many({}, {"$pull": {"AssignedSections": section_id}})
        except Exception as e:
            print(f"DB error removing section from instructors: {e}")

        try:
            db["Students"].update_many({}, {"$pull": {"EnrolledSections": section_id}})
        except Exception as e:
            print(f"DB error removing section from students: {e}")

        self._refresh_stack_page(1, self._build_courses_page)

    def _attendance_from_image(self, course):
        """Admin-only: pick an image file, detect faces, mark attendance for the course."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Attendance Image", "",
            "Images (*.png *.jpg *.jpeg *.bmp *.tiff *.webp)"
        )
        if not path:
            return

        section_id = str(course.get("SectionID", ""))
        enrolled_students = []
        if DB_CONNECTED:
            try:
                enrolled_ids      = course.get("EnrolledStudents", [])
                enrolled_students = list(db["Students"].find({"StudentID": {"$in": enrolled_ids}}))
            except Exception as e:
                QMessageBox.critical(self, "DB Error", str(e))
                return

        if not enrolled_students:
            QMessageBox.warning(self, "No Students", "No enrolled students found for this course.")
            return

        n_with_face = sum(1 for s in enrolled_students if s.get("FaceEmbedding"))
        if n_with_face == 0:
            QMessageBox.warning(
                self, "No Face Data",
                "None of the enrolled students have face embeddings uploaded.\n"
                "Please upload student photos first."
            )
            return

        progress_dlg = QDialog(self)
        progress_dlg.setWindowTitle("Processing Image...")
        progress_dlg.setFixedWidth(400)
        progress_dlg.setStyleSheet(f"background:{BG};")
        progress_dlg.setWindowFlags(progress_dlg.windowFlags() & ~Qt.WindowCloseButtonHint)

        pd_lay = QVBoxLayout(progress_dlg)
        pd_lay.setContentsMargins(24, 24, 24, 24)
        pd_lay.setSpacing(12)

        pd_title = QLabel("Analyzing Image...")
        pd_title.setStyleSheet(f"font-size:15px; font-weight:700; color:{TEXT_DARK};")
        pd_lay.addWidget(pd_title)

        pd_sub = QLabel("Running face detection and recognition. This may take a few seconds.")
        pd_sub.setStyleSheet(f"font-size:12px; color:{TEXT_GRAY};")
        pd_sub.setWordWrap(True)
        pd_lay.addWidget(pd_sub)

        pd_bar = QProgressBar()
        pd_bar.setRange(0, 0)
        pd_bar.setFixedHeight(8)
        pd_lay.addWidget(pd_bar)

        progress_dlg.setModal(True)
        progress_dlg.show()
        QApplication.processEvents()

        self._img_worker = ImageAttendanceWorker(path, enrolled_students)

        def _on_finished(recognized):
            progress_dlg.close()
            self._show_image_attendance_results(course, enrolled_students, recognized, path)

        def _on_error(msg):
            progress_dlg.close()
            QMessageBox.critical(self, "Processing Error", msg)

        self._img_worker.finished.connect(_on_finished)
        self._img_worker.error.connect(_on_error)
        self._img_worker.start()

    def _show_image_attendance_results(self, course, enrolled_students, recognized, image_path):
        """Show results dialog and allow admin to submit image-based attendance."""
        from datetime import datetime
        from PyQt5.QtWidgets import QTableWidget, QHeaderView, QAbstractItemView

        section_id   = str(course.get("SectionID", ""))
        course_title = course.get("CourseTitle", "")

        dlg = QDialog(self)
        dlg.setWindowTitle("Image Attendance Results")
        dlg.setMinimumWidth(580)
        dlg.setStyleSheet(f"background:{BG};")

        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(14)

        title_lbl = QLabel(f"Attendance Results — {section_id}  {course_title}")
        title_lbl.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title_lbl)

        img_name = os.path.basename(image_path)
        n_present = len(recognized)
        n_total   = len(enrolled_students)
        sub_lbl = QLabel(
            f"Image: {img_name}   •   {n_present} recognized / {n_total} enrolled"
        )
        sub_lbl.setStyleSheet(f"font-size:12px; color:{TEXT_GRAY};")
        lay.addWidget(sub_lbl)

        tbl = QTableWidget()
        tbl.setColumnCount(3)
        tbl.setHorizontalHeaderLabels(["Student ID", "Full Name", "Status"])
        tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        tbl.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tbl.setSelectionMode(QAbstractItemView.NoSelection)
        tbl.verticalHeader().setVisible(False)
        tbl.setStyleSheet(
            f"QTableWidget {{ border:1px solid {BORDER}; border-radius:8px; background:{WHITE}; }}"
            f"QHeaderView::section {{ background:{WHITE}; color:{TEXT_MED}; font-weight:600; "
            f"border:none; border-bottom:1px solid {BORDER}; padding:6px; }}"
        )

        rows = []
        for stu in enrolled_students:
            sid  = str(stu.get("StudentID", ""))
            name = stu.get("FullName", "")
            rows.append((sid, name, "Present" if sid in recognized else "Absent"))

        tbl.setRowCount(len(rows))
        for r, (sid, name, status) in enumerate(rows):
            tbl.setItem(r, 0, QTableWidgetItem(sid))
            tbl.setItem(r, 1, QTableWidgetItem(name))
            status_item = QTableWidgetItem(status)
            status_item.setForeground(
                QColor("#22C55E") if status == "Present" else QColor("#EF4444")
            )
            tbl.setItem(r, 2, status_item)

        tbl.setFixedHeight(min(40 * len(rows) + 42, 360))
        lay.addWidget(tbl)

        note_lbl = QLabel(
            "Review the results above, then click \"Submit Attendance\" to save to the database."
        )
        note_lbl.setStyleSheet(f"font-size:11px; color:{TEXT_GRAY};")
        note_lbl.setWordWrap(True)
        lay.addWidget(note_lbl)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; "
            "border-radius:6px; padding:8px 20px; font-weight:600;"
        )
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(dlg.reject)

        submit_btn = QPushButton("Submit Attendance")
        submit_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; "
            "padding:8px 20px; font-weight:700;"
        )
        submit_btn.setCursor(Qt.PointingHandCursor)

        def _submit():
            if not DB_CONNECTED:
                QMessageBox.warning(dlg, "No Database", "Database is not connected.")
                return
            try:
                now = datetime.now()
                session_doc = {
                    "CourseID":    course.get("CourseID", section_id),
                    "SectionID":   section_id,
                    "InstructorID": "admin",
                    "StartTime":   now,
                    "EndTime":     now,
                    "Status":      "completed",
                    "Source":      "image_upload",
                    "ImageFile":   os.path.basename(image_path),
                }
                result  = db["Sessions"].insert_one(session_doc)
                sess_id = result.inserted_id

                for sid, name, status in rows:
                    db["AttendanceLogs"].insert_one({
                        "SessionID":   sess_id,
                        "StudentID":   sid,
                        "Status":      status,
                        "FirstSeenAt": now if status == "Present" else None,
                        "LastSeenAt":  now if status == "Present" else None,
                    })

                QMessageBox.information(
                    dlg, "Attendance Saved",
                    f"Session created.\n"
                    f"{n_present} marked Present, {n_total - n_present} marked Absent."
                )
                dlg.accept()
            except Exception as e:
                QMessageBox.critical(dlg, "DB Error", f"Could not save attendance:\n{e}")

        submit_btn.clicked.connect(_submit)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(submit_btn)
        lay.addLayout(btn_row)

        dlg.exec_()

    def _upload_photo(self, student):
        import shutil

        sid = str(student.get("StudentID", ""))

        paths, _ = QFileDialog.getOpenFileNames(
            self, f"Select Photos (multiple allowed) for {student.get('FullName', '')}",
            "", "Images (*.jpg *.jpeg *.png)"
        )
        if not paths:
            return

        try:
            folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data", sid)
            os.makedirs(folder, exist_ok=True)

            existing = [
                f for f in os.listdir(folder)
                if f.lower().endswith((".jpg", ".jpeg", ".png"))
            ]
            start_idx = len(existing) + 1

            for i, path in enumerate(paths):
                ext = os.path.splitext(path)[1]
                dest = os.path.join(folder, f"img{start_idx + i}{ext}")
                shutil.copy2(path, dest)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not save photos:\n{e}")
            return

        all_folder_paths = [
            os.path.join(folder, f)
            for f in os.listdir(folder)
            if f.lower().endswith((".jpg", ".jpeg", ".png"))
        ]

        self._pending_student = student
        self._pending_folder  = folder

        self._embedding_worker = EmbeddingWorker(all_folder_paths)
        self._embedding_worker.finished.connect(self._on_embedding_done)
        self._embedding_worker.error.connect(self._on_embedding_error)
        self._embedding_worker.start()

        QMessageBox.information(
            self, "Processing...",
            f"{len(paths)} photo(s) saved for {student.get('FullName', '')}.\n"
            "Generating augmented variants and extracting embeddings...\n"
            "You'll see a confirmation when done."
        )

    def _on_embedding_done(self, embedding):
        import ai_models as _aim
        student = self._pending_student
        folder  = self._pending_folder
        sid     = str(student.get("StudentID", ""))

        all_image_paths = [
            os.path.abspath(os.path.join(folder, f))
            for f in os.listdir(folder)
            if f.lower().endswith((".jpg", ".jpeg", ".png"))
        ] if os.path.isdir(folder) else []

        if DB_CONNECTED:
            try:
                db["Students"].update_one(
                    {"StudentID": sid},
                    {"$set": {
                        "ImagePaths":    all_image_paths,
                        "FaceEmbedding": embedding,
                    }}
                )
            except Exception as e:
                QMessageBox.critical(self, "DB Error", f"Could not save embedding:\n{e}")
                return

        QMessageBox.information(
            self, "✅ Done",
            f"Embedding saved successfully for {student.get('FullName', '')}!"
        )

        self._refresh_stack_page(0, self._build_students_page)

        self._pending_student = None
        self._pending_folder  = None

    def _on_embedding_error(self, error_msg):
        QMessageBox.warning(self, "Embedding Failed", error_msg)
        self._pending_student = None
        self._pending_dest    = None

    def _view_photo(self, student):
        image_paths = student.get("ImagePaths", [])
        if not image_paths:
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Photo — {student.get('FullName', '')}")
        dialog.setFixedSize(320, 360)
        lay = QVBoxLayout(dialog)

        lbl = QLabel()
        pixmap = QPixmap(image_paths[0])
        if not pixmap.isNull():
            lbl.setPixmap(pixmap.scaled(280, 280, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        else:
            lbl.setText("Image not found")
        lbl.setAlignment(Qt.AlignCenter)

        name_lbl = QLabel(student.get("FullName", ""))
        name_lbl.setAlignment(Qt.AlignCenter)
        name_lbl.setStyleSheet("font-size:13px; font-weight:600;")

        lay.addWidget(lbl)
        lay.addWidget(name_lbl)
        dialog.exec_()

    def _add_course(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Course")
        dialog.setFixedWidth(460)
        dialog.setStyleSheet(f"background:{BG};")

        lay = QVBoxLayout(dialog)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        title_lbl = QLabel("Add New Course")
        title_lbl.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title_lbl)

        fields = {}
        for label, key, placeholder in [
            ("Section ID", "section_id", "e.g. 102"),
            ("Course Title", "course_title", "e.g. Data Science"),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
            field = QLineEdit()
            field.setPlaceholderText(placeholder)
            field.setFixedHeight(38)
            fields[key] = field
            lay.addWidget(lbl)
            lay.addWidget(field)

        days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
        times = [f"{h:02d}:{m:02d}" for h in range(8, 23) for m in [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55]]
        combos = {}
        for label, key, options in [
            ("Day", "day", days),
            ("Start Time", "start_time", times),
            ("End Time", "end_time", times),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
            cb = QComboBox()
            cb.addItems(options)
            cb.setFixedHeight(38)
            combos[key] = cb
            lay.addWidget(lbl)
            lay.addWidget(cb)

        instr_lbl = QLabel("Instructor")
        instr_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        lay.addWidget(instr_lbl)
        instr_combo = QComboBox()
        instr_combo.setFixedHeight(38)
        instructors = []
        instr_id_map = {}
        if DB_CONNECTED:
            try:
                instructors = list(db["Instructors"].find({"Username": {"$ne": "admin"}}))
            except Exception as _e:
                print(f"[AIAS] error: {_e}")
        for instr in instructors:
            username = instr.get("Username", "")
            instr_combo.addItem(username)
            instr_id_map[username] = instr.get("_id")
        if not instructors:
            instr_combo.addItem("(No instructors)")
        lay.addWidget(instr_combo)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(dialog.reject)
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:8px 20px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(dialog.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        if dialog.exec_() != QDialog.Accepted:
            return

        section_id = fields["section_id"].text().strip()
        course_title = fields["course_title"].text().strip()
        day = combos["day"].currentText()
        start_time = combos["start_time"].currentText()
        end_time = combos["end_time"].currentText()
        selected_instructor = instr_combo.currentText()

        if not all([section_id, course_title]) or selected_instructor == "(No instructors)":
            QMessageBox.warning(self, "Missing Fields", "Please fill in all fields.")
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        try:
            if db["Courses"].find_one({"SectionID": section_id}):
                QMessageBox.warning(self, "Duplicate Section ID", f"Section ID '{section_id}' already exists.")
                return
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        try:
            db["Courses"].insert_one({
                "SectionID": section_id,
                "CourseTitle": course_title,
                "InstructorUsername": selected_instructor,
                "InstructorID": instr_id_map.get(selected_instructor),
                "Schedule": {"Day": day, "StartTime": start_time, "EndTime": end_time},
                "EnrolledStudents": [],
            })
        except Exception as e:
            QMessageBox.critical(self, "DB Error", f"Could not insert course:\n{e}")
            return

        try:
            db["Instructors"].update_one(
                {"Username": selected_instructor},
                {"$addToSet": {"AssignedSections": section_id}},
            )
        except Exception as e:
            print(f"DB error updating instructor: {e}")

        QMessageBox.information(self, "Success", f"Course '{course_title}' added successfully.")
        self._refresh_stack_page(1, self._build_courses_page)

    def _build_instructors_page(self):
        page = QWidget()
        page.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(page)
        lay.setContentsMargins(32, 28, 32, 24)
        lay.setSpacing(18)

        hdr = QHBoxLayout()
        col = QVBoxLayout()
        col.setSpacing(2)
        t = QLabel("Instructors")
        t.setStyleSheet(f"font-size:20px; font-weight:800; color:{TEXT_DARK};")
        s = QLabel("Manage instructor accounts and section assignments")
        s.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY};")
        col.addWidget(t)
        col.addWidget(s)
        hdr.addLayout(col)
        hdr.addStretch()
        add_btn = QPushButton("+ Add Instructor")
        add_btn.setFixedHeight(36)
        add_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px; padding:6px 16px;"
        )
        add_btn.setCursor(Qt.PointingHandCursor)
        add_btn.clicked.connect(self._add_instructor)
        hdr.addWidget(add_btn)
        lay.addLayout(hdr)

        instructors = []
        if DB_CONNECTED:
            try:
                instructors = list(db["Instructors"].find({"Username": {"$ne": "admin"}}))
            except Exception as _e:
                print(f"[AIAS] error: {_e}")

        total = len(instructors)
        active_sections = len({
            sec
            for instr in instructors
            for sec in instr.get("AssignedSections", [])
        })

        sr = QHBoxLayout()
        sr.setSpacing(16)
        sr.addWidget(make_stat_card("Total instructors", total, PRIMARY))
        sr.addWidget(make_stat_card("Active sections", active_sections))
        sr.addStretch()
        lay.addLayout(sr)

        tbl = make_table(
            ["Username", "Full Name", "Email", "Sections", "Action"],
            instructors,
            col_widths={0: 140, 2: 200, 3: 100, 4: 180},
            stretch_col=1,
        )
        for r, instr in enumerate(instructors):
            username = instr.get("Username", "")
            fullname = instr.get("FullName", "")
            email = instr.get("UniversityEmail", "")
            sections = ", ".join(str(sec) for sec in instr.get("AssignedSections", []))

            tbl.setItem(r, 0, QTableWidgetItem(username))
            tbl.setItem(r, 1, QTableWidgetItem(fullname))
            tbl.setItem(r, 2, QTableWidgetItem(email))
            tbl.setItem(r, 3, QTableWidgetItem(sections))

            edit_btn = QPushButton("Edit")
            edit_btn.setCursor(Qt.PointingHandCursor)
            edit_btn.setStyleSheet(
                f"background:#E6F1FB; color:#0C447C; border-radius:5px;"
                "padding:4px 10px; font-size:12px; font-weight:600;"
            )
            edit_btn.clicked.connect(lambda _, i=instr: self._edit_instructor(i))
            del_btn = QPushButton("Delete")
            del_btn.setCursor(Qt.PointingHandCursor)
            del_btn.setStyleSheet(
                "background:#FCEBEB; color:#791F1F; border-radius:5px;"
                "padding:4px 10px; font-size:12px; font-weight:600;"
            )
            del_btn.clicked.connect(lambda _, u=username: self._delete_instructor(u))
            wrapper = QWidget()
            wrapper.setStyleSheet("background:transparent;")
            wl = QHBoxLayout(wrapper)
            wl.setContentsMargins(6, 4, 6, 4)
            wl.setSpacing(6)
            wl.addWidget(edit_btn, alignment=Qt.AlignCenter)
            wl.addWidget(del_btn, alignment=Qt.AlignCenter)
            tbl.setCellWidget(r, 4, wrapper)
            tbl.setRowHeight(r, 44)

        search_row = QHBoxLayout()
        search_field_i = QLineEdit()
        search_field_i.setPlaceholderText("🔍  Search by username or full name...")
        search_field_i.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:8px; padding:8px 14px; "
            f"font-size:13px; background:{WHITE}; color:{TEXT_DARK};"
        )
        search_field_i.setFixedHeight(38)
        search_row.addWidget(search_field_i)
        search_row.addStretch()

        def filter_instructors(text):
            text = text.lower().strip()
            for row in range(tbl.rowCount()):
                match = False
                for col in [0, 1]:
                    item = tbl.item(row, col)
                    if item and text in item.text().lower():
                        match = True
                        break
                tbl.setRowHidden(row, not match if text else False)

        search_field_i.textChanged.connect(filter_instructors)

        lay.addLayout(search_row)
        lay.addWidget(tbl)
        return page

    def _add_instructor(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Instructor")
        dialog.setFixedWidth(460)
        dialog.setStyleSheet(f"background:{BG};")

        lay = QVBoxLayout(dialog)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        title_lbl = QLabel("Add New Instructor")
        title_lbl.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title_lbl)

        fields = {}
        for label, key, placeholder, is_password in [
            ("Username", "username", "e.g. dr_ali", False),
            ("Full Name", "fullname", "e.g. Dr. Ali Al-Harbi", False),
            ("Password", "password", "Enter password", True),
            ("University Email", "email", "e.g. ali@qu.edu.sa", False),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
            field = QLineEdit()
            field.setPlaceholderText(placeholder)
            field.setFixedHeight(38)
            if is_password:
                field.setEchoMode(QLineEdit.Password)
            fields[key] = field
            lay.addWidget(lbl)
            lay.addWidget(field)

        sections_lbl = QLabel("Assign Sections")
        sections_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        lay.addWidget(sections_lbl)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(140)
        scroll.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:6px; background:{WHITE};"
        )
        scroll_widget = QWidget()
        scroll_widget.setStyleSheet(f"background:{WHITE};")
        scroll_lay = QVBoxLayout(scroll_widget)
        scroll_lay.setContentsMargins(10, 8, 10, 8)
        scroll_lay.setSpacing(6)

        checkboxes = []
        courses = []
        if DB_CONNECTED:
            try:
                courses = list(db["Courses"].find({}))
            except Exception as _e:
                print(f"[AIAS] error: {_e}")

        for course in courses:
            section_id = str(course.get("SectionID", ""))
            course_title = course.get("CourseTitle", "")
            cb = QCheckBox(f"{section_id} — {course_title}")
            cb.setStyleSheet(f"color:{TEXT_DARK}; font-size:12px; background:transparent;")
            checkboxes.append((cb, section_id))
            scroll_lay.addWidget(cb)

        if not courses:
            no_lbl = QLabel("No courses found")
            no_lbl.setStyleSheet(f"color:{TEXT_GRAY}; font-size:12px; background:transparent;")
            scroll_lay.addWidget(no_lbl)

        scroll_lay.addStretch()
        scroll.setWidget(scroll_widget)
        lay.addWidget(scroll)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(dialog.reject)
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:8px 20px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(dialog.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        if dialog.exec_() != QDialog.Accepted:
            return

        username = fields["username"].text().strip()
        fullname = fields["fullname"].text().strip()
        password = fields["password"].text()
        email = fields["email"].text().strip()

        if not username or not fullname or not password or not email:
            QMessageBox.warning(self, "Missing Fields", "Please fill in all fields.")
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        try:
            if db["Instructors"].find_one({"Username": username}):
                QMessageBox.warning(self, "Duplicate Username", f"Username '{username}' already exists.")
                return
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        hashed = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt())
        selected_sections = [sid for cb, sid in checkboxes if cb.isChecked()]

        try:
            db["Instructors"].insert_one({
                "Username": username,
                "FullName": fullname,
                "Password": hashed,
                "UniversityEmail": email,
                "AssignedSections": selected_sections,
            })
        except Exception as e:
            QMessageBox.critical(self, "DB Error", f"Could not insert instructor:\n{e}")
            return

        for sid in selected_sections:
            try:
                db["Courses"].update_one(
                    {"SectionID": sid},
                    {"$set": {"InstructorUsername": username}},
                )
            except Exception as e:
                print(f"DB error updating course: {e}")

        QMessageBox.information(self, "Success", f"Instructor '{fullname}' added successfully.")
        self._refresh_stack_page(2, self._build_instructors_page)

    def _edit_instructor(self, instructor):
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Instructor")
        dialog.setFixedWidth(460)
        dialog.setStyleSheet(f"background:{BG};")

        lay = QVBoxLayout(dialog)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(12)

        title_lbl = QLabel("Edit Instructor")
        title_lbl.setStyleSheet(f"font-size:16px; font-weight:700; color:{TEXT_DARK};")
        lay.addWidget(title_lbl)

        fields = {}
        for label, key, value, is_password, placeholder in [
            ("Full Name", "fullname", instructor.get("FullName", ""), False, ""),
            ("University Email", "email", instructor.get("UniversityEmail", ""), False, ""),
            ("New Password", "password", "", True, "Leave blank to keep current password"),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
            field = QLineEdit()
            field.setText(value)
            field.setFixedHeight(38)
            if is_password:
                field.setEchoMode(QLineEdit.Password)
            if placeholder:
                field.setPlaceholderText(placeholder)
            fields[key] = field
            lay.addWidget(lbl)
            lay.addWidget(field)

        sections_lbl = QLabel("Assigned Sections")
        sections_lbl.setStyleSheet(f"font-size:12px; font-weight:600; color:{TEXT_MED};")
        lay.addWidget(sections_lbl)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(140)
        scroll.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:6px; background:{WHITE};"
        )
        scroll_widget = QWidget()
        scroll_widget.setStyleSheet(f"background:{WHITE};")
        scroll_lay = QVBoxLayout(scroll_widget)
        scroll_lay.setContentsMargins(10, 8, 10, 8)
        scroll_lay.setSpacing(6)

        assigned = [str(s) for s in instructor.get("AssignedSections", [])]
        checkboxes = []
        courses = []
        if DB_CONNECTED:
            try:
                courses = list(db["Courses"].find({}))
            except Exception as _e:
                print(f"[AIAS] error: {_e}")

        for course in courses:
            section_id = str(course.get("SectionID", ""))
            course_title = course.get("CourseTitle", "")
            cb = QCheckBox(f"{section_id} — {course_title}")
            cb.setStyleSheet(f"color:{TEXT_DARK}; font-size:12px; background:transparent;")
            cb.setChecked(section_id in assigned)
            checkboxes.append((cb, section_id))
            scroll_lay.addWidget(cb)

        if not courses:
            no_lbl = QLabel("No courses found")
            no_lbl.setStyleSheet(f"color:{TEXT_GRAY}; font-size:12px; background:transparent;")
            scroll_lay.addWidget(no_lbl)

        scroll_lay.addStretch()
        scroll.setWidget(scroll_widget)
        lay.addWidget(scroll)

        btn_row = QHBoxLayout()
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(
            f"background:{WHITE}; color:{TEXT_MED}; border:1px solid {BORDER}; border-radius:6px; padding:8px 16px;"
        )
        cancel_btn.setCursor(Qt.PointingHandCursor)
        cancel_btn.clicked.connect(dialog.reject)
        save_btn = QPushButton("Save")
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:{WHITE}; border-radius:6px;"
            f"padding:8px 20px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(dialog.accept)
        btn_row.addWidget(cancel_btn)
        btn_row.addStretch()
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        if dialog.exec_() != QDialog.Accepted:
            return

        new_name = fields["fullname"].text().strip()
        new_email = fields["email"].text().strip()
        new_password = fields["password"].text()

        if not new_name or not new_email:
            QMessageBox.warning(self, "Missing Fields", "Full Name and Email cannot be empty.")
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        username = instructor.get("Username", "")
        new_sections = [sid for cb, sid in checkboxes if cb.isChecked()]
        update_fields = {
            "FullName": new_name,
            "UniversityEmail": new_email,
            "AssignedSections": new_sections,
        }
        if new_password:
            update_fields["Password"] = bcrypt.hashpw(new_password.encode("utf-8"), bcrypt.gensalt())

        try:
            db["Instructors"].update_one({"Username": username}, {"$set": update_fields})
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        try:
            db["Courses"].update_many(
                {"InstructorUsername": username, "SectionID": {"$nin": new_sections}},
                {"$unset": {"InstructorUsername": ""}},
            )
            for sid in new_sections:
                db["Courses"].update_one(
                    {"SectionID": sid},
                    {"$set": {"InstructorUsername": username}},
                )
        except Exception as e:
            print(f"DB error syncing courses: {e}")

        self._refresh_stack_page(2, self._build_instructors_page)

    def _delete_instructor(self, username):
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete instructor '{username}'?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        try:
            db["Instructors"].delete_one({"Username": username})
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return

        try:
            db["Courses"].update_many(
                {"InstructorUsername": username},
                {"$unset": {"InstructorUsername": ""}},
            )
        except Exception as e:
            print(f"DB error cleaning instructor from courses: {e}")

        self._refresh_stack_page(2, self._build_instructors_page)

    def _import_excel(self):
        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        path, _ = QFileDialog.getOpenFileName(
            self, "Import Students from Excel", "", "Excel Files (*.xlsx *.xls)"
        )
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active

            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

            # Accept flexible column names
            def _col(names):
                for n in names:
                    for i, h in enumerate(headers):
                        if n in h:
                            return i
                return None

            id_col   = _col(["student id", "studentid", "id", "student_id"])
            name_col = _col(["full name", "fullname", "name", "full_name"])
            sec_col  = _col(["section", "course", "enrolled"])

            if id_col is None or name_col is None:
                QMessageBox.critical(
                    self, "Import Failed",
                    "Could not find required columns.\n"
                    "Excel must have: 'Student ID' and 'Full Name' columns.\n"
                    f"Found columns: {', '.join(headers)}"
                )
                return

            inserted = skipped = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                sid  = str(row[id_col]).strip()   if row[id_col]   is not None else ""
                name = str(row[name_col]).strip()  if row[name_col] is not None else ""
                sec  = str(row[sec_col]).strip()   if sec_col is not None and row[sec_col] is not None else ""

                if not sid or not name or sid == "None":
                    continue

                if db["Students"].find_one({"StudentID": sid}):
                    skipped += 1
                    continue

                doc = {"StudentID": sid, "FullName": name, "EnrolledSections": []}
                if sec:
                    doc["EnrolledSections"] = [sec]
                    db["Courses"].update_one({"SectionID": sec}, {"$addToSet": {"EnrolledStudents": sid}})

                db["Students"].insert_one(doc)
                inserted += 1

            wb.close()
            QMessageBox.information(
                self, "Import Complete",
                f"Import finished.\n✔ Inserted: {inserted}\n⚠ Skipped (duplicate): {skipped}"
            )

            self._refresh_stack_page(0, self._build_students_page)

        except Exception as e:
            QMessageBox.critical(self, "Import Failed", f"Error reading Excel file:\n{e}")

    def _batch_embed_folder(self):
        if not DB_CONNECTED:
            QMessageBox.warning(self, "No Database", "Database is not connected.")
            return

        root_folder = QFileDialog.getExistingDirectory(
            self, "Select Root Folder (subfolders named by Student ID)"
        )
        if not root_folder:
            return

        data_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data")
        os.makedirs(data_folder, exist_ok=True)

        # Progress dialog
        progress_dialog = QDialog(self)
        progress_dialog.setWindowTitle("Batch Embedding")
        progress_dialog.setFixedSize(500, 380)
        progress_dialog.setStyleSheet(f"background:{BG};")
        pd_lay = QVBoxLayout(progress_dialog)
        pd_lay.setContentsMargins(24, 20, 24, 20)
        pd_lay.setSpacing(10)

        pd_title = QLabel("Processing student folders…")
        pd_title.setStyleSheet(f"font-size:15px; font-weight:700; color:{TEXT_DARK};")
        pd_lay.addWidget(pd_title)

        log_widget = QTableWidget(0, 1)
        log_widget.horizontalHeader().setVisible(False)
        log_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        log_widget.verticalHeader().setVisible(False)
        log_widget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        log_widget.setShowGrid(False)
        log_widget.setStyleSheet(f"background:{WHITE}; border:1px solid {BORDER}; border-radius:6px;")
        pd_lay.addWidget(log_widget)

        close_btn = QPushButton("Close")
        close_btn.setEnabled(False)
        close_btn.setFixedHeight(36)
        close_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; padding:6px 20px; font-weight:600;"
        )
        close_btn.clicked.connect(progress_dialog.accept)
        pd_lay.addWidget(close_btn, alignment=Qt.AlignRight)

        self._batch_worker = BatchEmbedWorker(root_folder, data_folder)

        def _on_progress(msg):
            r = log_widget.rowCount()
            log_widget.insertRow(r)
            item = QTableWidgetItem(msg)
            item.setFlags(Qt.ItemIsEnabled)
            log_widget.setItem(r, 0, item)
            log_widget.scrollToBottom()

        def _on_finished(processed, skipped):
            pd_title.setText(f"Done — ✔ {processed} embedded, ⚠ {skipped} skipped")
            close_btn.setEnabled(True)
            self._refresh_stack_page(0, self._build_students_page)

        self._batch_worker.progress.connect(_on_progress)
        self._batch_worker.finished.connect(_on_finished)
        self._batch_worker.start()
        progress_dialog.exec_()

    def _build_settings_page(self):
        page = QWidget()
        page.setStyleSheet(f"background:{BG};")
        lay = QVBoxLayout(page)
        lay.setContentsMargins(40, 32, 40, 32)
        lay.setSpacing(20)

        late_cutoff  = LATE_CUTOFF_MINUTES
        early_cutoff = EARLY_LEAVE_CUTOFF_MINUTES
        threshold    = RECOGNITION_THRESHOLD
        if DB_CONNECTED:
            try:
                doc = db["Settings"].find_one({"_id": "global"})
                if doc:
                    late_cutoff  = doc.get("LateCutoffMinutes",       10)
                    early_cutoff = doc.get("EarlyLeaveCutoffMinutes", 15)
                    threshold    = doc.get("RecognitionThreshold",    0.55)
            except Exception as e:
                print(f"[AIAS] Settings page load error: {e}")

        title = QLabel("System Settings")
        title.setStyleSheet(f"font-size:20px; font-weight:800; color:{TEXT_DARK};")
        sub = QLabel("Changes take effect immediately and persist across restarts.")
        sub.setStyleSheet(f"font-size:13px; color:{TEXT_GRAY};")
        lay.addWidget(title)
        lay.addWidget(sub)

        fields = {}
        for label, key, val, desc in [
            ("Late Arrival Cutoff (minutes)", "late", str(late_cutoff),
             "Students arriving after this many minutes are marked Late"),
            ("Early Leave Cutoff (minutes)", "early", str(early_cutoff),
             "Students not seen for this many minutes are marked Early Leave"),
            ("Face Recognition Threshold (0.0–1.0)", "threshold", str(threshold),
             "Minimum cosine similarity to identify a student (higher = stricter)"),
        ]:
            card = QFrame()
            card.setStyleSheet(
                f"QFrame{{background:{WHITE};border:1px solid {BORDER};border-radius:8px;}}"
            )
            cl = QVBoxLayout(card)
            cl.setContentsMargins(20, 16, 20, 16)
            cl.setSpacing(6)
            lbl = QLabel(label)
            lbl.setStyleSheet(f"font-size:13px; font-weight:700; color:{TEXT_DARK};")
            desc_lbl = QLabel(desc)
            desc_lbl.setStyleSheet(f"font-size:11px; color:{TEXT_GRAY};")
            field = QLineEdit(val)
            field.setFixedHeight(38)
            field.setFixedWidth(160)
            fields[key] = field
            cl.addWidget(lbl)
            cl.addWidget(desc_lbl)
            cl.addWidget(field)
            lay.addWidget(card)

        save_btn = QPushButton("Save Settings")
        save_btn.setFixedHeight(40)
        save_btn.setFixedWidth(160)
        save_btn.setStyleSheet(
            f"background:{PRIMARY}; color:white; border-radius:6px; font-weight:700;"
        )
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(lambda: self._save_settings(fields))
        lay.addWidget(save_btn)
        lay.addStretch()
        return page

    def _save_settings(self, fields):
        try:
            late   = int(fields["late"].text().strip())
            early  = int(fields["early"].text().strip())
            thresh = float(fields["threshold"].text().strip())
        except ValueError:
            QMessageBox.warning(self, "Invalid Input", "Please enter valid numeric values.")
            return
        if not (0.0 < thresh <= 1.0):
            QMessageBox.warning(self, "Invalid Threshold", "Threshold must be between 0.0 (exclusive) and 1.0 (inclusive).")
            return
        if late < 1 or early < 1:
            QMessageBox.warning(self, "Invalid Value", "Cutoff values must be at least 1 minute.")
            return

        import config as _cfg
        _cfg.LATE_CUTOFF_MINUTES        = late
        _cfg.EARLY_LEAVE_CUTOFF_MINUTES = early
        _cfg.RECOGNITION_THRESHOLD      = thresh

        if DB_CONNECTED:
            try:
                db["Settings"].update_one(
                    {"_id": "global"},
                    {"$set": {
                        "LateCutoffMinutes":       late,
                        "EarlyLeaveCutoffMinutes": early,
                        "RecognitionThreshold":    thresh,
                    }},
                    upsert=True,
                )
                QMessageBox.information(self, "Saved", "Settings saved and applied immediately.")
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))
        else:
            QMessageBox.information(self, "Saved", "Settings applied for this session (no DB).")

    def _logout(self):
        from ui_login import AppWindow
        for widget in QApplication.instance().topLevelWidgets():
            if isinstance(widget, AppWindow):
                widget._login_page.username_field.clear()
                widget._login_page.password_field.clear()
                widget._stack.setCurrentWidget(widget._login_page)
                widget.setStyleSheet("background:#0a0f0a;")
                break
        self.deleteLater()

    def keyPressEvent(self, event):
        from PyQt5.QtCore import Qt
        if event.key() == Qt.Key_Escape:
            pass
        else:
            super().keyPressEvent(event)
